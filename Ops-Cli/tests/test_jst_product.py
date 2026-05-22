import json

import pytest

from ops_cli.platforms.jst import product


def test_run_product_sync_requires_template(tmp_path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)

    with pytest.raises(RuntimeError, match="未找到商品同步模板"):
        product.run_product_sync()


def test_filter_workbook_by_brand(tmp_path) -> None:
    from openpyxl import Workbook, load_workbook

    source = tmp_path / "source.xlsx"
    output = tmp_path / "filtered.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "商品资料"
    ws.append(["商品编码", "品牌", "名称"])
    ws.append(["A1", "奥克斯", "商品A"])
    ws.append(["B1", "其他", "商品B"])
    ws.append(["C1", "苏泊尔", "商品C"])
    wb.save(source)

    summary = product.filter_workbook_by_brand(source, output, {"奥克斯", "苏泊尔"})

    assert summary[0]["kept_rows"] == 2
    assert summary[0]["deleted_rows"] == 1
    filtered = load_workbook(output)
    rows = list(filtered.active.iter_rows(values_only=True))
    assert rows[1][0] == "A1"
    assert rows[2][0] == "C1"


def test_run_product_sync_with_template(monkeypatch, tmp_path) -> None:
    from openpyxl import Workbook

    monkeypatch.chdir(tmp_path)
    (tmp_path / "data" / "jst").mkdir(parents=True, exist_ok=True)
    (tmp_path / "runtime" / "context").mkdir(parents=True, exist_ok=True)

    source = tmp_path / "聚水潭商品资料（最新）.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "商品资料"
    ws.append(["商品编码", "品牌", "名称"])
    ws.append(["A1", "奥克斯", "商品A"])
    ws.append(["B1", "其他", "商品B"])
    wb.save(source)

    template_path = tmp_path / "data" / "jst" / "product_sync_template.json"
    template_path.write_text(
        json.dumps(
            {
                "method": "POST",
                "url": "https://example.com",
                "headers": {"Cookie": "a=b"},
                "post_data_json": {"data": 160160444},
                "defaults": {
                    "source_path": str(source),
                    "keep_brands": ["奥克斯", "苏泊尔"],
                    "target_name": "聚水潭商品资料（最新）.xlsx",
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    class FakeResponse:
        status_code = 200

        def raise_for_status(self):
            return None

        def json(self):
            return {
                "code": 0,
                "data": {"url": "https://download.example.com/file.xlsx", "fileName": "聚水潭商品资料（最新）.xlsx"},
            }

    class FakeFileResponse:
        status_code = 200
        content = source.read_bytes()

    class FakeClient:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def request(self, method, url, headers=None, json=None):
            return FakeResponse()

        def get(self, url):
            return FakeFileResponse()

    monkeypatch.setattr(product, "build_client", lambda **kwargs: FakeClient())
    monkeypatch.setattr(product, "_scene_store_path", lambda site, scene: tmp_path / "scene.json")
    (tmp_path / "scene.json").write_text(
        json.dumps({"headers": {"Cookie": "a=b"}, "method": "POST", "url": "https://example.com"}),
        encoding="utf-8",
    )
    monkeypatch.setattr(product, "_scene_is_valid", lambda scene_data: {"valid": True, "reason": "ok"})

    result = product.run_product_sync(use_local_only=True)

    assert result.data["used_backend_export"] is False
    assert result.data["downloaded"] is False
    assert result.data["sheet_summary"][0]["kept_rows"] == 1
    assert result.data["scene"] == "product_export"


def test_run_product_sync_falls_back_to_local_source(monkeypatch, tmp_path) -> None:
    from openpyxl import Workbook

    monkeypatch.chdir(tmp_path)
    (tmp_path / "data" / "jst").mkdir(parents=True, exist_ok=True)
    (tmp_path / "runtime" / "context").mkdir(parents=True, exist_ok=True)

    source = tmp_path / "聚水潭商品资料（最新）.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "商品资料"
    ws.append(["商品编码", "品牌", "名称"])
    ws.append(["A1", "奥克斯", "商品A"])
    ws.append(["B1", "其他", "商品B"])
    wb.save(source)

    template_path = tmp_path / "data" / "jst" / "product_sync_template.json"
    template_path.write_text(
        json.dumps(
            {
                "method": "POST",
                "url": "https://example.com",
                "headers": {"Cookie": "a=b"},
                "post_data_json": {"data": 160160444},
                "defaults": {
                    "source_path": str(source),
                    "keep_brands": ["奥克斯", "苏泊尔"],
                    "target_name": "聚水潭商品资料（最新）.xlsx",
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    class FakeResponse:
        status_code = 200

        def raise_for_status(self):
            return None

        def json(self):
            return {
                "code": 0,
                "data": {"url": "https://download.example.com/file.xlsx", "fileName": "聚水潭商品资料（最新）.xlsx"},
            }

    class FakeFileResponse:
        status_code = 200
        content = '{"data":null,"code":10002,"message":"文件已过期"}'.encode("utf-8")

        def raise_for_status(self):
            return None

    class FakeClient:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def request(self, method, url, headers=None, json=None):
            return FakeResponse()

        def get(self, url):
            return FakeFileResponse()

    monkeypatch.setattr(product, "build_client", lambda **kwargs: FakeClient())
    monkeypatch.setattr(product, "_scene_store_path", lambda site, scene: tmp_path / "scene.json")
    (tmp_path / "scene.json").write_text(
        json.dumps({"headers": {"Cookie": "a=b"}, "method": "POST", "url": "https://example.com"}),
        encoding="utf-8",
    )
    monkeypatch.setattr(product, "_scene_is_valid", lambda scene_data: {"valid": True, "reason": "ok"})

    result = product.run_product_sync()

    assert result.data["downloaded"] is False
    assert result.data["fallback"] == "expired_or_invalid_export_url_used_local_source"


def test_run_product_sync_retries_after_auth_refresh(monkeypatch, tmp_path) -> None:
    from openpyxl import Workbook

    monkeypatch.chdir(tmp_path)
    (tmp_path / "data" / "jst").mkdir(parents=True, exist_ok=True)
    (tmp_path / "runtime" / "context").mkdir(parents=True, exist_ok=True)

    source = tmp_path / "聚水潭商品资料（最新）.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "商品资料"
    ws.append(["商品编码", "品牌", "名称"])
    ws.append(["A1", "奥克斯", "商品A"])
    wb.save(source)

    (tmp_path / "data" / "jst" / "product_sync_template.json").write_text(
        json.dumps(
            {
                "method": "POST",
                "url": "https://example.com",
                "headers": {"Cookie": "a=b"},
                "post_data_json": {"data": 160160444},
                "defaults": {"source_path": str(source), "keep_brands": ["奥克斯"], "target_name": "聚水潭商品资料（最新）.xlsx"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(product, "_scene_store_path", lambda site, scene: tmp_path / "scene.json")
    (tmp_path / "scene.json").write_text(json.dumps({"headers": {"Cookie": "a=b"}, "method": "POST", "url": "https://example.com"}), encoding="utf-8")
    monkeypatch.setattr(product, "_scene_is_valid", lambda scene_data: {"valid": True, "reason": "ok"})
    refresh_calls = {"count": 0}
    monkeypatch.setattr(product, "learn_jst_product_sync", lambda force=False: refresh_calls.__setitem__("count", refresh_calls["count"] + 1))

    attempts = {"count": 0}

    def fake_download_source(template, source_path):
        attempts["count"] += 1
        if attempts["count"] == 1:
            raise RuntimeError("401 Unauthorized")
        return {"download_url": "https://example.com/file.xlsx"}

    monkeypatch.setattr(product, "_download_source", fake_download_source)

    result = product.run_product_sync()

    assert result.data["auth_refresh_applied"] is True
    assert refresh_calls["count"] == 1
