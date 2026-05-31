from __future__ import annotations

import json
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

from ops_cli.capabilities import mark_scene_refreshed
from ops_cli.capabilities import require_interactive_recovery
from ops_cli.config import get_config
from ops_cli.integrations.sessionhub import get_scene_manager
from ops_cli.output import CommandResponse
from ops_cli.platforms.auth_shared import is_probable_auth_error
from ops_cli.platforms.jst.shared import ensure_scene_file_ready
from ops_cli.runtime_context import write_runtime_context
from ops_cli.utils.http import build_client


JST_SITE = "jst_erp"
PRODUCT_SCENE = "product_export"
TARGET_NAME = "聚水潭商品资料（最新）.xlsx"
TEMPLATE_PATH = Path("data/jst/product_sync_template.json")


def _sessionhub_root() -> Path:
    return Path(get_config().sessionhub_root).expanduser().resolve()


def _scene_store_path(site: str, scene: str) -> Path:
    return _sessionhub_root() / "data" / "sessions" / site / f"{scene}.json"


def _template_path() -> Path:
    return Path.cwd() / TEMPLATE_PATH


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


NOISY_HEADERS = {
    "accept-encoding",
    "content-length",
    "host",
    "cookie",
}


def _filter_cookies_for_url(cookies: list[dict[str, Any]] | None, url: str) -> list[dict[str, Any]]:
    if not cookies:
        return []
    host = (urlparse(url).hostname or "").lower()
    if not host:
        return []
    matched: list[dict[str, Any]] = []
    for cookie in cookies:
        domain = str(cookie.get("domain") or "").strip().lower()
        if not domain:
            continue
        normalized = domain.lstrip(".")
        if host == normalized or host.endswith(f".{normalized}"):
            matched.append(cookie)
    return matched


def _merge_cookie_header(
    headers: dict[str, Any],
    cookies: list[dict[str, Any]] | None,
    *,
    target_url: str = "",
) -> dict[str, str]:
    merged = {str(key): str(value) for key, value in headers.items() if str(key).lower() not in NOISY_HEADERS}
    cookies = _filter_cookies_for_url(cookies, target_url)
    if cookies:
        merged["cookie"] = "; ".join(
            f"{cookie.get('name')}={cookie.get('value')}"
            for cookie in cookies
            if cookie.get("name")
        )
    return merged


def _template_defaults() -> dict[str, Any]:
    config = get_config()
    return {
        "source_path": config.jst_product_source_path,
        "keep_brands": list(config.jst_product_keep_brands),
        "target_name": TARGET_NAME,
    }


def _scene_is_valid(scene_data: dict[str, Any]) -> dict[str, Any]:
    headers = _merge_cookie_header(
        dict(scene_data.get("headers") or {}),
        scene_data.get("cookies") or [],
        target_url=str(scene_data.get("url") or ""),
    )
    method = str(scene_data.get("method") or "POST").upper()
    url = str(scene_data.get("url") or "")
    payload = scene_data.get("post_data_json") or {}
    with build_client(follow_redirects=True, timeout=30.0) as client:
        response = client.request(method, url, headers=headers, json=payload)
    response.raise_for_status()
    data = response.json()
    export = data.get("data") or {}
    export_url = str(export.get("url") or "").strip()
    valid = response.status_code == 200 and data.get("code") == 0 and bool(export_url)
    reason = "接口返回 200，scene 可用" if valid else "导出接口未返回下载地址"
    return {
        "status_code": response.status_code,
        "valid": valid,
        "reason": reason,
        "export_url": export_url,
        "file_name": export.get("fileName"),
    }


def _write_template(*, scene_data: dict[str, Any]) -> Path:
    template = {
        "site": JST_SITE,
        "scene": PRODUCT_SCENE,
        "captured_at": datetime.now().isoformat(timespec="seconds"),
        "url": scene_data.get("url"),
        "method": scene_data.get("method"),
        "headers": _merge_cookie_header(
            dict(scene_data.get("headers") or {}),
            scene_data.get("cookies") or [],
            target_url=str(scene_data.get("url") or ""),
        ),
        "post_data_json": scene_data.get("post_data_json") or {},
        "defaults": _template_defaults(),
    }
    path = _template_path()
    _write_json(path, template)
    return path


def _load_template() -> dict[str, Any]:
    path = _template_path()
    if not path.exists():
        raise RuntimeError(f"未找到商品同步模板：{path}。请先运行 `ops jst product learn`。")
    return _read_json(path)


def filter_workbook_by_brand(source: Path, output: Path, keep_brands: set[str]) -> list[dict[str, Any]]:
    source_wb = load_workbook(source, read_only=False, data_only=False)
    output_wb = Workbook(write_only=True)
    summary: list[dict[str, Any]] = []

    for source_ws in source_wb.worksheets:
        rows = source_ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            output_wb.create_sheet(title=source_ws.title)
            summary.append(
                {
                    "sheet": source_ws.title,
                    "before_rows": 0,
                    "kept_rows": 0,
                    "deleted_rows": 0,
                    "brand_counts": {},
                }
            )
            continue

        if "品牌" not in headers:
            raise RuntimeError(f"工作表 {source_ws.title} 找不到【品牌】字段")

        brand_index = headers.index("品牌")
        output_ws = output_wb.create_sheet(title=source_ws.title)
        output_ws.append(headers)

        before_rows = 0
        kept_rows = 0
        brand_counts: dict[str, int] = {}
        for row in rows:
            before_rows += 1
            value = row[brand_index] if brand_index < len(row) else None
            brand = str(value).strip() if value is not None else ""
            if brand not in keep_brands:
                continue
            output_ws.append(row)
            kept_rows += 1
            brand_counts[brand] = brand_counts.get(brand, 0) + 1

        summary.append(
            {
                "sheet": source_ws.title,
                "before_rows": before_rows,
                "kept_rows": kept_rows,
                "deleted_rows": before_rows - kept_rows,
                "brand_counts": dict(sorted(brand_counts.items())),
            }
        )

    output.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output)
    source_wb.close()
    return summary


def _resolve_keep_brands(keep_brands: list[str] | None) -> list[str]:
    if keep_brands:
        return [brand.strip() for brand in keep_brands if brand.strip()]
    return list(get_config().jst_product_keep_brands)


def _is_valid_product_export(path: Path) -> bool:
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [
            str(worksheet.cell(1, column).value).strip() if worksheet.cell(1, column).value is not None else ""
            for column in range(1, worksheet.max_column + 1)
        ]
        valid = worksheet.max_row > 1 and "品牌" in headers and "商品编码" in headers
        workbook.close()
        return valid
    except Exception:
        return False


def _find_recent_browser_download(download_dir: Path, *, max_age_seconds: int = 3600) -> Path | None:
    if not download_dir.is_dir():
        return None
    cutoff = time.time() - max_age_seconds
    candidates = [
        path
        for path in download_dir.glob("商品资料_*.xlsx")
        if path.is_file() and path.stat().st_mtime >= cutoff and _is_valid_product_export(path)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _download_source(template: dict[str, Any], destination: Path) -> dict[str, Any]:
    headers = dict(template.get("headers") or {})
    method = str(template.get("method") or "POST").upper()
    url = str(template.get("url") or "").strip()
    payload = template.get("post_data_json") or {}
    with build_client(follow_redirects=True, timeout=120.0) as client:
        response = client.request(method, url, headers=headers, json=payload)
        response.raise_for_status()
        data = response.json()
        if data.get("code") != 0:
            raise RuntimeError(f"聚水潭商品资料导出接口返回异常：{data}")
        export = data.get("data") or {}
        export_url = str(export.get("url") or "").strip()
        if not export_url:
            raise RuntimeError("导出接口未返回下载地址")
        file_response = client.get(export_url)
        file_response.raise_for_status()
        if not file_response.content:
            raise RuntimeError("导出文件下载为空，已停止同步。")
        if not file_response.content.startswith(b"PK"):
            try:
                payload = json.loads(file_response.content.decode("utf-8"))
            except Exception:
                payload = None
            if isinstance(payload, dict) and str(payload.get("message") or "").strip():
                raise RuntimeError(f"导出文件下载失败：{payload.get('message')}")
            raise RuntimeError("导出文件不是合法的 xlsx 压缩包")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(file_response.content)
    return {
        "export_url": export_url,
        "file_name": export.get("fileName"),
        "download_size": len(file_response.content),
    }


def learn_jst_product_sync(*, force: bool = False) -> CommandResponse:
    scene_path = _scene_store_path(JST_SITE, PRODUCT_SCENE)
    inputs = {"site": JST_SITE, "scene": PRODUCT_SCENE, "force": force}
    manager = get_scene_manager()

    if scene_path.exists() and not force:
        scene_data = _read_json(scene_path)
        try:
            check = _scene_is_valid(scene_data)
            if check["valid"]:
                template_path = _write_template(scene_data=scene_data)
                context_path = write_runtime_context(
                    task_name="jst_product_learn",
                    status="success",
                    inputs=inputs,
                    outputs={"scene_path": str(scene_path), "template_path": str(template_path), "reuse": True},
                    artifacts=[str(scene_path), str(template_path)],
                )
                return CommandResponse(
                    success=True,
                    platform="jst",
                    command="product learn",
                    data={
                        "site": JST_SITE,
                        "scene": PRODUCT_SCENE,
                        "source": "existing_scene",
                        "scene_path": str(scene_path),
                        "template_path": str(template_path),
                        "context_path": str(context_path),
                        "next_command": "ops --json jst product sync",
                    },
                )
        except Exception:
            pass

    session = manager.capture_scene(JST_SITE, PRODUCT_SCENE) if force else manager.ensure_scene(JST_SITE, PRODUCT_SCENE)
    scene_data = _read_json(scene_path) if scene_path.exists() else session
    template_path = _write_template(scene_data=scene_data)
    check = _scene_is_valid(scene_data)
    context_path = write_runtime_context(
        task_name="jst_product_learn",
        status="success" if check["valid"] else "failed",
        inputs=inputs,
        outputs={"scene_path": str(scene_path), "template_path": str(template_path), "check": check},
        artifacts=[str(scene_path), str(template_path)],
    )
    if not check["valid"]:
        raise RuntimeError(f"scene 已捕获，但复检失败：{check['reason']}")
    return CommandResponse(
        success=True,
        platform="jst",
        command="product learn",
        data={
            "site": JST_SITE,
            "scene": PRODUCT_SCENE,
            "source": "sessionhub",
            "scene_path": str(scene_path),
            "template_path": str(template_path),
            "context_path": str(context_path),
            "next_command": "ops --json jst product sync",
        },
    )


def run_product_sync(
    *,
    dry_run: bool = False,
    use_local_only: bool = False,
    keep_brands: list[str] | None = None,
) -> CommandResponse:
    keep_brand_values = _resolve_keep_brands(keep_brands)
    keep_brand_set = set(keep_brand_values)

    retried_for_auth = False
    auth_refresh_applied = False
    while True:
        template = _load_template()
        scene_path = _scene_store_path(JST_SITE, PRODUCT_SCENE)
        ensure_scene_file_ready(
            scene_path=scene_path,
            read_scene=_read_json,
            validate_scene=_scene_is_valid,
            refresh_scene=learn_jst_product_sync,
            next_command="ops jst product learn",
            missing_label="商品导出 scene",
            invalid_label="商品导出 scene",
        )

        defaults = template.get("defaults") or {}
        source_path = Path(str(defaults.get("source_path") or get_config().jst_product_source_path)).expanduser()
        download_dir = Path(str(defaults.get("download_dir") or Path.home() / "Downloads")).expanduser()
        download_meta: dict[str, Any] = {"used_backend_export": not use_local_only, "downloaded": False}
        if use_local_only:
            if not source_path.is_file():
                raise RuntimeError(f"源文件不存在：{source_path}")
        elif not dry_run:
            try:
                download_meta = {
                    "used_backend_export": True,
                    "downloaded": True,
                    **_download_source(template, source_path),
                }
            except RuntimeError as exc:
                if not retried_for_auth and is_probable_auth_error(exc):
                    require_interactive_recovery(PRODUCT_SCENE)
                    learn_jst_product_sync(force=True)
                    mark_scene_refreshed(PRODUCT_SCENE)
                    retried_for_auth = True
                    auth_refresh_applied = True
                    continue
                if source_path.is_file() and "导出文件下载失败" in str(exc):
                    recent_download = _find_recent_browser_download(download_dir)
                    if recent_download is not None:
                        source_path.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(recent_download, source_path)
                        download_meta = {
                            "used_backend_export": True,
                            "downloaded": True,
                            "fallback": "recent_browser_download_used",
                            "warning": str(exc),
                            "source": str(recent_download),
                            "file_name": recent_download.name,
                            "download_size": recent_download.stat().st_size,
                        }
                    else:
                        download_meta = {
                            "used_backend_export": True,
                            "downloaded": False,
                            "fallback": "expired_or_invalid_export_url_used_local_source",
                            "warning": str(exc),
                        }
                else:
                    raise
        break

    if auth_refresh_applied:
        download_meta["auth_refresh_applied"] = True

    if dry_run:
        context_path = write_runtime_context(
            task_name="jst_product_sync_run",
            status="success",
            inputs={
                "dry_run": True,
                "use_local_only": use_local_only,
                "keep_brands": keep_brand_values,
                "source_path": str(source_path),
            },
            outputs={
                "used_backend_export": not use_local_only,
                "downloaded": False,
                "output_path": str(source_path),
            },
        )
        return CommandResponse(
            success=True,
            platform="jst",
            command="product sync",
            data={
                "source": str(source_path),
                "used_backend_export": not use_local_only,
                "downloaded": False,
                "keep_brands": keep_brand_values,
                "sheet_summary": [],
                "output_path": str(source_path),
                "scene": PRODUCT_SCENE,
                "context_path": str(context_path),
                "dry_run": True,
            },
        )

    with tempfile.TemporaryDirectory(prefix="ops_cli_jst_product_") as temp_dir:
        filtered_path = Path(temp_dir) / TARGET_NAME
        sheet_summary = filter_workbook_by_brand(source_path, filtered_path, keep_brand_set)
        source_path.write_bytes(filtered_path.read_bytes())

    context_path = write_runtime_context(
        task_name="jst_product_sync_run",
        status="success",
        inputs={
            "dry_run": False,
            "use_local_only": use_local_only,
            "keep_brands": keep_brand_values,
            "source_path": str(source_path),
        },
        outputs={
            "used_backend_export": download_meta["used_backend_export"],
            "downloaded": download_meta["downloaded"],
            "sheet_summary": sheet_summary,
            "output_path": str(source_path),
        },
    )
    return CommandResponse(
        success=True,
        platform="jst",
        command="product sync",
        data={
            "source": str(source_path),
            "used_backend_export": download_meta["used_backend_export"],
            "downloaded": download_meta["downloaded"],
            "keep_brands": keep_brand_values,
            "sheet_summary": sheet_summary,
            "output_path": str(source_path),
            "scene": PRODUCT_SCENE,
            "context_path": str(context_path),
            **{key: value for key, value in download_meta.items() if key not in {"used_backend_export", "downloaded"}},
        },
    )


def list_products() -> CommandResponse:
    return CommandResponse(
        success=True,
        platform="jst",
        command="product list",
        data={"items": [], "total": 0, "mode": "mock"},
    )
