from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ops_cli.capabilities import mark_scene_refreshed
from ops_cli.capabilities import require_interactive_recovery
from ops_cli.config import get_config
from ops_cli.output import CommandResponse
from ops_cli.platforms.auth_shared import is_probable_auth_error
from ops_cli.runtime_context import write_runtime_context

from ops_cli.platforms.tmcs.shared import TMCS_PRODUCT_EXPORT_FILENAME
from ops_cli.platforms.tmcs.shared import TMCS_PRODUCT_EXPORT_SCENE
from ops_cli.platforms.tmcs.shared import TMCS_PRODUCT_LATEST_FILENAME
from ops_cli.platforms.tmcs.shared import TMCS_PRODUCT_SEARCH_SCENE
from ops_cli.platforms.tmcs.shared import TMCS_SITE
from ops_cli.platforms.tmcs.shared import check_scene_or_fail
from ops_cli.platforms.tmcs.shared import ensure_scene_assets
from ops_cli.platforms.tmcs.shared import extract_export_task_id
from ops_cli.platforms.tmcs.shared import find_download_url
from ops_cli.platforms.tmcs.shared import form_encode
from ops_cli.platforms.tmcs.shared import gei_task_download_url
from ops_cli.platforms.tmcs.shared import is_probably_excel
from ops_cli.platforms.tmcs.shared import load_scene_or_fail
from ops_cli.platforms.tmcs.shared import read_json
from ops_cli.platforms.tmcs.shared import resolve_download_content
from ops_cli.platforms.tmcs.shared import sanitize_replay_headers
from ops_cli.platforms.tmcs.shared import tmcs_download
from ops_cli.platforms.tmcs.shared import tmcs_request
from ops_cli.platforms.tmcs.shared import write_json
from ops_cli.platforms.tmcs.shared import merge_cookie_header


TEMPLATE_PATH = Path("data/tmcs/product_sync_template.json")


def _template_path() -> Path:
    return Path.cwd() / TEMPLATE_PATH


def _write_template(*, export_scene: dict[str, Any], search_scene: dict[str, Any]) -> Path:
    config = get_config()
    template = {
        "site": TMCS_SITE,
        "scenes": {
            "search": TMCS_PRODUCT_SEARCH_SCENE,
            "export": TMCS_PRODUCT_EXPORT_SCENE,
        },
        "captured_at": datetime.now().isoformat(timespec="seconds"),
        "defaults": {
            "import_path": config.tmcs_product_import_path,
            "latest_path": config.tmcs_product_latest_path,
            "jst_path": config.jst_product_source_path,
            "import_file_name": TMCS_PRODUCT_EXPORT_FILENAME,
            "latest_file_name": TMCS_PRODUCT_LATEST_FILENAME,
        },
        "search": {
            "url": search_scene.get("url"),
            "method": search_scene.get("method"),
            "headers": _sanitize_tmcs_headers(search_scene.get("headers") or {}, search_scene.get("cookies") or []),
            "post_data_form": search_scene.get("post_data_form") or {},
        },
        "export": {
            "url": export_scene.get("url"),
            "method": export_scene.get("method"),
            "headers": _sanitize_tmcs_headers(export_scene.get("headers") or {}, export_scene.get("cookies") or []),
            "post_data_form": export_scene.get("post_data_form") or {},
        },
    }
    path = _template_path()
    write_json(path, template)
    return path


def _sanitize_tmcs_headers(headers: dict[str, Any], cookies: list[dict[str, Any]] | None = None) -> dict[str, str]:
    cleaned = sanitize_replay_headers(headers, [])
    cookie_header = merge_cookie_header({}, cookies).get("cookie")
    if cookie_header:
        cleaned["cookie"] = cookie_header
    return cleaned


def _load_template() -> dict[str, Any]:
    path = _template_path()
    if not path.exists():
        raise RuntimeError(f"未找到猫超商品同步模板：{path}。请先运行 `ops tmcs product learn`。")
    return read_json(path)


def _norm(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return text.upper()


def _load_sheet_data(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header = [
        str(worksheet.cell(1, column).value).strip() if worksheet.cell(1, column).value is not None else ""
        for column in range(1, worksheet.max_column + 1)
    ]
    rows: list[list[object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        row = [worksheet.cell(row_index, column).value for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    workbook.close()
    return header, rows


def _require_columns(header: list[str], required: list[str], file_name: str) -> None:
    missing = [column for column in required if column not in header]
    if missing:
        raise RuntimeError(f"{file_name} 缺少必要字段: {', '.join(missing)}")


def _column_index(header: list[str], candidates: list[str], file_name: str) -> int:
    for candidate in candidates:
        if candidate in header:
            return header.index(candidate)
    raise RuntimeError(f"{file_name} 缺少必要字段: {'/'.join(candidates)}")


def _build_row_by_latest_header(latest_header: list[str], import_header: list[str], import_row: list[object]) -> list[object]:
    import_index = {name: idx for idx, name in enumerate(import_header)}
    row: list[object] = []
    for column_name in latest_header:
        idx = import_index.get(column_name)
        row.append(import_row[idx] if idx is not None and idx < len(import_row) else None)
    return row


def _build_jst_code_pool(jst_rows: list[list[object]], jst_header: list[str]) -> tuple[dict[str, str], list[str]]:
    product_code_idx = jst_header.index("商品编码")
    exact_map: dict[str, str] = {}
    normalized_codes: list[str] = []
    seen: set[str] = set()
    for row in jst_rows:
        code = _norm(row[product_code_idx] if product_code_idx < len(row) else None)
        if not code:
            continue
        exact_map.setdefault(code, code)
        if code not in seen:
            normalized_codes.append(code)
            seen.add(code)
    return exact_map, normalized_codes


def _find_replacement_code(barcode: object, exact_map: dict[str, str], normalized_codes: list[str]) -> tuple[str | None, str]:
    normalized_barcode = _norm(barcode)
    if not normalized_barcode:
        return None, "empty"
    exact_hit = exact_map.get(normalized_barcode)
    if exact_hit is not None:
        return exact_hit, "exact"
    fuzzy_hits = [code for code in normalized_codes if normalized_barcode in code or code in normalized_barcode]
    if len(fuzzy_hits) == 1:
        return fuzzy_hits[0], "fuzzy"
    if len(fuzzy_hits) > 1:
        return None, "multiple"
    return None, "miss"


def _save_latest_rows(path: Path, rows: list[list[object]]) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]
    max_existing_row = worksheet.max_row
    if max_existing_row > 1:
        worksheet.delete_rows(2, max_existing_row - 1)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _flatten_row(row: Any, *, prefix: str = "") -> dict[str, Any]:
    if not isinstance(row, dict):
        return {prefix or "value": row}
    flattened: dict[str, Any] = {}
    for key, value in row.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            flattened.update(_flatten_row(value, prefix=name))
        elif isinstance(value, list):
            flattened[name] = json.dumps(value, ensure_ascii=False)
        else:
            flattened[name] = value
    return flattened


def _extract_rows(payload: Any) -> tuple[list[dict[str, Any]], int | None]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)], None
    if not isinstance(payload, dict):
        return [], None
    data = payload.get("data", payload)
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)], None
    if not isinstance(data, dict):
        return [], None
    total = None
    for key in ("total", "totalCount", "count"):
        if isinstance(data.get(key), int):
            total = int(data[key])
            break
    for key in ("data", "list", "rows", "records", "items", "dataSource"):
        rows = data.get(key)
        if isinstance(rows, list):
            return [row for row in rows if isinstance(row, dict)], total
    return [], total


def _download_goods_from_search(*, search_scene: dict[str, Any], destination: Path) -> dict[str, Any]:
    from openpyxl import Workbook

    form_data = dict(search_scene.get("post_data_form") or {})
    headers = dict(search_scene.get("headers") or {})
    method = str(search_scene.get("method") or "POST").upper()
    url = str(search_scene.get("url") or "").strip()
    form_data["pageSize"] = "100"
    all_rows: list[dict[str, Any]] = []
    total: int | None = None
    for page_index in range(1, 1000):
        form_data["pageIndex"] = str(page_index)
        _, payload, _ = tmcs_request(method, url, headers=headers, data_body=form_encode(form_data), timeout=120.0)
        rows, detected_total = _extract_rows(payload)
        total = detected_total if detected_total is not None else total
        if not rows:
            break
        all_rows.extend(rows)
        if total is not None and len(all_rows) >= total:
            break
        if len(rows) < int(form_data["pageSize"]):
            break
    if not all_rows:
        raise RuntimeError("猫超商品搜索接口未返回可写入 Excel 的数据")
    flattened = [_flatten_row(row) for row in all_rows]
    headers_out = sorted({key for row in flattened for key in row})
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品列表"
    sheet.append(headers_out)
    for row in flattened:
        sheet.append([row.get(key) for key in headers_out])
    workbook.save(destination)
    return {
        "status_code": 200,
        "export_task_id": None,
        "export_url": None,
        "file_name": destination.name,
        "download_size": destination.stat().st_size,
        "source": "product_search_api_fallback",
        "row_count": len(all_rows),
    }


def _download_goods_export(*, export_scene: dict[str, Any], search_scene: dict[str, Any], destination: Path) -> dict[str, Any]:
    form_data = dict(export_scene.get("post_data_form") or {})
    if "_scm_token_" not in form_data or "query" not in form_data:
        raise RuntimeError("商品导出 scene 缺少 _scm_token_ 或 query")
    headers = dict(export_scene.get("headers") or {})
    method = str(export_scene.get("method") or "POST").upper()
    url = str(export_scene.get("url") or "").strip()
    status_code, payload, _ = tmcs_request(method, url, headers=headers, data_body=form_encode(form_data))
    if isinstance(payload, dict) and payload.get("success") is False and payload.get("errorCode") == "PL_GEI_U00001":
        return _download_goods_from_search(search_scene=search_scene, destination=destination)
    task_id = extract_export_task_id(payload)
    download_url = find_download_url(payload)
    file_name = None
    if isinstance(payload, dict):
        data_node = payload.get("data")
        if isinstance(data_node, dict):
            file_name = data_node.get("fileName")
    if task_id:
        gei_url = gei_task_download_url(url, task_id)
        try:
            _, nested_payload, nested_content = tmcs_download(gei_url, headers=headers)
            content, _ = resolve_download_content(content=nested_content, parsed_payload=nested_payload, headers=headers)
        except RuntimeError:
            return _download_goods_from_search(search_scene=search_scene, destination=destination)
    elif download_url:
        _, nested_payload, nested_content = tmcs_download(download_url, headers=headers)
        content, _ = resolve_download_content(content=nested_content, parsed_payload=nested_payload, headers=headers)
    else:
        raise RuntimeError(f"商品导出接口未返回 taskId 或下载地址：{json.dumps(payload, ensure_ascii=False)[:500]}")
    if not content or not is_probably_excel(content):
        raise RuntimeError("商品导出返回的不是合法 Excel 文件")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(content)
    return {
        "status_code": status_code,
        "export_task_id": task_id,
        "export_url": download_url,
        "file_name": file_name,
        "download_size": len(content),
    }


def _sync_product_workbook(*, latest_path: Path, import_path: Path, jst_path: Path, dry_run: bool) -> dict[str, Any]:
    latest_header, latest_rows = _load_sheet_data(latest_path)
    import_header, import_rows = _load_sheet_data(import_path)
    jst_header, jst_rows = _load_sheet_data(jst_path)

    _require_columns(jst_header, ["商品编码"], jst_path.name)

    goods_candidates = ["货品编码", "erpCode", "itemId", "storageGoodsId"]
    barcode_candidates = ["条码", "barcode", "barCode"]
    latest_goods_idx = _column_index(latest_header, goods_candidates, latest_path.name)
    latest_barcode_idx = _column_index(latest_header, barcode_candidates, latest_path.name)
    import_goods_idx = _column_index(import_header, goods_candidates, import_path.name)

    latest_goods_codes = {
        _norm(row[latest_goods_idx] if latest_goods_idx < len(row) else None)
        for row in latest_rows
        if _norm(row[latest_goods_idx] if latest_goods_idx < len(row) else None)
    }

    appended_rows: list[list[object]] = []
    for row in import_rows:
        goods_code = _norm(row[import_goods_idx] if import_goods_idx < len(row) else None)
        if not goods_code or goods_code in latest_goods_codes:
            continue
        latest_goods_codes.add(goods_code)
        appended_rows.append(_build_row_by_latest_header(latest_header, import_header, row))

    exact_map, normalized_codes = _build_jst_code_pool(jst_rows, jst_header)
    stats = {
        "exact_replaced": 0,
        "fuzzy_replaced": 0,
        "unchanged": 0,
        "empty_barcode": 0,
        "missed": 0,
        "multiple_candidates": 0,
    }

    for row in appended_rows:
        current_barcode = row[latest_barcode_idx] if latest_barcode_idx < len(row) else None
        replacement, match_type = _find_replacement_code(current_barcode, exact_map, normalized_codes)
        if match_type == "empty":
            stats["empty_barcode"] += 1
            continue
        if match_type == "multiple":
            stats["multiple_candidates"] += 1
            continue
        if match_type == "miss" or replacement is None:
            stats["missed"] += 1
            continue
        if _norm(current_barcode) == replacement:
            stats["unchanged"] += 1
            continue
        row[latest_barcode_idx] = replacement
        if match_type == "exact":
            stats["exact_replaced"] += 1
        else:
            stats["fuzzy_replaced"] += 1

    if not dry_run:
        _save_latest_rows(latest_path, latest_rows + appended_rows)

    return {
        "original_latest_rows": len(latest_rows),
        "import_rows": len(import_rows),
        "new_rows": len(appended_rows),
        "final_latest_rows": len(latest_rows) + (0 if dry_run else len(appended_rows)),
        **stats,
    }


def learn_product_sync(*, force: bool = False) -> CommandResponse:
    inputs = {"site": TMCS_SITE, "scenes": [TMCS_PRODUCT_SEARCH_SCENE, TMCS_PRODUCT_EXPORT_SCENE], "force": force}
    try:
        search_scene, search_check, search_path = ensure_scene_assets(
            site=TMCS_SITE,
            scene=TMCS_PRODUCT_SEARCH_SCENE,
            force=force,
            next_command="ops tmcs auth capture",
        )
    except RuntimeError:
        search_path = Path(get_config().sessionhub_root).expanduser() / "data" / "sessions" / TMCS_SITE / f"{TMCS_PRODUCT_SEARCH_SCENE}.json"
        search_scene = read_json(search_path)
        search_check = {"status": search_scene.get("status", "unknown"), "fallback": "loaded_existing_scene"}
    try:
        export_scene, export_check, export_path = ensure_scene_assets(
            site=TMCS_SITE,
            scene=TMCS_PRODUCT_EXPORT_SCENE,
            force=force,
            next_command="ops tmcs auth capture",
        )
    except RuntimeError:
        export_path = Path(get_config().sessionhub_root).expanduser() / "data" / "sessions" / TMCS_SITE / f"{TMCS_PRODUCT_EXPORT_SCENE}.json"
        export_scene = read_json(export_path)
        export_check = {"status": export_scene.get("status", "unknown"), "fallback": "loaded_existing_scene"}
    template_path = _write_template(export_scene=export_scene, search_scene=search_scene)
    context_path = write_runtime_context(
        task_name="tmcs_product_learn",
        status="success",
        inputs=inputs,
        outputs={
            "template_path": str(template_path),
            "search_scene_path": str(search_path),
            "export_scene_path": str(export_path),
            "search_check": search_check,
            "export_check": export_check,
        },
        artifacts=[str(search_path), str(export_path), str(template_path)],
    )
    return CommandResponse(
        success=True,
        platform="tmcs",
        command="product learn",
        data={
            "site": TMCS_SITE,
            "search_scene": TMCS_PRODUCT_SEARCH_SCENE,
            "export_scene": TMCS_PRODUCT_EXPORT_SCENE,
            "template_path": str(template_path),
            "context_path": str(context_path),
            "next_command": "ops --json tmcs product sync",
        },
    )


def run_product_sync(
    *,
    dry_run: bool = False,
    use_local_only: bool = False,
    force_refresh: bool = False,
) -> CommandResponse:
    retried_for_auth = False
    auth_refresh_applied = False
    while True:
        template = _load_template()
        defaults = template.get("defaults") or {}
        import_path = Path(str(defaults.get("import_path") or get_config().tmcs_product_import_path)).expanduser()
        latest_path = Path(str(defaults.get("latest_path") or get_config().tmcs_product_latest_path)).expanduser()
        jst_path = Path(str(defaults.get("jst_path") or get_config().jst_product_source_path)).expanduser()

        search_scene_data = template.get("search") or {}
        export_scene_data = template.get("export") or load_scene_or_fail(TMCS_SITE, TMCS_PRODUCT_EXPORT_SCENE, next_command="ops tmcs product learn")
        if not use_local_only:
            check_scene_or_fail(TMCS_SITE, TMCS_PRODUCT_SEARCH_SCENE, next_command="ops tmcs product learn")

        should_auto_download = (not use_local_only) and (force_refresh or not import_path.exists())
        if should_auto_download and not dry_run:
            try:
                download_meta = {
                    "used_backend_export": True,
                    "downloaded": True,
                    **_download_goods_export(export_scene=export_scene_data, search_scene=search_scene_data, destination=import_path),
                }
            except RuntimeError as exc:
                if not retried_for_auth and is_probable_auth_error(exc):
                    require_interactive_recovery(TMCS_PRODUCT_EXPORT_SCENE)
                    learn_product_sync(force=True)
                    mark_scene_refreshed(TMCS_PRODUCT_EXPORT_SCENE)
                    retried_for_auth = True
                    auth_refresh_applied = True
                    continue
                raise
        else:
            download_meta = {
                "used_backend_export": should_auto_download,
                "downloaded": False,
                "auto_download_reason": (
                    "use_local_only"
                    if use_local_only
                    else "existing_import_file"
                    if import_path.exists() and not force_refresh
                    else "dry_run"
                    if dry_run
                    else "not_needed"
                ),
            }
        break

    if auth_refresh_applied:
        download_meta["auth_refresh_applied"] = True

    if not import_path.exists():
        raise RuntimeError(f"未找到猫超导入文件：{import_path}")
    if not jst_path.exists():
        raise RuntimeError(f"未找到聚水潭商品资料：{jst_path}")

    effective_latest_path = latest_path
    if not latest_path.exists():
        if dry_run:
            effective_latest_path = import_path
        else:
            latest_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(import_path, latest_path)

    sync_summary = _sync_product_workbook(
        latest_path=effective_latest_path,
        import_path=import_path,
        jst_path=jst_path,
        dry_run=dry_run,
    )
    context_path = write_runtime_context(
        task_name="tmcs_product_sync_run",
        status="success",
        inputs={
            "dry_run": dry_run,
            "use_local_only": use_local_only,
            "force_refresh": force_refresh,
            "import_path": str(import_path),
            "latest_path": str(latest_path),
            "jst_path": str(jst_path),
        },
        outputs={**download_meta, **sync_summary, "output_path": str(latest_path)},
        artifacts=[str(import_path), str(latest_path)],
    )
    return CommandResponse(
        success=True,
        platform="tmcs",
        command="product sync",
        data={
            "source": str(import_path),
            "output_path": str(latest_path),
            "jst_file": str(jst_path),
            "scene": TMCS_PRODUCT_EXPORT_SCENE,
            "search_scene": TMCS_PRODUCT_SEARCH_SCENE,
            "context_path": str(context_path),
            "dry_run": dry_run,
            **download_meta,
            **sync_summary,
        },
    )


def list_products() -> CommandResponse:
    return CommandResponse(
        success=True,
        platform="tmcs",
        command="product list",
        data={"next_command": "ops --json tmcs product sync", "mode": "cli"},
    )
