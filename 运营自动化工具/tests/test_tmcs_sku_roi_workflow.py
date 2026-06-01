from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.runtime import WorkflowRunner
from core.runtime.registry import discover_workflow
from tasks.tmcs_sku_roi import main as task_entry
from workflows.tmcs_sku_roi import steps
from workflows.tmcs_sku_roi.excel_lookup import load_roi_config
from workflows.tmcs_sku_roi.workflow import build_workflow


def _make_tmcs_file(path: Path, *, sku_code: str = "SKU001", barcode: str = "BAR001", product_code: str = "SPU001") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品编码", "商品名称", "商品上下架状态", "SKU编码", "SKU上下架状态", "生产厂家", "条码"])
    sheet.append([product_code, "测试商品", "上架", sku_code, "上架", "厂商", barcode])
    workbook.save(path)
    workbook.close()
    return path


def _make_jst_file(path: Path, *, product_code: str = "BAR001", price: str = "799", cost: float = 361) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["市场|吊牌价", "基本售价", "图片", "款式编码", "商品编码", "商品名称", "商品简称", "颜色及规格", "颜色", "规格", "实际库存数", "订单占有数", "淘系控价", "成本价"])
    sheet.append([None, None, None, "STYLE", product_code, "测试商品", None, None, None, None, 0, 0, price, cost])
    workbook.save(path)
    workbook.close()
    return path


def _make_roi_config(path: Path, **overrides: float) -> Path:
    payload = {
        "supply_price_factor": 0.9,
        "vip_discount_rate": 0.0,
        "general_fee_rate": 0.007,
        "other_fee_rate": 0.02,
        "storage_fee_rate": 0.0,
        "tax_rate": 0.03,
        "management_fee_rate": 0.048,
        "refund_rate": 0.1,
        "refund_flat_fee": 5.0,
        "domestic_shipping_fee": 5.0,
        "gift_cost": 0.0,
        "safe_profit_rate": 0.1,
        "ideal_promotion_ratio": 0.12,
    }
    payload.update(overrides)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def test_workflow_registers() -> None:
    workflow = discover_workflow("tmcs_sku_roi")
    assert workflow.id == "tmcs_sku_roi"
    assert [step.id for step in workflow.steps] == [
        "check_inputs",
        "lookup_tmcs_barcode",
        "lookup_jst_product",
        "calculate_roi",
        "collect_outputs",
    ]


def test_main_routes_to_workflow(monkeypatch) -> None:
    calls: list[list[str]] = []
    monkeypatch.setattr(sys, "argv", ["tmcs_sku_roi", "--sku-code", "SKU001", "--dry-run"])
    monkeypatch.setattr(task_entry, "_run_workflow", lambda args: calls.append(list(args)) or 0, raising=False)

    assert task_entry.main() == 0
    assert calls == [["tmcs_sku_roi", "--sku-code", "SKU001", "--dry-run"]]


def test_dry_run_is_safe_and_outputs_preview(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    output_file = tmp_path / "result.json"
    run = runner.run(
        build_workflow(),
        inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "SKU001", "--output", str(output_file)]},
        dry_run=True,
    )

    assert run.status == "dry_run_success"
    assert not output_file.exists()
    collect_step = json.loads((runner.last_run_dir / "steps" / "collect_outputs.json").read_text(encoding="utf-8"))
    assert collect_step["outputs"]["理想ROI"] == "8.3333"
    run_json = json.loads((runner.last_run_dir / "run.json").read_text(encoding="utf-8"))
    assert set(run_json["outputs"]) == {"保本ROI", "安全ROI", "理想ROI"}


def test_product_code_query_uses_first_barcode_when_multiple_rows(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = tmp_path / "tmcs.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品编码", "商品名称", "商品上下架状态", "SKU编码", "SKU上下架状态", "生产厂家", "条码"])
    sheet.append(["SPU001", "测试商品A", "上架", "SKU001", "上架", "厂商", "BAR001"])
    sheet.append(["SPU001", "测试商品B", "上架", "SKU002", "上架", "厂商", "BAR002"])
    workbook.save(tmcs_file)
    workbook.close()
    jst_file = _make_jst_file(tmp_path / "jst.xlsx", product_code="BAR001")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(
        build_workflow(),
        inputs={"dry_run": True, "args": ["--dry-run", "--product-code", "SPU001"]},
        dry_run=True,
    )

    assert run.status == "dry_run_success"
    run_json = json.loads((runner.last_run_dir / "run.json").read_text(encoding="utf-8"))
    assert run_json["outputs"]["理想ROI"] == "8.3333"


def test_sku_not_found_returns_clear_error(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "MISS"]}, dry_run=True)

    assert run.status == "failed"
    assert any("未找到 SKU 编码" in error for error in run.errors)


def test_product_code_not_found_returns_clear_error(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": True, "args": ["--dry-run", "--product-code", "SPU404"]}, dry_run=True)

    assert run.status == "failed"
    assert any("未找到 商品编码" in error for error in run.errors)


def test_requires_exactly_one_lookup_key(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(
        build_workflow(),
        inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "SKU001", "--product-code", "SPU001"]},
        dry_run=True,
    )

    assert run.status == "failed"
    assert any("必须且只能提供一个查询参数" in error for error in run.errors)


def test_barcode_missing_returns_clear_error(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx", barcode="")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "SKU001"]}, dry_run=True)

    assert run.status == "failed"
    assert any("条码为空" in error for error in run.errors)


def test_jst_product_not_found_returns_clear_error(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx", barcode="BAR404")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx", product_code="BAR001")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "SKU001"]}, dry_run=True)

    assert run.status == "failed"
    assert any("未找到商品编码" in error for error in run.errors)


def test_real_run_writes_json_artifact(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    output_file = tmp_path / "roi.json"
    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(
        build_workflow(),
        inputs={"dry_run": False, "args": ["--sku-code", "SKU001", "--output", str(output_file)]},
        dry_run=False,
    )

    assert run.status == "success"
    assert output_file.exists()
    payload = json.loads(output_file.read_text(encoding="utf-8"))
    assert set(payload) == {"保本ROI", "安全ROI", "理想ROI"}
    run_json = json.loads((runner.last_run_dir / "run.json").read_text(encoding="utf-8"))
    assert set(run_json["outputs"]) == {"保本ROI", "安全ROI", "理想ROI"}
    artifacts = json.loads((runner.last_run_dir / "artifacts.json").read_text(encoding="utf-8"))
    assert any(artifact["role"] == "output" and artifact["type"] == "json" for artifact in artifacts)


def test_real_run_does_not_modify_source_excels(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json")
    before = {path: path.stat().st_mtime_ns for path in (tmcs_file, jst_file)}
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": False, "args": ["--sku-code", "SKU001"]}, dry_run=False)

    assert run.status == "success"
    after = {path: path.stat().st_mtime_ns for path in (tmcs_file, jst_file)}
    assert before == after


def test_load_roi_config_success(tmp_path: Path) -> None:
    config_path = _make_roi_config(tmp_path / "tmcs_sku_roi.json")

    loaded = load_roi_config(config_path)

    assert loaded["safe_profit_rate"] == 0.1
    assert loaded["ideal_promotion_ratio"] == 0.12


def test_load_roi_config_missing_file_fails(tmp_path: Path) -> None:
    missing_path = tmp_path / "missing.json"

    try:
        load_roi_config(missing_path)
    except ValueError as exc:
        assert "配置文件不存在" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def test_load_roi_config_invalid_json_fails(tmp_path: Path) -> None:
    config_path = tmp_path / "tmcs_sku_roi.json"
    config_path.write_text("{bad json", encoding="utf-8")

    try:
        load_roi_config(config_path)
    except ValueError as exc:
        assert "不是合法 JSON" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def test_load_roi_config_missing_required_keys_fails(tmp_path: Path) -> None:
    config_path = tmp_path / "tmcs_sku_roi.json"
    config_path.write_text(json.dumps({"supply_price_factor": 0.9}, ensure_ascii=False), encoding="utf-8")

    try:
        load_roi_config(config_path)
    except ValueError as exc:
        assert "缺少字段" in str(exc)
        assert "ideal_promotion_ratio" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def test_dry_run_uses_external_config_values(tmp_path: Path, monkeypatch) -> None:
    tmcs_file = _make_tmcs_file(tmp_path / "tmcs.xlsx")
    jst_file = _make_jst_file(tmp_path / "jst.xlsx")
    roi_config = _make_roi_config(tmp_path / "tmcs_sku_roi.json", ideal_promotion_ratio=0.2)
    monkeypatch.setattr(steps, "DEFAULT_TMCS_FILE", tmcs_file)
    monkeypatch.setattr(steps, "DEFAULT_JST_FILE", jst_file)
    monkeypatch.setattr(steps, "DEFAULT_ROI_CONFIG_FILE", roi_config)

    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(
        build_workflow(),
        inputs={"dry_run": True, "args": ["--dry-run", "--sku-code", "SKU001"]},
        dry_run=True,
    )

    assert run.status == "dry_run_success"
    run_json = json.loads((runner.last_run_dir / "run.json").read_text(encoding="utf-8"))
    assert run_json["outputs"]["理想ROI"] == "5.0000"
