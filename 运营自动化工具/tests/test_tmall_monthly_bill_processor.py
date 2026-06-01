from __future__ import annotations
import argparse
import sys

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from tasks.tmall_monthly_bill import main as tmall_main
from tasks.tmall_monthly_bill import processor as tmall_processor
from tasks.tmall_monthly_bill.services.profit_summary_service import _centered_summary_start_row, render_profit_summary
from tasks.tmall_monthly_bill.services.promotion_service import write_promotion_sheet
from tasks.tmall_monthly_bill.services.reconciliation_service import write_reconciliation_sheet


def sheet_values(sheet) -> list[tuple[object, ...]]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def test_rebuilds_reconciliation_sheet_with_original_header_order(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "开票表"
    stale = workbook.create_sheet("对账单列表")
    stale.append(["旧列"])

    source = tmp_path / "对账单列表.xlsx"
    data = Workbook()
    sheet = data.active
    sheet.append(["账单编号", "创建时间", "状态"])
    sheet.append(["B001", "2026-04-01", "成功"])
    data.save(source)

    write_reconciliation_sheet(workbook, source)

    assert sheet_values(workbook["对账单列表"]) == [
        ("账单编号", "创建时间", "状态"),
        ("B001", "2026-04-01", "成功"),
    ]


def test_statement_matches_periods_requires_all_bill_periods(tmp_path: Path) -> None:
    source = tmp_path / "对账单列表.xlsx"
    data = Workbook()
    sheet = data.active
    sheet.append(["账单周期", "商家开票含税总额"])
    sheet.append(["2026-04-01~2026-04-10", 100])
    sheet.append(["2026-04-11~2026-04-20", 200])
    data.save(source)

    class FakeSource:
        load_workbook = staticmethod(__import__("openpyxl").load_workbook)

    assert (
        tmall_main.statement_matches_periods(
            FakeSource,
            source,
            {"2026-04-01~2026-04-10", "2026-04-11~2026-04-20"},
        )
        is True
    )
    assert (
        tmall_main.statement_matches_periods(
            FakeSource,
            source,
            {"2026-04-01~2026-04-10", "2026-04-11~2026-04-20", "2026-04-21~2026-04-30"},
        )
        is False
    )


def test_writes_promotion_sheet_from_csv_without_changing_column_order(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "开票表"

    source = tmp_path / "万象台推广账单_2026-04.csv"
    source.write_text("收支类型,金额,备注\n支出,12.50,投放A\n", encoding="utf-8-sig")

    write_promotion_sheet(workbook, "万相台推广数据表格", source)

    assert sheet_values(workbook["万相台推广数据表格"]) == [
        ("收支类型", "金额", "备注"),
        ("支出", 12.5, "投放A"),
    ]


def test_build_invoice_sheet_rounds_bill_and_ticket_amounts_to_cents() -> None:
    cargo_header = ["后端商品编码", "商品编码", "品名", "商品数量", "含税单价"]
    cargo_rows = [
        ["B1", "P1", "商品1", 1, "10.005"],
        ["B2", "P1", "商品1", 1, "20.005"],
    ]
    ticket_header = ["后端商品编码", "商品编码", "含税金额"]
    ticket_rows = [
        ["B1", "P1", "0.004"],
        ["B2", "P1", "0.006"],
    ]

    invoice_header, invoice_rows, _ = tmall_processor.build_invoice_sheet(cargo_header, cargo_rows, ticket_header, ticket_rows)
    bill_idx = invoice_header.index("账单金额")
    ticket_idx = invoice_header.index("票扣")
    invoice_idx = invoice_header.index("开票金额")

    bill_amounts = [Decimal(str(row[bill_idx])) for row in invoice_rows]
    ticket_amounts = [Decimal(str(row[ticket_idx])) for row in invoice_rows]
    invoice_amounts = [Decimal(str(row[invoice_idx])) for row in invoice_rows]

    assert bill_amounts == [Decimal("10"), Decimal("20.01")]
    assert ticket_amounts == [Decimal("0"), Decimal("0.01")]
    assert invoice_amounts == [Decimal("10"), Decimal("20.02")]


def test_build_invoice_sheet_rebalances_ticket_total_to_match_grand_total_cents() -> None:
    cargo_header = ["后端商品编码", "商品编码", "品名", "商品数量", "含税单价"]
    cargo_rows = [
        ["B1", "P1", "商品1", 1, "10.00"],
        ["B2", "P2", "商品2", 1, "10.00"],
        ["B3", "P3", "商品3", 1, "10.00"],
        ["B4", "P4", "商品4", 1, "10.00"],
        ["B5", "P5", "商品5", 1, "10.00"],
    ]
    ticket_header = ["后端商品编码", "商品编码", "含税金额"]
    ticket_rows = [
        ["B1", "P1", "0.005"],
        ["B2", "P2", "0.005"],
        ["B3", "P3", "0.005"],
        ["B4", "P4", "0.005"],
        ["B5", "P5", "0.005"],
    ]

    invoice_header, invoice_rows, _ = tmall_processor.build_invoice_sheet(cargo_header, cargo_rows, ticket_header, ticket_rows)
    ticket_idx = invoice_header.index("票扣")
    invoice_idx = invoice_header.index("开票金额")

    ticket_amounts = [Decimal(str(row[ticket_idx])) for row in invoice_rows]
    invoice_amounts = [Decimal(str(row[invoice_idx])) for row in invoice_rows]

    assert sum(ticket_amounts) == Decimal("0.02")
    assert sum(invoice_amounts) == Decimal("50.02")


def test_build_invoice_sheet_falls_back_to_backend_code_for_blank_product_code_ticket_rows() -> None:
    cargo_header = ["后端商品编码", "商品编码", "品名", "商品数量", "含税单价"]
    cargo_rows = [["B1", None, "商品1", 2, "100.00"]]
    ticket_header = ["后端商品编码", "商品编码", "含税金额"]
    ticket_rows = [["B1", None, "-12.34"]]

    invoice_header, invoice_rows, _ = tmall_processor.build_invoice_sheet(cargo_header, cargo_rows, ticket_header, ticket_rows)
    ticket_idx = invoice_header.index("票扣")
    invoice_idx = invoice_header.index("开票金额")

    assert Decimal(str(invoice_rows[0][ticket_idx])) == Decimal("-12.34")
    assert Decimal(str(invoice_rows[0][invoice_idx])) == Decimal("187.66")


def test_renders_profit_summary_in_invoice_sheet_right_side() -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "开票表"
    invoice.append(["后端商品编码", "商品数量", "票扣", "账单金额", "开票金额", "成本"])
    invoice.append(["A1", 2, -10.50, 118.80, 120.30, 40])
    invoice.append(["A2", 3, "-5.00", "199.20", "200.20", "50"])

    cost = workbook.create_sheet("成本表")
    cost.append(["后端商品编码", "金额"])
    cost.append(["A1", 80])
    cost.append(["A2", "150.00"])

    charge = workbook.create_sheet("账扣表格")
    charge.append(["含税金额"])
    charge.append([-12.34])

    wxt = workbook.create_sheet("万相台推广数据表格")
    wxt.append(["收支类型", "金额"])
    wxt.append(["支出", "30.25"])
    wxt.append(["收入", "5.00"])

    zdx = workbook.create_sheet("智多星推广数据表格")
    zdx.append(["类型", "收支金额"])
    zdx.append(["从冻结中转出", "18.75"])
    zdx.append(["其他", "1.00"])

    start_col = invoice.max_column + 2
    render_profit_summary(workbook, month_label="4月份利润表")

    start_row = _centered_summary_start_row(invoice, 7)
    assert invoice.cell(start_row, start_col).value == "4月份利润表"
    assert invoice.cell(start_row + 1, start_col).value == "销售金额（开票金额）"
    assert invoice.cell(start_row + 1, start_col + 1).value == 5
    assert invoice.cell(start_row + 1, start_col + 2).value == 318
    assert invoice.cell(start_row + 4, start_col + 2).value == 15.5
    assert invoice.cell(start_row + 5, start_col + 2).value == 12.34
    assert invoice.cell(start_row + 6, start_col + 2).value == 11.16


def test_profit_summary_falls_back_to_invoice_cost_when_cost_sheet_missing() -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "开票表"
    invoice.append(["商品数量", "票扣", "账单金额", "开票金额", "成本"])
    invoice.append([2, 0, 100, 101, 20])
    invoice.append([1, 0, 60, 61, 30])

    charge = workbook.create_sheet("账扣表格")
    charge.append(["含税金额"])
    charge.append([0])

    wxt = workbook.create_sheet("万相台推广数据表格")
    wxt.append(["收支类型", "金额"])
    wxt.append(["支出", 10])

    zdx = workbook.create_sheet("智多星推广数据表格")
    zdx.append(["类型", "金额"])
    zdx.append(["从冻结中转出", 5])

    start_col = invoice.max_column + 2
    render_profit_summary(workbook, month_label="4月份利润表")

    start_row = _centered_summary_start_row(invoice, 7)
    assert invoice.cell(start_row + 2, start_col + 2).value == 70


def test_profit_summary_limits_zdx_full_ledger_to_bill_period() -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "开票表"
    invoice.append(["商品数量", "票扣", "账单金额", "开票金额", "成本"])
    invoice.append([1, 0, 100, 100, 20])

    cost = workbook.create_sheet("成本表")
    cost.append(["金额"])
    cost.append([20])

    charge = workbook.create_sheet("账扣表格")
    charge.append(["含税金额"])
    charge.append([0])

    wxt = workbook.create_sheet("万相台推广数据表格")
    wxt.append(["收支类型", "金额"])
    wxt.append(["支出", 0])

    zdx = workbook.create_sheet("智多星推广数据表格")
    zdx.append(["类型", "资金明细", "时间"])
    zdx.append(["从冻结中转出", 5, "2026-04-23 09:00:43"])
    zdx.append(["从冻结中转出", 99, "2026-05-07 14:00:07"])

    start_col = invoice.max_column + 2
    render_profit_summary(
        workbook,
        month_label="4月份利润表",
        period_start="2026-04-01",
        period_end="2026-04-30",
    )

    start_row = _centered_summary_start_row(invoice, 7)
    assert invoice.cell(start_row + 3, start_col + 2).value == 5
    assert invoice.cell(start_row + 6, start_col + 2).value == 75


def test_profit_summary_starts_from_center_of_invoice_data_area() -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "开票表"
    invoice.append(["商品数量", "票扣", "账单金额", "开票金额", "成本"])
    for _ in range(20):
        invoice.append([1, 0, 10, 11, 3])

    charge = workbook.create_sheet("账扣表格")
    charge.append(["含税金额"])
    charge.append([0])

    wxt = workbook.create_sheet("万相台推广数据表格")
    wxt.append(["收支类型", "金额"])
    wxt.append(["支出", 0])

    zdx = workbook.create_sheet("智多星推广数据表格")
    zdx.append(["类型", "金额"])
    zdx.append(["从冻结中转出", 0])

    start_col = invoice.max_column + 2
    render_profit_summary(workbook, month_label="4月份利润表")

    start_row = _centered_summary_start_row(invoice, 7)
    assert start_row == 8
    assert invoice.cell(start_row, start_col).value == "4月份利润表"


def test_readme_mentions_new_output_sheets() -> None:
    readme_path = Path(__file__).resolve().parents[1] / "tasks" / "tmall_monthly_bill" / "README.md"
    text = readme_path.read_text(encoding="utf-8")
    assert "对账单列表" in text
    assert "万相台推广数据表格" in text
    assert "智多星推广数据表格" in text
    assert "利润汇总" in text
    assert "不再移动" in text


def test_statement_auto_download_uses_last_month_for_full_previous_month_range(monkeypatch) -> None:
    recorded = {}
    calls = {"count": 0}

    def fake_run(command, text, capture_output):
        recorded["command"] = command

        class Result:
            returncode = 0
            stdout = ""
            stderr = ""

        return Result()

    from datetime import date as real_date

    class FakeDate(real_date):
        @classmethod
        def today(cls):
            return real_date(2026, 5, 15)

    monkeypatch.setattr(tmall_main, "date", FakeDate)
    monkeypatch.setattr(tmall_main, "statement_list_candidates", lambda path: [])
    monkeypatch.setattr(tmall_main, "infer_bill_periods", lambda bill_files: ["2026-04-01~2026-04-10"])
    monkeypatch.setattr(tmall_main, "infer_bill_date_range", lambda bill_files: ("2026-04-01", "2026-04-30"))
    monkeypatch.setattr(tmall_main, "choose_downloader_python", lambda: "/usr/bin/python3")
    monkeypatch.setattr(tmall_main.subprocess, "run", fake_run)
    def fake_find_matching_statement_list(source, statement_path, bill_periods):
        calls["count"] += 1
        if calls["count"] == 1:
            return None
        return Path("/tmp/对账单列表.xlsx")

    monkeypatch.setattr(tmall_main, "find_matching_statement_list", fake_find_matching_statement_list)

    args = argparse.Namespace(
        skip_auto_download=False,
        dry_run=False,
        downloader_script="/tmp/downloader.py",
    )

    result = tmall_main.auto_download_statement_list_if_needed(
        args,
        source=None,
        statement_path=Path("/tmp/对账单列表.xlsx"),
        bill_files=[Path("/tmp/HDB202604012026041033586806223.xlsx")],
        bill_dir=Path("/tmp"),
    )

    assert "--last-month" in recorded["command"]
    assert "--start" not in recorded["command"]
    assert result["statement_auto_download_attempted"] is True


def test_process_keeps_source_files_in_place_after_generating_archive_workbook(tmp_path: Path, monkeypatch) -> None:
    bill_dir = tmp_path / "downloads"
    work_dir = tmp_path / "workdir"
    output_dir = tmp_path / "desktop"
    bill_dir.mkdir()
    work_dir.mkdir()

    bill_file = bill_dir / "HDB202604012026041033586806223-1603202.xlsx"
    bill_file.write_text("placeholder", encoding="utf-8")
    statement_file = bill_dir / "对账单列表.xlsx"
    statement_book = Workbook()
    statement_sheet = statement_book.active
    statement_sheet.append(["账单周期", "商家开票含税总额"])
    statement_sheet.append(["2026-04-01~2026-04-10", 100])
    statement_book.save(statement_file)
    table1 = work_dir / "table1.xlsx"
    table2 = work_dir / "table2.xlsx"
    Workbook().save(table1)
    Workbook().save(table2)

    wxt_file = bill_dir / "万象台推广账单_2026-04.csv"
    wxt_file.write_text("收支类型,金额\n支出,10\n", encoding="utf-8-sig")
    zdx_file = bill_dir / "场景智投资金账户明细导出.xlsx"
    zdx_book = Workbook()
    zdx_sheet = zdx_book.active
    zdx_sheet.append(["类型", "金额", "时间"])
    zdx_sheet.append(["从冻结中转出", 5, "2026-04-02 09:00:00"])
    zdx_book.save(zdx_file)

    class FakeSource:
        Workbook = Workbook
        MAIN_SHEET_TEMPLATE = "猫超{month}月账单数据表格"
        OUTPUT_TEMPLATE = "猫超{month}月账单数据表格.xlsx"
        ARCHIVE_ROOT_NAME = "猫超月账单数据"
        ARCHIVE_SUBDIR_TEMPLATE = "{month}月对账数据"

        @staticmethod
        def get_bill_files(path: Path):
            return [path / bill_file.name]

        @staticmethod
        def infer_month_from_bills(_bill_files):
            return 4

        @staticmethod
        def build_combined_rows(_bill_files):
            return ["对账单号", "商品数量", "账单金额", "票扣", "成本", "含税金额"], [["HDB202604012026041033586806223", 1, 100, -8, 30, -3]]

        @staticmethod
        def build_table1_mapping(_path):
            return {}

        @staticmethod
        def build_table2_mapping(_path):
            return {}

        @staticmethod
        def enrich_rows(raw_header, raw_rows, *_args):
            return raw_header, raw_rows, {"mapped_table1": 1, "unmatched_table1": 0, "mapped_table2": 1, "unmatched_table2": 0}

        @staticmethod
        def sort_rows_by_backend_code(_raw_header, raw_rows):
            return raw_rows

        @staticmethod
        def build_sub_sheet_rows(header, rows, sheet_name):
            if sheet_name == "货款表格":
                return [rows[0]]
            if sheet_name == "票扣表格":
                return [rows[0]]
            if sheet_name == "账扣表格":
                return [rows[0]]
            return []

        @staticmethod
        def build_invoice_sheet(*_args):
            return ["商品数量", "票扣", "账单金额", "开票金额", "成本"], [[1, -8, 100, 100, 30]], []

        @staticmethod
        def build_cost_sheet(*_args):
            return ["金额"], [[30]]

        @staticmethod
        def ensure_clean_target(path: Path):
            if path.exists():
                path.unlink()

        @staticmethod
        def append_sheet(workbook, title, header, rows):
            sheet = workbook.create_sheet(title)
            sheet.append(header)
            for row in rows:
                sheet.append(row)

    monkeypatch.setattr(tmall_main, "auto_download_bills_if_needed", lambda args, bill_dir: {"auto_download_attempted": False})
    monkeypatch.setattr(tmall_main, "DEFAULT_OUTPUT_DIR", output_dir)
    monkeypatch.setattr(tmall_main, "load_source_module", lambda path: FakeSource)
    monkeypatch.setattr(
        tmall_main,
        "auto_download_statement_list_if_needed",
        lambda args, source, statement_path, bill_files, bill_dir: {"statement_auto_download_attempted": False, "statement_list_before_process": str(statement_file)},
    )
    monkeypatch.setattr(
        tmall_main,
        "compare_invoice_amount",
        lambda source, statement_path, invoice_header, invoice_rows, bill_files: {
            "statement_list": str(statement_file),
            "bill_year": 2026,
            "bill_month": 4,
            "bill_periods": ["2026-04-01~2026-04-10"],
            "matched_statement_periods": ["2026-04-01~2026-04-10"],
            "invoice_amount_sum": "100",
            "statement_invoice_tax_total_sum": "100",
            "invoice_compare_amount": 100,
            "statement_compare_amount": 100,
            "invoice_amount_check": "开票金额正确",
        },
    )
    monkeypatch.setattr(tmall_main, "infer_bill_date_range", lambda bill_files: ("2026-04-01", "2026-04-10"))
    monkeypatch.setattr(tmall_main, "download_promotion_bill", lambda source_name, start, end, bill_dir: wxt_file if source_name == "wxt" else zdx_file)

    args = argparse.Namespace(
        bill_dir=str(bill_dir),
        work_dir=str(work_dir),
        source_script=str(tmp_path / "processor.py"),
        statement_list=str(statement_file),
        table1_file=str(table1),
        table2_file=str(table2),
        downloader_script=str(tmp_path / "downloader.py"),
        skip_auto_download=False,
        dry_run=False,
    )

    payload = tmall_main.process(args)

    output_path = output_dir / "猫超4月账单数据表格.xlsx"
    assert output_path.exists()
    assert bill_file.exists()
    assert statement_file.exists()
    assert payload["bill_files_moved"] is False
    assert payload["statement_list_moved"] is False
    assert payload["statement_list_archive_path"] is None


def test_find_existing_promotion_bill_prefers_current_bill_dir(tmp_path: Path) -> None:
    bill_dir = tmp_path / "downloads"
    bill_dir.mkdir()

    expected = bill_dir / "万相台推广账单_2026-04.csv"
    expected.write_text("收支类型,金额\n支出,10\n", encoding="utf-8-sig")

    other_dir = tmp_path / "other"
    other_dir.mkdir()
    (other_dir / "万相台推广账单_2026-04(1).csv").write_text("收支类型,金额\n支出,20\n", encoding="utf-8-sig")

    found = tmall_main.find_existing_promotion_bill("wxt", "2026-04-01", bill_dir)

    assert found == expected.resolve()


def test_download_promotion_bill_uses_existing_download_before_cli(tmp_path: Path, monkeypatch) -> None:
    bill_dir = tmp_path / "downloads"
    bill_dir.mkdir()
    expected = bill_dir / "智多星推广账单_2026-04.xlsx"
    expected.write_bytes(b"existing")
    calls = {"count": 0}

    def fake_run_ops_json(*_args, **_kwargs):
        calls["count"] += 1
        return {"data": {"downloaded_files": []}}

    monkeypatch.setattr(tmall_main, "run_ops_json", fake_run_ops_json)

    found = tmall_main.download_promotion_bill("zdx", "2026-04-01", "2026-04-30", bill_dir)

    assert found == expected.resolve()
    assert calls["count"] == 0


def test_download_promotion_bill_enables_interactive_recovery_for_cli(tmp_path: Path, monkeypatch) -> None:
    bill_dir = tmp_path / "downloads"
    bill_dir.mkdir()
    generated_dir = tmp_path / "generated"
    generated_dir.mkdir()
    expected = generated_dir / "智多星推广账单_2026-04.xlsx"
    expected.write_bytes(b"downloaded")
    observed: dict[str, object] = {}

    def fake_run_ops_json(args, **kwargs):
        observed["args"] = args
        observed["kwargs"] = kwargs
        return {"data": {"downloaded_files": [str(expected)]}}

    monkeypatch.setattr(tmall_main, "run_ops_json", fake_run_ops_json)

    found = tmall_main.download_promotion_bill("zdx", "2026-04-01", "2026-04-30", bill_dir)

    assert found == expected.resolve()
    assert observed["kwargs"] == {"interactive_recovery": True}


def test_legacy_main_routes_to_workflow_without_processing(monkeypatch) -> None:
    calls: list[list[str]] = []
    argv = ["tmall_monthly_bill", "--dry-run"]
    monkeypatch.setattr(sys, "argv", argv)
    monkeypatch.setattr(tmall_main, "_run_workflow", lambda args: calls.append(list(args)) or 0, raising=False)
    monkeypatch.setattr(tmall_main, "process", lambda *a, **k: (_ for _ in ()).throw(AssertionError("旧入口不应直接整理账单")))

    assert tmall_main.main() == 0
    assert calls == [["tmall_monthly_bill", "--dry-run"]]
