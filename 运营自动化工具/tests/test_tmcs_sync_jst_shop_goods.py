from __future__ import annotations

import sys
import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook


SKILL_DIR = Path(__file__).resolve().parents[1] / "skills" / "tmcs_sync_jst_shop_goods"
sys.path.insert(0, str(SKILL_DIR))

from excel_builder import IMPORT_HEADERS, build_import_workbooks, build_rows  # noqa: E402
from input_loader import load_item_ids_from_excel, parse_item_ids  # noqa: E402
import cli_client as skill_cli_client  # noqa: E402

PROJECT_ROOT = SKILL_DIR.parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from core.task_registry import resolve_task, task_scripts  # noqa: E402


def test_skill_does_not_contain_platform_browser_automation_code() -> None:
    forbidden = ["playwright", "connect_over_cdp", "cookie", "localStorage", "sessionStorage", "http://", "https://", "selector"]
    for path in SKILL_DIR.glob("*.py"):
        if path.name == "cli_client.py":
            continue
        text = path.read_text(encoding="utf-8")
        lowered = text.lower()
        for token in forbidden:
            assert token.lower() not in lowered, f"{path.name} should not contain platform-side token {token}"


def test_parse_item_ids_dedupes_and_preserves_order() -> None:
    assert parse_item_ids("123, 234,123,,345") == ["123", "234", "345"]


def test_load_item_ids_from_excel_accepts_alias_headers(tmp_path: Path) -> None:
    path = tmp_path / "商品ID列表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["商品ID", "备注"])
    ws.append(["1052534376394", "a"])
    ws.append(["1052534376394", "duplicate"])
    ws.append(["6247519890566", "b"])
    wb.save(path)

    assert load_item_ids_from_excel(path) == ["1052534376394", "6247519890566"]


def test_build_rows_maps_stock_to_jst_import_shape() -> None:
    import_rows, failures = build_rows(
        requested_item_ids=["1052534376394"],
        stock_rows=[
            {
                "platform_item_id": "1052534376394",
                "platform_sku_id": "6247519890565",
                "supplier_goods_id": "SUP-001",
                "merchant_goods_code": "MGC-001",
            }
        ],
    )

    assert failures == []
    assert import_rows == [
        {
            "线上款式编码": "1052534376394",
            "线上商品编码": "MGC-001",
            "线上国标码": "",
            "平台店铺款式编码": "1052534376394",
            "平台店铺商品编码": "SUP-001",
            "原始商品编码": "MGC-001",
            "线上商品名称": "",
            "线上颜色规格": "",
            "商品标识": "Retail",
        }
    ]


def test_build_import_workbooks_writes_text_cells(tmp_path: Path) -> None:
    result = build_import_workbooks(
        import_rows=[
            {
                "线上款式编码": "1052534376394",
                "线上商品编码": "MGC-001",
                "线上国标码": "",
                "平台店铺款式编码": "1052534376394",
                "平台店铺商品编码": "SUP-001",
                "原始商品编码": "MGC-001",
                "线上商品名称": "",
                "线上颜色规格": "",
                "商品标识": "Retail",
            }
        ],
        failures=[],
        output_dir=tmp_path,
        timestamp="20260518_120000",
    )

    wb = load_workbook(result["import_path"])
    ws = wb.active
    assert [cell.value for cell in ws[1]] == IMPORT_HEADERS
    assert ws["A2"].value == "1052534376394"
    assert ws["A2"].number_format == "@"


def test_skill_real_platform_call_uses_shared_interactive_recovery(monkeypatch) -> None:
    observed: dict[str, object] = {}

    def fake_run_ops_json(args, **kwargs):
        observed["args"] = args
        observed["kwargs"] = kwargs
        return {"success": True, "data": {"rows": []}}

    monkeypatch.setattr(skill_cli_client, "run_ops_json", fake_run_ops_json)

    assert skill_cli_client.query_tmcs_stock(item_ids=["1001"], warehouse_code="WH") == []
    assert observed["kwargs"] == {"interactive_recovery": True}


def test_formal_task_entry_resolves_chinese_triggers() -> None:
    assert resolve_task("聚水潭商品信息同步猫超") == "tmcs_sync_jst_shop_goods"
    assert resolve_task("猫超商品信息同步聚水潭") == "tmcs_sync_jst_shop_goods"
    assert resolve_task("平台商品ID同步聚水潭") == "tmcs_sync_jst_shop_goods"
    assert task_scripts()["tmcs_sync_jst_shop_goods"] == PROJECT_ROOT / "tasks" / "tmcs_sync_jst_shop_goods" / "main.py"


def test_formal_skill_uses_run_py_entry_and_trigger_document() -> None:
    skill_doc = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    metadata = (SKILL_DIR / "skill.yaml").read_text(encoding="utf-8")

    assert "聚水潭商品信息同步猫超" in skill_doc
    assert "1052305450766 聚水潭商品信息同步猫超" in skill_doc
    assert "python3 run.py 聚水潭商品信息同步猫超" in metadata


def test_formal_task_entry_defaults_to_run_subcommand() -> None:
    result = subprocess.run(
        [sys.executable, str(task_scripts()["tmcs_sync_jst_shop_goods"]), "--help"],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert "--item-ids" in result.stdout
    assert "--import-jst" in result.stdout
