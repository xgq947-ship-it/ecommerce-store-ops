from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


TMCS_REQUIRED_HEADERS = ("商品编码", "SKU编码", "条码")
JST_REQUIRED_HEADERS = ("商品编码", "淘系控价", "成本价")
TEMPLATE_REQUIRED_ITEMS = {
    "消费者到手价": "C4",
    "供货价系数": "C5",
    "产品成本": "C7",
    "国内运费/发仓": "C8",
    "赠品成本": "C9",
    "88VIP折扣承担率": "C10",
    "通用收费率": "C11",
    "其他收费率": "C12",
    "仓储/物流费率": "C13",
    "税点": "C14",
    "公司管理费用率": "C15",
    "退款率": "C16",
    "单笔退款固定损耗": "C17",
    "目标保留利润率": "C27",
}
TEMPLATE_REQUIRED_FORMULAS = {
    "供货价": "=C4*C5",
    "真实经营利润": "=E23-E22",
    "盈亏平衡ROI": '=IF(E29>0,C4/E29,"无利润")',
    "安全ROI": '=IF(E28>0,C4/E28,"不建议推广")',
}


@dataclass
class TemplateInfo:
    path: Path
    config: dict[str, float]
    verified_items: dict[str, str]


def _normalize_header(value) -> str:
    return str(value or "").strip()


def _load_sheet(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=False)
    return workbook, workbook[workbook.sheetnames[0]]


def _header_map(sheet) -> dict[str, int]:
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for column, raw_value in enumerate(header_row, start=1):
        header = _normalize_header(raw_value)
        if header:
            headers[header] = column
    return headers


def _require_headers(headers: dict[str, int], required: tuple[str, ...], file_label: str) -> None:
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"{file_label} 缺少字段：{', '.join(missing)}")


def find_tmcs_barcode(path: Path, *, sku_code: str | None = None, product_code: str | None = None) -> dict[str, str]:
    workbook, sheet = _load_sheet(path)
    try:
        headers = _header_map(sheet)
        _require_headers(headers, TMCS_REQUIRED_HEADERS, "猫超商品列表")
        product_index = headers["商品编码"] - 1
        sku_index = headers["SKU编码"] - 1
        barcode_index = headers["条码"] - 1
        if sku_code:
            mode = "sku_code"
            lookup_value = sku_code
        elif product_code:
            mode = "product_code"
            lookup_value = product_code
        else:
            raise ValueError("缺少猫超查询键：sku_code 或 product_code")

        first_match_barcode = None
        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            if mode == "sku_code":
                value = _normalize_header(row_values[sku_index] if sku_index < len(row_values) else None)
            else:
                value = _normalize_header(row_values[product_index] if product_index < len(row_values) else None)
            if value != lookup_value:
                continue
            barcode = _normalize_header(row_values[barcode_index] if barcode_index < len(row_values) else None)
            if not barcode:
                raise ValueError(f"{'SKU' if mode == 'sku_code' else '商品编码'} {lookup_value} 已找到，但条码为空")
            if mode == "sku_code":
                return {"sku_code": lookup_value, "barcode": barcode}
            if first_match_barcode is None:
                first_match_barcode = barcode
        if first_match_barcode is not None:
            return {"product_code": lookup_value, "barcode": first_match_barcode}
    finally:
        workbook.close()
    if sku_code:
        raise ValueError(f"猫超商品列表未找到 SKU 编码：{sku_code}")
    raise ValueError(f"猫超商品列表未找到 商品编码：{product_code}")


def _parse_control_price(raw_value) -> float:
    text = str(raw_value or "").strip()
    if not text:
        raise ValueError("聚水潭商品资料中的淘系控价为空")
    parts = [part for part in text.replace("\n", " ").split() if part]
    if len(parts) > 1:
        raise ValueError(f"聚水潭商品资料中的淘系控价存在多个值：{text}")
    try:
        return float(parts[0])
    except ValueError as exc:
        raise ValueError(f"聚水潭商品资料中的淘系控价不是有效数字：{text}") from exc


def find_jst_product(path: Path, barcode_as_product_code: str) -> dict[str, float | str]:
    workbook, sheet = _load_sheet(path)
    try:
        headers = _header_map(sheet)
        _require_headers(headers, JST_REQUIRED_HEADERS, "聚水潭商品资料")
        code_index = headers["商品编码"] - 1
        price_index = headers["淘系控价"] - 1
        cost_index = headers["成本价"] - 1
        matches = []
        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            code = _normalize_header(row_values[code_index] if code_index < len(row_values) else None)
            if code != barcode_as_product_code:
                continue
            raw_cost = row_values[cost_index] if cost_index < len(row_values) else None
            matches.append(
                {
                    "product_code": code,
                    "price": _parse_control_price(row_values[price_index] if price_index < len(row_values) else None),
                    "cost": float(raw_cost),
                }
            )
    finally:
        workbook.close()

    if not matches:
        raise ValueError(f"聚水潭商品资料未找到商品编码：{barcode_as_product_code}")
    if len(matches) > 1:
        raise ValueError(f"聚水潭商品资料中商品编码重复：{barcode_as_product_code}")
    match = matches[0]
    if match["cost"] <= 0:
        raise ValueError(f"聚水潭商品资料中的成本价无效：{match['cost']}")
    return match


def load_roi_template(path: Path) -> TemplateInfo:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if "猫超ROI测算" not in workbook.sheetnames:
            raise ValueError("ROI 模板缺少工作表：猫超ROI测算")
        sheet = workbook["猫超ROI测算"]
        item_to_row = {}
        for row in range(1, (sheet.max_row or 0) + 1):
            item = _normalize_header(sheet.cell(row, 2).value)
            if item:
                item_to_row[item] = row

        verified_items: dict[str, str] = {}
        config = {}
        missing_items = [name for name in TEMPLATE_REQUIRED_ITEMS if name not in item_to_row]
        if missing_items:
            raise ValueError(f"ROI 模板缺少项目：{', '.join(missing_items)}")

        for item_name, expected_cell in TEMPLATE_REQUIRED_ITEMS.items():
            row = item_to_row[item_name]
            cell_ref = f"C{row}"
            if cell_ref != expected_cell:
                raise ValueError(f"ROI 模板项目位置变更：{item_name} 期望 {expected_cell}，实际 {cell_ref}")
            verified_items[item_name] = cell_ref
            config_key = {
                "供货价系数": "supply_price_factor",
                "88VIP折扣承担率": "vip_discount_rate",
                "通用收费率": "general_fee_rate",
                "其他收费率": "other_fee_rate",
                "仓储/物流费率": "storage_fee_rate",
                "税点": "tax_rate",
                "公司管理费用率": "management_fee_rate",
                "退款率": "refund_rate",
                "单笔退款固定损耗": "refund_flat_fee",
                "国内运费/发仓": "domestic_shipping_fee",
                "赠品成本": "gift_cost",
                "目标保留利润率": "safe_profit_rate",
            }.get(item_name)
            if config_key:
                config[config_key] = float(sheet[f"C{row}"].value)

        for item_name, expected_formula in TEMPLATE_REQUIRED_FORMULAS.items():
            row = item_to_row[item_name]
            actual_formula = str(sheet[f"E{row}"].value or "").strip()
            if actual_formula != expected_formula:
                raise ValueError(f"ROI 模板公式不匹配：{item_name} 期望 {expected_formula}，实际 {actual_formula}")

        return TemplateInfo(path=path, config=config, verified_items=verified_items)
    finally:
        workbook.close()


def write_result_json(path: Path, payload: dict) -> Path:
    import json

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def write_result_excel(path: Path, payload: dict) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ROI结果"
    sheet.append(["字段", "值"])
    for key in ("保本ROI", "安全ROI", "理想ROI"):
        sheet.append([key, payload[key]])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path
