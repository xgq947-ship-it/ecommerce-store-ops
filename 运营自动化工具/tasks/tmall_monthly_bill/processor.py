from __future__ import annotations

import argparse
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook


TABLE1_FILE_NAME = "猫超商品列表导出 (最新）.xlsx"
TABLE2_FILE_NAME = "聚水潭商品资料（最新）.xlsx"
TABLE2_ALLOWED_BRANDS = {"奥克斯", "苏泊尔"}
OUTPUT_TEMPLATE = "猫超{month}月账单数据表格.xlsx"
MAIN_SHEET_TEMPLATE = "猫超{month}月账单数据表格"
ARCHIVE_ROOT_NAME = "猫超月账单数据"
ARCHIVE_SUBDIR_TEMPLATE = "{month}月对账数据"
EXTRA_FIELDS = ["商品编码", "成本", "品名"]

SUB_SHEET_RULES = {
    "货款表格": {
        "fees": {"货款"},
        "sort_by_backend_code": True,
        "with_extras": True,
    },
    "票扣表格": {
        "fees": {"88VIP用户权益折扣", "TOB销售补差", "价保补差", "毛保", "特殊商品折扣"},
        "sort_by_backend_code": False,
        "with_extras": True,
    },
    "账扣表格": {
        "fees": {"供应商违规处罚", "消费者退款赔付", "售后客服服务费", "渠道推广服务费"},
        "sort_by_backend_code": False,
        "with_extras": False,
    },
}


@dataclass
class SheetData:
    header: list[str]
    rows: list[list[object]]


def norm(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return text


def to_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def to_excel_number(value: Decimal | None) -> object:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def round_currency(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"))


def backend_code_sort_key(value: object) -> tuple[int, str]:
    code = norm(value)
    return (1, "") if code is None else (0, code)


def load_sheet_data(path: Path, sheet_name: str | None = None) -> SheetData:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header = [norm(worksheet.cell(1, column).value) or "" for column in range(1, worksheet.max_column + 1)]
    rows: list[list[object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        row = [worksheet.cell(row_index, column).value for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return SheetData(header=header, rows=rows)


def get_bill_files(folder: Path) -> list[Path]:
    files = sorted(path for path in folder.glob("HDB*.xlsx") if path.is_file())
    if not files:
        raise FileNotFoundError("当前目录未找到 HDB*.xlsx 账单文件。")
    return files


def infer_month_from_bills(bill_files: list[Path]) -> str:
    for path in bill_files:
        stem = path.stem
        if len(stem) >= 9 and stem.startswith("HDB"):
            month = stem[7:9]
            if month.isdigit():
                return str(int(month))
    raise ValueError("无法从账单文件名推断月份。")


def build_combined_rows(bill_files: list[Path]) -> tuple[list[str], list[list[object]]]:
    combined_header: list[str] | None = None
    combined_rows: list[list[object]] = []
    for path in bill_files:
        sheet = load_sheet_data(path)
        if combined_header is None:
            combined_header = sheet.header
        elif sheet.header != combined_header:
            raise ValueError(f"表头不一致，无法合并: {path.name}")
        combined_rows.extend(sheet.rows)
    if combined_header is None:
        raise ValueError("账单文件为空，无法处理。")
    return combined_header, combined_rows


def set_column_widths(worksheet, header: list[str], rows: list[list[object]]) -> None:
    sample_rows = rows[:300]
    for column_index, title in enumerate(header, start=1):
        values = ["" if title is None else str(title)]
        for row in sample_rows:
            value = row[column_index - 1] if column_index - 1 < len(row) else None
            values.append("" if value is None else str(value))
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = min(max(max_length + 2, 10), 28)


def append_sheet(workbook: Workbook, title: str, header: list[str], rows: list[list[object]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    set_column_widths(worksheet, header, rows)


def build_table1_mapping(table1_path: Path) -> dict[str, object]:
    sheet = load_sheet_data(table1_path)
    goods_code_idx = sheet.header.index("货品编码")
    barcode_idx = sheet.header.index("条码")

    mapping: dict[str, object] = {}
    for row in sheet.rows:
        goods_code = norm(row[goods_code_idx] if goods_code_idx < len(row) else None)
        if not goods_code:
            continue
        mapping[goods_code] = row[barcode_idx] if barcode_idx < len(row) else None
    return mapping


def build_table2_mapping(table2_path: Path) -> dict[str, dict[str, object]]:
    sheet = load_sheet_data(table2_path)
    code_idx = sheet.header.index("商品编码")
    cost_idx = sheet.header.index("成本价")
    brand_idx = sheet.header.index("品牌")
    category_idx = sheet.header.index("分类")

    mapping: dict[str, dict[str, object]] = {}
    for row in sheet.rows:
        brand = norm(row[brand_idx] if brand_idx < len(row) else None)
        if brand not in TABLE2_ALLOWED_BRANDS:
            continue
        product_code = norm(row[code_idx] if code_idx < len(row) else None)
        if not product_code:
            continue
        mapping[product_code] = {
            "成本": row[cost_idx] if cost_idx < len(row) else None,
            "品名": row[category_idx] if category_idx < len(row) else None,
        }
    return mapping


def enrich_rows(
    header: list[str],
    rows: list[list[object]],
    table1_mapping: dict[str, object],
    table2_mapping: dict[str, dict[str, object]],
) -> tuple[list[str], list[list[object]], dict[str, int]]:
    backend_code_idx = header.index("后端商品编码")
    enriched_rows: list[list[object]] = []
    mapped_table1 = 0
    mapped_table2 = 0
    unmatched_table1 = 0
    unmatched_table2 = 0

    for row in rows:
        new_row = row[:]
        backend_code = norm(row[backend_code_idx] if backend_code_idx < len(row) else None)
        product_code = table1_mapping.get(backend_code)
        cost = None
        product_name = None
        if product_code is None:
            unmatched_table1 += 1
        else:
            mapped_table1 += 1
            table2_record = table2_mapping.get(norm(product_code))
            if table2_record is None:
                unmatched_table2 += 1
            else:
                cost = table2_record["成本"]
                product_name = table2_record["品名"]
                mapped_table2 += 1
        new_row.extend([product_code, cost, product_name])
        enriched_rows.append(new_row)

    return header + EXTRA_FIELDS, enriched_rows, {
        "mapped_table1": mapped_table1,
        "unmatched_table1": unmatched_table1,
        "mapped_table2": mapped_table2,
        "unmatched_table2": unmatched_table2,
    }


def sort_rows_by_backend_code(header: list[str], rows: list[list[object]]) -> list[list[object]]:
    backend_code_idx = header.index("后端商品编码")
    return sorted(rows, key=lambda row: backend_code_sort_key(row[backend_code_idx] if backend_code_idx < len(row) else None))


def build_sub_sheet_rows(header: list[str], rows: list[list[object]], sheet_name: str) -> list[list[object]]:
    fee_idx = header.index("费用类型")
    backend_code_idx = header.index("后端商品编码")
    rule = SUB_SHEET_RULES[sheet_name]
    filtered = [row[:] for row in rows if row[fee_idx] in rule["fees"]]
    if rule["sort_by_backend_code"]:
        filtered.sort(key=lambda row: backend_code_sort_key(row[backend_code_idx] if backend_code_idx < len(row) else None))
    return filtered


def choose_first_non_empty(values: list[object]) -> object:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def allocate_amount_by_ratio(amount: Decimal, weights: list[Decimal]) -> list[Decimal]:
    if not weights:
        return []
    amount = round_currency(amount)
    total_weight = sum(weights, Decimal("0"))
    if total_weight == Decimal("0"):
        allocations = [Decimal("0")] * len(weights)
        if amount != Decimal("0"):
            allocations[0] = amount
        return allocations

    allocations: list[Decimal] = []
    allocated_sum = Decimal("0")
    for index, weight in enumerate(weights):
        if index == len(weights) - 1:
            share = amount - allocated_sum
        else:
            share = round_currency(amount * weight / total_weight)
            allocated_sum += share
        allocations.append(share)
    return allocations


def build_invoice_sheet(
    cargo_header: list[str],
    cargo_rows: list[list[object]],
    ticket_header: list[str],
    ticket_rows: list[list[object]],
) -> tuple[list[str], list[list[object]], list[str]]:
    cargo_backend_idx = cargo_header.index("后端商品编码")
    cargo_product_idx = cargo_header.index("商品编码")
    cargo_name_idx = cargo_header.index("品名")
    cargo_qty_idx = cargo_header.index("商品数量")
    cargo_price_idx = cargo_header.index("含税单价")
    ticket_backend_idx = ticket_header.index("后端商品编码") if "后端商品编码" in ticket_header else None
    ticket_product_idx = ticket_header.index("商品编码") if "商品编码" in ticket_header else None
    ticket_amount_idx = ticket_header.index("含税金额") if "含税金额" in ticket_header else None

    cargo_groups: dict[tuple[str, str | None], dict[str, object]] = {}
    price_conflict_codes: set[str] = set()
    backend_price_values: dict[str, set[str]] = {}
    for row in cargo_rows:
        backend_code = norm(row[cargo_backend_idx] if cargo_backend_idx < len(row) else None)
        if not backend_code:
            continue
        price_value = row[cargo_price_idx] if cargo_price_idx < len(row) else None
        price_norm = norm(price_value)
        if price_norm is not None:
            backend_price_values.setdefault(backend_code, set()).add(price_norm)
        group_key = (backend_code, price_norm)
        group = cargo_groups.setdefault(
            group_key,
            {
                "后端商品编码": backend_code,
                "商品编码": None,
                "品名": None,
                "商品数量": Decimal("0"),
                "含税单价": price_value,
            },
        )
        if group["商品编码"] in (None, ""):
            group["商品编码"] = row[cargo_product_idx]
        if group["品名"] in (None, ""):
            group["品名"] = row[cargo_name_idx]

        qty = to_decimal(row[cargo_qty_idx] if cargo_qty_idx < len(row) else None)
        if qty is not None:
            group["商品数量"] += qty
        if group["含税单价"] is None:
            group["含税单价"] = price_value

    for backend_code, price_values in backend_price_values.items():
        if len(price_values) > 1:
            price_conflict_codes.add(backend_code)

    ticket_by_product_code: dict[str, Decimal] = {}
    ticket_backend_groups: dict[str, dict[str, object]] = {}
    if ticket_backend_idx is not None and ticket_product_idx is not None and ticket_amount_idx is not None:
        for row in ticket_rows:
            backend_code = norm(row[ticket_backend_idx] if ticket_backend_idx < len(row) else None)
            if not backend_code:
                continue
            group = ticket_backend_groups.setdefault(
                backend_code,
                {
                    "商品编码": None,
                    "含税金额": Decimal("0"),
                },
            )
            if group["商品编码"] in (None, ""):
                group["商品编码"] = row[ticket_product_idx]
            amount = to_decimal(row[ticket_amount_idx] if ticket_amount_idx < len(row) else None)
            if amount is not None:
                group["含税金额"] += amount

    for group in ticket_backend_groups.values():
        product_code = norm(group["商品编码"])
        if not product_code:
            continue
        ticket_by_product_code[product_code] = ticket_by_product_code.get(product_code, Decimal("0")) + group["含税金额"]

    invoice_header = ["后端商品编码", "商品编码", "品名", "商品数量", "含税单价", "账单金额", "票扣", "开票金额"]
    invoice_rows: list[list[object]] = []
    product_group_rows: dict[str, list[dict[str, object]]] = {}
    group_order: list[tuple[str, str | None]] = []

    for group_key in sorted(cargo_groups.keys(), key=lambda item: (item[0], item[1] or "")):
        group = cargo_groups[group_key]
        qty = group["商品数量"]
        price = to_decimal(group["含税单价"])
        product_code = norm(group["商品编码"])
        bill_amount = qty * price if price is not None else Decimal("0")
        row_data = {
            "后端商品编码": group["后端商品编码"],
            "商品编码": group["商品编码"],
            "品名": group["品名"],
            "商品数量": qty,
            "含税单价": group["含税单价"],
            "账单金额": bill_amount,
            "原始账单金额": bill_amount,
        }
        if product_code:
            product_group_rows.setdefault(product_code, []).append(row_data)
        else:
            product_group_rows.setdefault(f"__missing__:{group['后端商品编码']}:{len(group_order)}", []).append(row_data)
        group_order.append(group_key)

    bill_allocations: dict[int, Decimal] = {}
    ticket_allocations: dict[int, Decimal] = {}
    for rows in product_group_rows.values():
        weights = [(row["原始账单金额"] if row["原始账单金额"] is not None else Decimal("0")) for row in rows]
        bill_total = sum(weights, Decimal("0"))
        product_code = norm(rows[0]["商品编码"])
        ticket_total = ticket_by_product_code.get(product_code, Decimal("0")) if product_code else Decimal("0")
        allocated_bills = allocate_amount_by_ratio(bill_total, weights)
        allocated = allocate_amount_by_ratio(ticket_total, weights)
        for index, amount in enumerate(allocated_bills):
            bill_allocations[id(rows[index])] = amount
        for index, amount in enumerate(allocated):
            ticket_allocations[id(rows[index])] = amount

    for group_key in group_order:
        group = cargo_groups[group_key]
        qty = group["商品数量"]
        price = to_decimal(group["含税单价"])
        bill_amount = qty * price if price is not None else Decimal("0")
        row_data = None
        product_code = norm(group["商品编码"])
        for candidate_rows in product_group_rows.values():
            for candidate in candidate_rows:
                if (
                    candidate["后端商品编码"] == group["后端商品编码"]
                    and candidate["含税单价"] == group["含税单价"]
                    and candidate["商品数量"] == qty
                ):
                    row_data = candidate
                    break
            if row_data is not None:
                break
        bill_amount = bill_allocations.get(id(row_data), round_currency(bill_amount)) if row_data is not None else round_currency(bill_amount)
        ticket_amount = ticket_allocations.get(id(row_data), Decimal("0")) if row_data is not None else Decimal("0")
        if qty == Decimal("0") and ticket_amount == Decimal("0"):
            continue
        invoice_amount = bill_amount + ticket_amount
        invoice_rows.append(
            [
                group["后端商品编码"],
                group["商品编码"],
                group["品名"],
                to_excel_number(qty),
                group["含税单价"],
                to_excel_number(bill_amount),
                to_excel_number(ticket_amount),
                to_excel_number(invoice_amount),
            ]
        )

    return invoice_header, invoice_rows, sorted(price_conflict_codes)


def build_cost_sheet(cargo_header: list[str], cargo_rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    backend_idx = cargo_header.index("后端商品编码")
    product_idx = cargo_header.index("商品编码")
    name_idx = cargo_header.index("品名")
    qty_idx = cargo_header.index("商品数量")
    cost_idx = cargo_header.index("成本")

    groups: dict[str, dict[str, object]] = {}
    for row in cargo_rows:
        backend_code = norm(row[backend_idx] if backend_idx < len(row) else None)
        if not backend_code:
            continue
        group = groups.setdefault(
            backend_code,
            {
                "后端商品编码": backend_code,
                "商品编码": None,
                "品名": None,
                "商品数量": Decimal("0"),
                "成本": None,
            },
        )
        if group["商品编码"] in (None, ""):
            group["商品编码"] = row[product_idx]
        if group["品名"] in (None, ""):
            group["品名"] = row[name_idx]
        if group["成本"] in (None, ""):
            group["成本"] = row[cost_idx]
        qty = to_decimal(row[qty_idx] if qty_idx < len(row) else None)
        if qty is not None:
            group["商品数量"] += qty

    cost_header = ["后端商品编码", "商品编码", "品名", "商品数量", "成本", "金额"]
    cost_rows: list[list[object]] = []
    for backend_code in sorted(groups.keys()):
        group = groups[backend_code]
        qty = group["商品数量"]
        if qty == Decimal("0"):
            continue
        cost = to_decimal(group["成本"])
        amount = qty * cost if cost is not None else None
        cost_rows.append(
            [
                backend_code,
                group["商品编码"],
                group["品名"],
                to_excel_number(qty),
                group["成本"],
                to_excel_number(amount),
            ]
        )
    return cost_header, cost_rows


def ensure_clean_target(path: Path) -> None:
    if path.exists():
        if path.is_dir():
            raise IsADirectoryError(f"目标路径已存在同名文件夹，无法覆盖: {path}")
        path.unlink()


def process(folder: Path, table1_file_name: str, table2_file_name: str) -> Path:
    bill_files = get_bill_files(folder)
    table1_path = folder / table1_file_name
    table2_path = folder / table2_file_name
    if not table1_path.exists():
        raise FileNotFoundError(f"未找到猫超商品列表导出 (最新）.xlsx 文件: {table1_file_name}")
    if not table2_path.exists():
        raise FileNotFoundError(f"未找到聚水潭商品资料（最新）.xlsx 文件: {table2_file_name}")

    month = infer_month_from_bills(bill_files)
    main_sheet_name = MAIN_SHEET_TEMPLATE.format(month=month)
    archive_dir = folder / ARCHIVE_ROOT_NAME / ARCHIVE_SUBDIR_TEMPLATE.format(month=month)
    archive_dir.mkdir(parents=True, exist_ok=True)
    output_file_name = OUTPUT_TEMPLATE.format(month=month)
    output_path = archive_dir / output_file_name
    stale_output_path = folder / output_file_name
    ensure_clean_target(output_path)
    if stale_output_path != output_path and stale_output_path.exists():
        ensure_clean_target(stale_output_path)

    raw_header, raw_rows = build_combined_rows(bill_files)
    table1_mapping = build_table1_mapping(table1_path)
    table2_mapping = build_table2_mapping(table2_path)
    enriched_header, enriched_rows, stats = enrich_rows(raw_header, raw_rows, table1_mapping, table2_mapping)

    main_rows_sorted = sort_rows_by_backend_code(raw_header, raw_rows)
    cargo_rows = build_sub_sheet_rows(enriched_header, enriched_rows, "货款表格")
    ticket_rows = build_sub_sheet_rows(enriched_header, enriched_rows, "票扣表格")
    charge_rows = build_sub_sheet_rows(raw_header, raw_rows, "账扣表格")
    invoice_header, invoice_rows, price_conflict_codes = build_invoice_sheet(enriched_header, cargo_rows, enriched_header, ticket_rows)
    cost_header, cost_rows = build_cost_sheet(enriched_header, cargo_rows)

    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, main_sheet_name, raw_header, main_rows_sorted)
    append_sheet(workbook, "货款表格", enriched_header, cargo_rows)
    append_sheet(workbook, "票扣表格", enriched_header, ticket_rows)
    append_sheet(workbook, "账扣表格", raw_header, charge_rows)
    append_sheet(workbook, "开票表", invoice_header, invoice_rows)
    append_sheet(workbook, "成本表", cost_header, cost_rows)
    workbook.save(output_path)

    print(f"输出文件: {output_path.name}")
    print(f"归档目录: {archive_dir}")
    print(f"账单文件数: {len(bill_files)}")
    print(f"总表数据行: {len(raw_rows)}")
    print(f"猫超商品列表导出 (最新）.xlsx 匹配成功行数: {stats['mapped_table1']}")
    print(f"猫超商品列表导出 (最新）.xlsx 未匹配行数: {stats['unmatched_table1']}")
    print(f"聚水潭商品资料（最新）.xlsx 匹配成功行数: {stats['mapped_table2']}")
    print(f"聚水潭商品资料（最新）.xlsx 未匹配行数: {stats['unmatched_table2']}")
    print(f"货款表格行数: {len(cargo_rows)}")
    print(f"票扣表格行数: {len(ticket_rows)}")
    print(f"账扣表格行数: {len(charge_rows)}")
    print(f"开票表行数: {len(invoice_rows)}")
    print(f"成本表行数: {len(cost_rows)}")
    if price_conflict_codes:
        print("以下后端商品编码存在多个不同的含税单价，开票表已按含税单价拆分记录，并按账单金额占比分摊票扣：")
        for code in price_conflict_codes:
            print(code)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="处理猫超账单并生成总表、子表、开票表和成本表。")
    parser.add_argument("--dir", default=".", help="账单和商品表所在目录，默认当前目录。")
    parser.add_argument("--table1-file", default=TABLE1_FILE_NAME, help=f"猫超商品列表导出 (最新）.xlsx 文件名，默认 {TABLE1_FILE_NAME}")
    parser.add_argument("--table2-file", default=TABLE2_FILE_NAME, help=f"聚水潭商品资料（最新）.xlsx 文件名，默认 {TABLE2_FILE_NAME}")
    args = parser.parse_args()

    folder = Path(args.dir).expanduser().resolve()
    process(folder, args.table1_file, args.table2_file)


if __name__ == "__main__":
    main()
