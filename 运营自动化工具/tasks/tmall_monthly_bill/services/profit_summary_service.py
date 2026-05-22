from __future__ import annotations

from decimal import Decimal, InvalidOperation

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


LABEL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
MONEY_FORMAT = "#,##0.00"


def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if text == "":
        return Decimal("0")
    normalized = text.replace(",", "")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _header_index(sheet, name: str) -> int:
    header = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    for index, value in enumerate(header, start=1):
        if str(value).strip() == name:
            return index
    raise ValueError(f"{sheet.title} 未找到字段：{name}")


def _optional_header_index(sheet, names: list[str]) -> int | None:
    header = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    normalized = [str(value).strip() if value is not None else "" for value in header]
    for name in names:
        if name in normalized:
            return normalized.index(name) + 1
    return None


def _sum_column(sheet, column_index: int) -> Decimal:
    total = Decimal("0")
    for row in range(2, sheet.max_row + 1):
        amount = _to_decimal(sheet.cell(row, column_index).value)
        total += amount
    return total


def _sum_filtered(sheet, filter_column: int, filter_value: str, amount_column: int) -> Decimal:
    total = Decimal("0")
    for row in range(2, sheet.max_row + 1):
        current = sheet.cell(row, filter_column).value
        if str(current).strip() != filter_value:
            continue
        total += _to_decimal(sheet.cell(row, amount_column).value)
    return total


def _cost_total_from_invoice(invoice_sheet) -> Decimal:
    qty_index = _header_index(invoice_sheet, "商品数量")
    cost_index = _header_index(invoice_sheet, "成本")
    total = Decimal("0")
    for row in range(2, invoice_sheet.max_row + 1):
        total += _to_decimal(invoice_sheet.cell(row, qty_index).value) * _to_decimal(invoice_sheet.cell(row, cost_index).value)
    return total


def _write_cell(sheet, row: int, column: int, value: object, *, bold: bool = False, fill: bool = False, align: str = "left", number_format: str | None = None) -> None:
    cell = sheet.cell(row, column, value)
    cell.border = THIN_BORDER
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        cell.fill = LABEL_FILL
    if number_format:
        cell.number_format = number_format


def _centered_summary_start_row(invoice_sheet, summary_row_count: int) -> int:
    data_start_row = 2
    data_end_row = max(invoice_sheet.max_row, data_start_row)
    data_midpoint = (data_start_row + data_end_row) / 2
    summary_offset = (summary_row_count - 1) / 2
    return max(data_start_row, int(round(data_midpoint - summary_offset)))


def render_profit_summary(workbook: Workbook, month_label: str) -> None:
    invoice_sheet = workbook["开票表"]
    cost_sheet = workbook["成本表"] if "成本表" in workbook.sheetnames else None
    charge_sheet = workbook["账扣表格"]
    wxt_sheet = workbook["万相台推广数据表格"]
    zdx_sheet = workbook["智多星推广数据表格"]

    start_col = invoice_sheet.max_column + 2
    qty_total = _sum_column(invoice_sheet, _header_index(invoice_sheet, "商品数量"))
    sales_total = _sum_column(invoice_sheet, _header_index(invoice_sheet, "账单金额"))
    ticket_total = abs(_sum_column(invoice_sheet, _header_index(invoice_sheet, "票扣")))
    charge_total = abs(_sum_column(charge_sheet, _header_index(charge_sheet, "含税金额")))
    cost_total = _sum_column(cost_sheet, _header_index(cost_sheet, "金额")) if cost_sheet else _cost_total_from_invoice(invoice_sheet)
    wxt_amount_index = _optional_header_index(wxt_sheet, ["金额", "操作金额(元)", "操作金额", "收支金额", "发生金额"])
    if wxt_amount_index is None:
        raise ValueError("万相台推广数据表格 未找到金额列")
    wxt_total = _sum_filtered(wxt_sheet, _header_index(wxt_sheet, "收支类型"), "支出", wxt_amount_index)
    zdx_amount_index = _optional_header_index(zdx_sheet, ["金额", "收支金额", "发生金额", "资金明细"])
    if zdx_amount_index is None:
        raise ValueError("智多星推广数据表格 未找到金额列")
    zdx_total = _sum_filtered(zdx_sheet, _header_index(zdx_sheet, "类型"), "从冻结中转出", zdx_amount_index)
    marketing_total = wxt_total + zdx_total
    profit_total = sales_total - cost_total - marketing_total - ticket_total - charge_total

    rows = [
        (month_label, "台数", "含税金额"),
        ("销售金额（开票金额）", int(qty_total), float(sales_total)),
        ("实际成本", None, float(cost_total)),
        ("营销推广", None, float(marketing_total)),
        ("票扣", None, float(ticket_total)),
        ("账扣", None, float(charge_total)),
        ("最终利润", None, float(profit_total)),
    ]
    start_row = _centered_summary_start_row(invoice_sheet, len(rows))

    widths = [22, 10, 14]
    for offset, width in enumerate(widths):
        invoice_sheet.column_dimensions[get_column_letter(start_col + offset)].width = width

    for row_offset, row_values in enumerate(rows):
        row_index = start_row + row_offset
        is_header = row_offset == 0
        is_final = row_offset == len(rows) - 1
        _write_cell(invoice_sheet, row_index, start_col, row_values[0], bold=is_header or is_final, fill=True)
        _write_cell(
            invoice_sheet,
            row_index,
            start_col + 1,
            row_values[1],
            bold=is_header,
            align="center" if row_values[1] is not None else "left",
            number_format="0" if isinstance(row_values[1], int) else None,
        )
        _write_cell(
            invoice_sheet,
            row_index,
            start_col + 2,
            row_values[2],
            bold=is_header or is_final,
            align="right",
            number_format=MONEY_FORMAT if isinstance(row_values[2], (int, float)) else None,
        )
