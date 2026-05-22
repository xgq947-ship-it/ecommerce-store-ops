from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def _to_numeric_if_possible(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "":
        return ""
    normalized = text.replace(",", "")
    try:
        number = float(normalized)
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def _replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    return workbook.create_sheet(sheet_name)


def _set_column_widths(worksheet, header: list[str], rows: list[list[object]]) -> None:
    sample_rows = rows[:300]
    for column_index, title in enumerate(header, start=1):
        values = [title]
        for row in sample_rows:
            value = row[column_index - 1] if column_index - 1 < len(row) else None
            values.append("" if value is None else str(value))
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = min(max(max_length + 2, 10), 28)


def _read_excel(path: Path) -> tuple[list[str], list[list[object]]]:
    source = load_workbook(path, data_only=True)
    worksheet = source[source.sheetnames[0]]
    raw_rows = list(worksheet.iter_rows(values_only=True))
    if not raw_rows:
        return [], []
    width = max((len(row) for row in raw_rows), default=0)
    if width <= 0:
        return [], []
    header = [_normalize_header(raw_rows[0][column] if column < len(raw_rows[0]) else None) for column in range(width)]
    rows: list[list[object]] = []
    for raw_row in raw_rows[1:]:
        row = [_to_numeric_if_possible(raw_row[column] if column < len(raw_row) else None) for column in range(width)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return header, rows


def _read_csv(path: Path) -> tuple[list[str], list[list[object]]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise UnicodeDecodeError("csv", b"", 0, 1, "无法识别 CSV 编码")
    if not rows:
        return [], []
    header = [_normalize_header(value) for value in rows[0]]
    body = [[_to_numeric_if_possible(value) for value in row] for row in rows[1:] if any(str(cell).strip() for cell in row if cell is not None)]
    return header, body


def read_promotion_source(source_path: Path) -> tuple[list[str], list[list[object]]]:
    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(source_path)
    return _read_excel(source_path)


def write_promotion_sheet(workbook: Workbook, sheet_name: str, source_path: Path) -> None:
    header, rows = read_promotion_source(source_path)
    worksheet = _replace_sheet(workbook, sheet_name)
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    _set_column_widths(worksheet, header, rows)
