from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def _coerce_cell_value(value: object) -> object:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return ""
    return value


def _replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    return workbook.create_sheet(sheet_name)


def _set_column_widths(worksheet, header: list[str], rows: list[list[object]]) -> None:
    sample_rows = rows[:300]
    for column_index, title in enumerate(header, start=1):
        values = [title]
        for row in sample_rows:
            value = row[column_index - 1] if column_index - 1 < len(row) else None
            values.append("" if value is None else str(value))
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = min(max(max_length + 2, 10), 28)


def _read_first_sheet(path: Path) -> tuple[list[str], list[list[object]]]:
    source = load_workbook(path, data_only=True)
    worksheet = source[source.sheetnames[0]]
    header = [_normalize_header(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]
    rows: list[list[object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        row = [_coerce_cell_value(worksheet.cell(row_index, column).value) for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return header, rows


def write_reconciliation_sheet(workbook: Workbook, source_path: Path, sheet_name: str = "对账单列表") -> None:
    header, rows = _read_first_sheet(source_path)
    worksheet = _replace_sheet(workbook, sheet_name)
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    _set_column_widths(worksheet, header, rows)
