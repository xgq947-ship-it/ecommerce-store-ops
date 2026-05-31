#!/usr/bin/env python3
"""Package buyer-show images by order id and update the brushing workbook.

This script intentionally updates xlsx files by patching sheet XML inside the
zip package instead of saving through openpyxl, because these workbooks may use
WPS DISPIMG/cellimages resources that normal workbook round-trips can break.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.config_loader import get_path  # noqa: E402


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".heic", ".webp"}
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", MAIN_NS)
MIN_IMAGES_PER_GROUP = 4


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Auto package buyer-show images.")
    parser.add_argument("--buyer-show-path", required=True)
    parser.add_argument("--model", required=True)
    parser.add_argument("--workbook")
    parser.add_argument("--groups", help="Comma-separated group folder names or numbers, e.g. 6,7,8 or 外拍1,外拍2")
    parser.add_argument("--batch", help="Only use grouped folders whose names contain this text, e.g. 外拍")
    parser.add_argument("--images-per-group", type=int, default=5, help="Maximum images copied per order folder. Groups with at least 4 images are accepted by default.")
    parser.add_argument("--allow-total-shortage", type=int, default=0, help="Deprecated compatibility option. Buyer-show groups now require more than 3 images instead of an exact target count.")
    parser.add_argument("--desktop", default=str(get_path("buyer_show_output_dir")))
    parser.add_argument("--reset-rotation", action="store_true")
    parser.add_argument("--rotation-key", help="Override the default rotation scope key.")
    parser.add_argument("--contact-sheet-only", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def latest_workbook() -> Path:
    folder = get_path("brush_register_dir")
    pattern = str(get_path("brush_register_pattern"))
    candidates = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise SystemExit(f"未找到刷单登记表：{folder}/{pattern}")
    return candidates[0]


def normalize_order_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, (int, float)):
        return from_excel(value).strftime("%Y%m%d")
    raw = str(value).strip()
    digits = re.findall(r"\d+", raw)
    if len(digits) >= 3:
        year, month, day = digits[:3]
        return f"{int(year):04d}{int(month):02d}{int(day):02d}"
    raise SystemExit(f"无法识别登记表日期：{raw}")


def read_matches(workbook: Path, model: str) -> tuple[list[dict], str, dict, dict]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb.active
    header_row_num = 2
    headers = [cell.value for cell in next(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))]
    ci = {h: i for i, h in enumerate(headers) if h is not None}
    for required in ["订单编号", "名称"]:
        if required not in ci:
            raise SystemExit(f"登记表缺少字段：{required}")

    date_col_name = next((name for name in ("刷单日期", "订单日期", "下单日期", "日期") if name in ci), None)
    status_col_name = "买家秀是否自动生成" if "买家秀是否自动生成" in ci else None

    records = []
    skipped_generated = 0
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
        name = row[ci["名称"]]
        order_id = row[ci["订单编号"]]
        if name and model in str(name) and order_id:
            record = {"row": row_num, "order_id": str(order_id).strip(), "name": str(name).strip()}
            if "刷手" in ci:
                brusher = row[ci["刷手"]]
                record["brusher"] = str(brusher).strip() if brusher not in (None, "") else ""
            if date_col_name:
                record["order_date"] = row[ci[date_col_name]]
                record["order_date_key"] = normalize_order_date(record["order_date"])
            else:
                record["order_date"] = None
                record["order_date_key"] = None
            if status_col_name:
                status_value = row[ci[status_col_name]]
                if str(status_value).strip() == "是":
                    skipped_generated += 1
                    continue
            records.append(record)

    seen = set()
    unique = []
    for record in records:
        if record["order_id"] not in seen:
            seen.add(record["order_id"])
            unique.append(record)
    if not unique and skipped_generated:
        raise SystemExit(f"该型号匹配到的订单都已生成买家秀：{model}")
    if not unique:
        raise SystemExit(f"未在登记表名称字段中找到型号：{model}")
    if not date_col_name:
        raise SystemExit("登记表缺少可用于分批的日期字段，请检查“刷单日期/订单日期/下单日期/日期”列")

    grouped: dict[str, list[dict]] = split_records_by_date(unique)
    date_keys = sorted(grouped)
    product_name = unique[0]["name"]
    summary = {
        "date_column": date_col_name,
        "skipped_generated_count": skipped_generated,
        "pending_date_keys": date_keys,
        "pending_records_by_date": {date_key: len(grouped[date_key]) for date_key in date_keys},
        "selected_order_ids": [record["order_id"] for record in unique],
    }
    return unique, product_name, ci, summary


def split_records_by_date(records: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for record in records:
        date_key = record.get("order_date_key")
        if not date_key:
            raise SystemExit(f"订单缺少可用于分批的日期：{record['order_id']}")
        grouped.setdefault(date_key, []).append(record)
    return {date_key: grouped[date_key] for date_key in sorted(grouped)}


def natural_key(text: str):
    return [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", text)]


def image_files(folder: Path) -> list[Path]:
    return sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTS], key=lambda p: natural_key(p.name.lower()))


def grouped_sources(base: Path, batch: str | None = None) -> list[tuple[str, list[Path]]]:
    groups = []
    for child in sorted([p for p in base.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
        if batch and batch not in child.name:
            continue
        imgs = image_files(child)
        if imgs:
            groups.append((child.name, imgs))
    return groups


def explicit_groups(base: Path, groups_arg: str) -> list[tuple[str, list[Path]]]:
    groups = []
    for raw in [g.strip() for g in groups_arg.split(",") if g.strip()]:
        folder = base / raw
        if not folder.is_dir() and raw.isdigit():
            folder = base / raw
        if not folder.is_dir():
            raise SystemExit(f"指定分组不存在：{raw}")
        imgs = image_files(folder)
        if not imgs:
            raise SystemExit(f"指定分组没有图片：{folder}")
        groups.append((folder.name, imgs))
    return groups


def verify_group_image_counts(groups: list[tuple[str, list[Path]]], images_per_group: int, allow_total_shortage: int = 0) -> None:
    min_images_per_group = min(MIN_IMAGES_PER_GROUP, images_per_group)
    for name, imgs in groups:
        count = len(imgs)
        if count < min_images_per_group:
            raise SystemExit(f"图片不足：分组 {name} 至少需要 {min_images_per_group} 张，当前 {count} 张")


def flat_sources(base: Path, count: int, images_per_group: int, allow_total_shortage: int = 0) -> list[tuple[str, list[Path]]]:
    min_images_per_group = min(MIN_IMAGES_PER_GROUP, images_per_group)
    flat = sorted(
        [p for p in base.rglob("*") if p.is_file() and p.suffix.lower() in IMAGE_EXTS],
        key=lambda p: natural_key(str(p.relative_to(base)).lower()),
    )
    minimum_needed = count * min_images_per_group
    if len(flat) < minimum_needed:
        raise SystemExit(f"图片不足：至少需要 {minimum_needed} 张，当前 {len(flat)} 张")

    groups = []
    cursor = 0
    remaining = len(flat)
    remaining_groups = count
    for idx in range(count):
        take = min(images_per_group, remaining - min_images_per_group * (remaining_groups - 1))
        groups.append((f"flat-{idx + 1}", flat[cursor : cursor + take]))
        cursor += take
        remaining -= take
        remaining_groups -= 1
    verify_group_image_counts(groups, images_per_group, allow_total_shortage)
    return groups


def rotation_state_path() -> Path:
    return get_path("runtime_dir") / "buyer_show_rotation_state.json"


def default_rotation_key(base: Path, model: str, batch: str | None) -> str:
    raw = json.dumps({
        "buyer_show_path": str(base.resolve()),
        "model": model.strip(),
        "batch": (batch or "").strip(),
    }, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def load_rotation_state() -> dict[str, dict]:
    path = rotation_state_path()
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def save_rotation_state(state: dict[str, dict]) -> Path:
    path = rotation_state_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def get_rotation_cursor(rotation_key: str) -> int:
    state = load_rotation_state()
    cursor = state.get(rotation_key, {}).get("cursor", 0)
    return int(cursor) if isinstance(cursor, int) or str(cursor).isdigit() else 0


def reset_rotation_cursor(rotation_key: str) -> Path:
    state = load_rotation_state()
    if rotation_key in state:
        del state[rotation_key]
    return save_rotation_state(state)


def set_rotation_cursor(rotation_key: str, cursor: int, base: Path, model: str, batch: str | None, group_names: list[str]) -> Path:
    state = load_rotation_state()
    state[rotation_key] = {
        "cursor": cursor,
        "buyer_show_path": str(base),
        "model": model,
        "batch": batch or "",
        "group_names": group_names,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    return save_rotation_state(state)


def plan_group_batches(
    records: list[dict],
    groups: list[tuple[str, list[Path]]],
    start_cursor: int,
    images_per_group: int,
    allow_total_shortage: int = 0,
) -> tuple[list[dict], int]:
    if not groups:
        raise SystemExit("没有可用的买家秀分组")

    buckets = split_records_by_date(records)
    total_needed = sum(len(items) for items in buckets.values())
    if total_needed > len(groups):
        raise SystemExit(f"分组不足：当前需要 {total_needed} 组，实际只有 {len(groups)} 组")

    ordered_groups = groups[start_cursor:] + groups[:start_cursor]
    cursor = 0
    batches = []
    for date_key, bucket_records in buckets.items():
        count = len(bucket_records)
        chosen_groups = ordered_groups[cursor: cursor + count]
        if len(chosen_groups) != count:
            raise SystemExit(f"分组不足：日期 {date_key} 需要 {count} 组，实际只剩 {len(chosen_groups)} 组")
        verify_group_image_counts(chosen_groups, images_per_group, allow_total_shortage)
        batches.append({
            "date_key": date_key,
            "records": bucket_records,
            "groups": chosen_groups,
        })
        cursor += count
    return batches, (start_cursor + total_needed) % len(groups)


def select_group_batches(
    base: Path,
    records: list[dict],
    groups_arg: str | None,
    batch: str | None,
    images_per_group: int,
    allow_total_shortage: int,
    rotation_key: str | None,
) -> tuple[list[dict], dict]:
    total_needed = len(records)
    grouped = grouped_sources(base, batch=batch)
    metadata = {
        "source_mode": "grouped" if grouped else "flat",
        "rotation_key": rotation_key,
        "rotation_cursor_before": 0,
        "rotation_cursor_after": 0,
    }

    if groups_arg:
        explicit = explicit_groups(base, groups_arg)
        if total_needed > len(explicit):
            raise SystemExit(f"分组不足：当前需要 {total_needed} 组，显式只给了 {len(explicit)} 组")
        verify_group_image_counts(explicit[:total_needed], images_per_group, allow_total_shortage)
        batches, _ = plan_group_batches(records, explicit[:total_needed], 0, images_per_group, allow_total_shortage)
        metadata["source_mode"] = "explicit_groups"
        return batches, metadata

    if grouped:
        cursor_before = get_rotation_cursor(rotation_key or "")
        if cursor_before >= len(grouped):
            cursor_before = 0
        batches, cursor_after = plan_group_batches(records, grouped, cursor_before, images_per_group, allow_total_shortage)
        metadata["rotation_cursor_before"] = cursor_before
        metadata["rotation_cursor_after"] = cursor_after
        return batches, metadata

    flat = flat_sources(base, total_needed, images_per_group, allow_total_shortage)
    batches, _ = plan_group_batches(records, flat, 0, images_per_group, allow_total_shortage)
    return batches, metadata


def safe_filename(name: str) -> str:
    return re.sub(r"[/:]", "-", name).strip()


def brusher_prefix(records: list[dict]) -> str:
    brushers = []
    for record in records:
        brusher = str(record.get("brusher", "")).strip()
        if brusher and brusher not in brushers:
            brushers.append(brusher)
    if not brushers:
        return ""
    return f"【{'+'.join(brushers)}】"


def model_code_for_filename(model: str, product_name: str) -> str:
    for text in (model, product_name):
        match = re.search(r"[A-Z0-9]+(?:-[A-Z0-9]+)+", text, flags=re.IGNORECASE)
        if match:
            return match.group(0).upper()
    return safe_filename(model or product_name)


def date_suffix_for_filename(records: list[dict]) -> str:
    normalized = []
    for record in records:
        text = record.get("order_date_key")
        if not text:
            continue
        if text not in normalized:
            normalized.append(text)
    if not normalized:
        raise SystemExit("登记表缺少可用于命名的订单日期，请检查“刷单日期/订单日期/下单日期/日期”列")
    if len(normalized) > 1:
        raise SystemExit(f"匹配到多个订单日期，无法确定压缩包命名：{', '.join(normalized)}")
    return normalized[0]


def bucket_assignments_by_brusher(assignments: list[tuple[dict, tuple[str, list[Path]]]]) -> list[tuple[str, list[tuple[dict, tuple[str, list[Path]]]]]]:
    buckets: dict[str, list[tuple[dict, tuple[str, list[Path]]]]] = {}
    order: list[str] = []
    for record, group in assignments:
        brusher = str(record.get("brusher", "")).strip()
        if brusher not in buckets:
            buckets[brusher] = []
            order.append(brusher)
        buckets[brusher].append((record, group))
    return [(brusher, buckets[brusher]) for brusher in order]


def comment_for(product_name: str, model: str, index: int) -> str:
    if "刮痧" in product_name:
        options = [
            "用了几天才来评价，实物挺有质感，握着顺手，热敷刮痧的时候温度舒服，肩颈酸的时候用一会儿会放松不少。",
            "收到后试了几次，机器不算笨重，自己拿着也方便，红光热敷看着挺明显，搭配精油用更顺滑，日常放松够用了。",
            "给家里人也试了一下，操作不复杂，按键清楚，肩膀和腿上都能用，力度自己控制，比手动刮省力很多。",
            "外观颜色挺耐看，做工也还可以，开机升温速度比较快，刮完局部会热热的，久坐后拿出来用一下挺舒服。",
            "天猫超市发货快，包装完整，奥克斯这个刮痧仪整体比预期好，平时肩颈紧、腿酸的时候用用很方便。",
        ]
    elif "手部按摩" in product_name or "手部" in product_name:
        options = [
            "用了几天才来评价，外观看着很干净，手放进去包裹感不错，按完手掌和手指会放松很多，平时电脑打字多的人用着挺合适。",
            "收到后试了一下，机器质感比想象中好，按摩力度可以调，低档比较舒服，高档按得更明显，晚上边追剧边用很方便。",
            "给家里人也试了，操作比较简单，手放进去按一下就能用，热敷温温的不会烫，用完手掌没有那么僵。",
            "外观挺耐看，放在桌面也不突兀，按摩的时候声音能接受，不影响看电视，手酸的时候用十几分钟很舒服。",
            "天猫超市发货很快，包装也完整，奥克斯这个手部按摩器做工还可以，手腕和手掌位置都能照顾到，日常放松够用了。",
        ]
    else:
        options = [
            f"收到后用了几天才来评价，{model}实物质感不错，使用起来比较方便，日常放松的时候拿出来用一会儿挺舒服。",
            "包装完整，发货也快，产品和页面描述基本一致，操作不复杂，家里人试了也觉得挺顺手。",
            "整体做工比想象中好，使用时声音能接受，放在家里日常用比较方便，属于会经常拿出来用的东西。",
            "用了几次感觉还可以，外观耐看，功能够用，不是花里胡哨的那种，日常放松和护理比较实用。",
            "天猫超市买东西比较放心，物流快，收到没有破损，试用下来符合预期，后面再看看耐用情况。",
        ]
    return options[index % len(options)]


def package_zip(records: list[dict], product_name: str, model: str, groups: list[tuple[str, list[Path]]], desktop: Path, images_per_group: int) -> tuple[Path, list[dict]]:
    date_suffix = date_suffix_for_filename(records)
    root_name = safe_filename(f"{brusher_prefix(records)}{model_code_for_filename(model, product_name)}买家秀{date_suffix}")
    zip_path = desktop / f"{root_name}.zip"
    manifest = []
    with tempfile.TemporaryDirectory(prefix="buyer_show_pack_") as td:
        root = Path(td) / root_name
        root.mkdir()
        for idx, (record, (group_name, imgs)) in enumerate(zip(records, groups)):
            dst = root / record["order_id"]
            dst.mkdir()
            chosen = imgs[:images_per_group]
            for img in chosen:
                shutil.copy2(img, dst / img.name)
            (dst / "买家秀评价.txt").write_text(comment_for(product_name, model, idx) + "\n", encoding="utf-8")
            manifest.append({"row": record["row"], "order_id": record["order_id"], "source_group": group_name, "images": [p.name for p in chosen]})
        if zip_path.exists():
            zip_path.unlink()
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            for p in sorted(root.rglob("*")):
                if p.is_file():
                    zf.write(p, p.relative_to(root.parent))
    return zip_path, manifest


def verify_zip(zip_path: Path, records: list[dict], product_name: str, images_per_group: int, allow_total_shortage: int = 0) -> dict:
    root_name = zip_path.stem
    counts = {}
    min_images_per_group = min(MIN_IMAGES_PER_GROUP, images_per_group)
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        for record in records:
            prefix = f"{root_name}/{record['order_id']}/"
            folder_names = [n for n in names if n.startswith(prefix)]
            img_count = sum(Path(n).suffix.lower() in IMAGE_EXTS for n in folder_names)
            txt_count = sum(Path(n).name == "买家秀评价.txt" for n in folder_names)
            counts[record["order_id"]] = {"img": img_count, "txt": txt_count}
            if img_count < min_images_per_group or img_count > images_per_group or txt_count != 1:
                raise SystemExit(f"压缩包核对失败：{record['order_id']} 图片 {img_count} TXT {txt_count}")
    return counts


def col_index(ref: str) -> int:
    letters = re.match(r"([A-Z]+)", ref).group(1)
    value = 0
    for ch in letters:
        value = value * 26 + ord(ch) - 64
    return value


def patch_workbook(workbook: Path, records: list[dict], ci: dict) -> tuple[Path, list[dict]]:
    backup_dir = get_path("backup_dir")
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{workbook.stem}_买家秀回写前_{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}"
    shutil.copy2(workbook, backup)

    if "买家秀是否自动生成" in ci:
        status_col = get_column_letter(ci["买家秀是否自动生成"] + 1)
        add_status_header = False
    else:
        status_col = get_column_letter(max(ci.values()) + 2)
        add_status_header = True

    rows = {r["row"] for r in records}
    with tempfile.TemporaryDirectory(prefix="buyer_show_xlsx_patch_") as td:
        tmp = Path(td) / workbook.name
        with zipfile.ZipFile(workbook, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            sheet1_seen = False
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    sheet1_seen = True
                    root = ET.fromstring(data)
                    sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
                    if sheet_data is None:
                        raise SystemExit("sheet1.xml 缺少 sheetData")
                    row_map = {int(row.attrib["r"]): row for row in sheet_data.findall(f"{{{MAIN_NS}}}row") if "r" in row.attrib}
                    if add_status_header:
                        patch_cell(row_map, sheet_data, 2, status_col, "买家秀是否自动生成")
                    for row_num in rows:
                        patch_cell(row_map, sheet_data, row_num, status_col, "是")
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                zout.writestr(item, data)
            if not sheet1_seen:
                raise SystemExit("当前仅支持首个业务工作表 sheet1.xml，未找到可回写的首个工作表")
        shutil.copy2(tmp, workbook)

    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb.active
    verify = []
    for record in records:
        row = record["row"]
        order_col = get_column_letter(ci["订单编号"] + 1)
        name_col = get_column_letter(ci["名称"] + 1)
        verify.append({
            "row": row,
            "order_id": ws[f"{order_col}{row}"].value,
            "name": ws[f"{name_col}{row}"].value,
            "status": ws[f"{status_col}{row}"].value,
        })
        if (
            str(ws[f"{order_col}{row}"].value).strip() != record["order_id"]
            or str(ws[f"{name_col}{row}"].value).strip() != record["name"]
            or ws[f"{status_col}{row}"].value != "是"
        ):
            raise SystemExit(f"登记表回写核对失败：row {row}")
    return backup, verify


def ensure_row(row_map, sheet_data, row_num: int):
    row = row_map.get(row_num)
    if row is None:
        row = ET.SubElement(sheet_data, f"{{{MAIN_NS}}}row", {"r": str(row_num)})
        row_map[row_num] = row
    return row


def clear_cell(row_map, sheet_data, row_num: int, col_letter: str) -> None:
    row = ensure_row(row_map, sheet_data, row_num)
    ref = f"{col_letter}{row_num}"
    for cell in list(row.findall(f"{{{MAIN_NS}}}c")):
        if cell.attrib.get("r") == ref:
            row.remove(cell)


def patch_cell(row_map, sheet_data, row_num: int, col_letter: str, text: str) -> None:
    row = ensure_row(row_map, sheet_data, row_num)
    ref = f"{col_letter}{row_num}"
    clear_cell(row_map, sheet_data, row_num, col_letter)
    cell = ET.Element(f"{{{MAIN_NS}}}c", {"r": ref, "t": "inlineStr"})
    is_el = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
    t_el = ET.SubElement(is_el, f"{{{MAIN_NS}}}t")
    t_el.text = text
    row.append(cell)
    cells = sorted(list(row.findall(f"{{{MAIN_NS}}}c")), key=lambda c: col_index(c.attrib.get("r", "A1")))
    for c in list(row.findall(f"{{{MAIN_NS}}}c")):
        row.remove(c)
    for c in cells:
        row.append(c)


def make_contact_sheet(base: Path, batch: str | None, out: Path) -> Path:
    from PIL import Image, ImageDraw

    groups = grouped_sources(base, batch=batch)[:20]
    if not groups:
        raise SystemExit("没有可生成缩略图的图片分组")
    thumb_w, thumb_h = 220, 220
    cols = 5
    rows = len(groups)
    canvas = Image.new("RGB", (cols * thumb_w, rows * (thumb_h + 34)), "white")
    draw = ImageDraw.Draw(canvas)
    for r, (name, imgs) in enumerate(groups):
        for c, p in enumerate(imgs[:cols]):
            try:
                im = Image.open(p).convert("RGB")
                im.thumbnail((thumb_w, thumb_h))
                x = c * thumb_w + (thumb_w - im.width) // 2
                y = r * (thumb_h + 34) + (thumb_h - im.height) // 2
                canvas.paste(im, (x, y))
                draw.text((c * thumb_w + 4, r * (thumb_h + 34) + thumb_h + 5), f"{name}:{p.name[:18]}", fill=(0, 0, 0))
            except Exception:
                draw.text((c * thumb_w + 4, r * (thumb_h + 34) + 4), f"ERR {name}", fill=(255, 0, 0))
    canvas.save(out, quality=90)
    return out


def _run_workflow(workflow_args: list[str]) -> int:
    from run import run_workflow

    return run_workflow(workflow_args)


def main(argv: list[str] | None = None) -> int:
    args = list(sys.argv[1:] if argv is None else argv)
    return _run_workflow(["buyer_show", *args])


if __name__ == "__main__":
    raise SystemExit(main())
