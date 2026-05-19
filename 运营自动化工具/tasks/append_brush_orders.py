#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请用配套的 .command 运行。") from exc

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.config_loader import get_path  # noqa: E402
from core.task_context import TaskContext  # noqa: E402


DEFAULT_WORK_DIR = get_path("brush_register_dir")
DEFAULT_PRODUCT_FILE = get_path("jst_product_master_file")
DEFAULT_BRUSH_PRODUCT_FILE = get_path("brush_product_file")
DEFAULT_WECHAT_FILE_DIR = get_path("wechat_file_dir")
DEFAULT_WECHAT_TARGET_DIR = get_path("brush_orders_dir")
RUNTIME_DIR = get_path("runtime_dir")
LATEST_BRUSH_ORDERS_PATH = RUNTIME_DIR / "latest_brush_orders.json"
TAG_JST_TASK_PATH = Path(__file__).resolve().parent / "jst_order_label" / "main.py"
BUNDLED_PYTHON = Path("/Users/dasheng/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3")

BASE_DIR = DEFAULT_WORK_DIR
SOURCE_DIR = get_path("brush_orders_dir")
BACKUP_DIR = get_path("backup_dir")
PRODUCT_FILE = DEFAULT_PRODUCT_FILE
BRUSH_PRODUCT_FILE = DEFAULT_BRUSH_PRODUCT_FILE
WECHAT_FILE_DIR = DEFAULT_WECHAT_FILE_DIR
WECHAT_TARGET_DIR = DEFAULT_WECHAT_TARGET_DIR
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML_NS = {"a": MAIN_NS}

TANGYANG_BRUSHER = "唐杨"
TANGYANG_COMMISSION = Decimal("14")
XIAOHUDIE_BRUSHER = "小蝴蝶"
TOTAL_SOURCE_BRUSHER = "小蝴蝶"
TARGET_HEADERS = [
    "序号",
    "刷手",
    "刷单日期",
    "订单编号",
    "订单金额",
    "佣金金额",
    "给刷手的转款截图1",
    "给刷手的转款截图2",
    "商品编码",
    "名称",
    "买家秀是否自动生成",
]


@dataclass(frozen=True)
class SourceRecord:
    order_no: str
    amount: Decimal
    commission: Decimal
    brusher: str
    product_code: str
    product_name: str
    source_file: str
    source_mtime: float
    source_row: int = 0


@dataclass(frozen=True)
class SourceBatch:
    source_file: Path
    source_type: str
    sheet_name: str
    month: int
    day: int
    date_text: str
    records: list[SourceRecord]


def excel_serial(dt: datetime) -> int:
    return (dt - datetime(1899, 12, 30)).days


def to_decimal(value: float | int | str | Decimal) -> Decimal:
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise RuntimeError(f"金额格式不正确：{value}") from exc


def money_text(value: Decimal) -> str:
    text = f"{value.quantize(Decimal('0.01'))}"
    return text.rstrip("0").rstrip(".") if "." in text else text


def cell_text(value: object) -> str:
    return str(value).strip() if value not in (None, "") else ""


def parse_date_marker(value: object) -> tuple[int, int, str] | None:
    text = cell_text(value)
    match = re.search(r"(\d{1,2})月\s*(\d{1,2})", text)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2)), text


def parse_standard_row_date(value: object) -> tuple[int, int, str]:
    if isinstance(value, datetime):
        return value.month, value.day, value.strftime("%-m月%-d日")
    if isinstance(value, date):
        return value.month, value.day, value.strftime("%-m月%-d日")
    if isinstance(value, (int, float)) and value not in (None, ""):
        excel_base = datetime(1899, 12, 30)
        parsed = excel_base + timedelta(days=float(value))
        return parsed.month, parsed.day, parsed.strftime("%-m月%-d日")

    text = cell_text(value)
    if not text:
        raise RuntimeError("日期为空")

    patterns = (
        r"^(\d{4})-(\d{1,2})-(\d{1,2})$",
        r"^(\d{1,2})月\s*(\d{1,2})日?$",
        r"^(\d{1,2})[.-](\d{1,2})$",
    )
    for pattern in patterns:
        match = re.match(pattern, text)
        if not match:
            continue
        if len(match.groups()) == 3:
            month = int(match.group(2))
            day = int(match.group(3))
        else:
            month = int(match.group(1))
            day = int(match.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return month, day, f"{month}月{day}日"
    raise RuntimeError(f"日期格式不支持：{text}")


def parse_month_day(value: str) -> tuple[int, int]:
    text = value.strip()
    patterns = (
        r"^(\d{1,2})(\d{2})$",
        r"^(\d{1,2})[.-](\d{1,2})$",
        r"^(\d{1,2})月\s*(\d{1,2})$",
        r"^(\d{1,2})🈷️?\s*(\d{1,2})$",
    )
    for pattern in patterns:
        match = re.match(pattern, text)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            if 1 <= month <= 12 and 1 <= day <= 31:
                return month, day
    raise argparse.ArgumentTypeError(f"日期格式不支持：{value}，可用 430、4.30、4-30、4月30")


def parse_natural_month_day(words: list[str]) -> tuple[int, int]:
    text = "".join(words).strip()
    if not text:
        today = date.today()
        return today.month, today.day
    text = text.replace("的", "").replace("日", "").strip()
    if "昨天" in text:
        target = date.today() - timedelta(days=1)
        return target.month, target.day
    if "今天" in text or "当天" in text:
        target = date.today()
        return target.month, target.day

    patterns = (
        r"(\d{1,2})月\s*(\d{1,2})",
        r"(\d{1,2})[.-](\d{1,2})",
        r"\b(\d{1,2})(\d{2})\b",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            month = int(match.group(1))
            day = int(match.group(2))
            if 1 <= month <= 12 and 1 <= day <= 31:
                return month, day
    raise argparse.ArgumentTypeError(f"无法识别日期：{text}，可说 昨天的、4月29、4.29、4-29、429")


def normalized_filename_stem(path: Path) -> str:
    return re.sub(r"\s+", "", path.stem)


def starts_with_date_prefix(text: str, prefixes: list[str]) -> bool:
    for prefix in prefixes:
        if not text.startswith(prefix):
            continue
        if len(text) == len(prefix):
            return True
        next_char = text[len(prefix)]
        if not next_char.isdigit():
            return True
    return False


def matched_wechat_source_label(path: Path, month: int, day: int) -> str | None:
    if not path.is_file():
        return None
    if path.suffix.lower() != ".xlsx":
        return None
    if path.name.startswith(("~$", ".~")):
        return None

    stem = normalized_filename_stem(path)
    total_prefix = f"【总表】天猫超市{month}.{day}账"
    if stem.startswith(total_prefix) and re.search(r"元\d+单", stem[len(total_prefix) :]):
        return "天猫超市总表"

    if starts_with_date_prefix(
        stem,
        [
            f"奥克斯索隆{month}🈷️{day}",
            f"奥克斯索隆{month}月{day}",
            f"奥克斯索隆{month}🈷{day}",
        ],
    ):
        return "奥克斯索隆"
    return None


def file_mtime_matches_date(path: Path, month: int, day: int) -> bool:
    modified = datetime.fromtimestamp(path.stat().st_mtime)
    return modified.month == month and modified.day == day


def is_standard_brush_template_file(path: Path) -> bool:
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        return False
    if path.name.startswith(("~$", ".~")):
        return False
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell_text(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
        return detect_source_type(headers) == "标准刷单模板"
    except Exception:
        return False


def copy_standard_wechat_templates(month: int, day: int, print_skipped: bool = False) -> list[Path]:
    copied: list[Path] = []
    for path in sorted(WECHAT_FILE_DIR.rglob("*.xlsx")):
        if path.name.startswith(("~$", ".~")):
            continue
        if not file_mtime_matches_date(path, month, day):
            if print_skipped:
                print(f"跳过微信文件（日期不匹配）：{path}")
            continue
        if not is_standard_brush_template_file(path):
            continue
        target = unique_copy_target(WECHAT_TARGET_DIR, path.name)
        shutil.copy2(path, target)
        copied.append(target)
        print(f"已复制微信标准模板：{path} -> {target}")
    return copied


def has_xlsx_files(folder: Path) -> bool:
    return folder.exists() and any(
        path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
        for path in folder.iterdir()
    )


def unique_copy_target(target_dir: Path, filename: str) -> Path:
    target = target_dir / filename
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    candidate = target_dir / f"{stem}-{stamp}{suffix}"
    counter = 1
    while candidate.exists():
        candidate = target_dir / f"{stem}-{stamp}-{counter}{suffix}"
        counter += 1
    return candidate
def copy_wechat_source_files(month: int, day: int, print_skipped: bool = False) -> list[Path]:
    if not WECHAT_FILE_DIR.exists():
        raise FileNotFoundError(f"找不到微信文件目录：{WECHAT_FILE_DIR}")

    WECHAT_TARGET_DIR.mkdir(parents=True, exist_ok=True)
    latest_by_label: dict[str, Path] = {}
    standard_templates = copy_standard_wechat_templates(month, day, print_skipped=print_skipped)
    copied: list[Path] = []
    skipped = 0
    for path in sorted(WECHAT_FILE_DIR.rglob("*.xlsx")):
        if path.name.startswith(("~$", ".~")):
            continue
        label = matched_wechat_source_label(path, month, day)
        if label:
            previous = latest_by_label.get(label)
            if previous is None or path.stat().st_mtime > previous.stat().st_mtime:
                latest_by_label[label] = path
        else:
            skipped += 1
            if print_skipped:
                print(f"跳过微信文件：{path}")

    for label in ("天猫超市总表", "奥克斯索隆"):
        path = latest_by_label.get(label)
        if not path:
            print(f"未找到微信源表：{label} {month}月{day}日")
            continue
        target = unique_copy_target(WECHAT_TARGET_DIR, path.name)
        shutil.copy2(path, target)
        copied.append(target)
        print(f"已复制微信源表（{label} 最新）：{path} -> {target}")

    copied = standard_templates + copied
    if not copied:
        raise FileNotFoundError(
            f"微信文件目录里没有找到 {month}月{day}日源表："
            f"标准刷单模板，或 【总表】天猫超市{month}.{day}账xxxx元xx单.xlsx，或 奥克斯索隆{month}🈷️{day}.xlsx"
        )
    if not print_skipped:
        print(f"微信文件扫描完成：复制 {len(copied)} 个，跳过 {skipped} 个")
    return copied


def source_files() -> list[Path]:
    if not SOURCE_DIR.exists():
        raise FileNotFoundError(f"找不到今日刷单表格文件夹：{SOURCE_DIR}")
    files = []
    for path in sorted(SOURCE_DIR.glob("*.xlsx")):
        if path.name.startswith(("~$", ".~")):
            continue
        files.append(path)
    if not files:
        raise FileNotFoundError(f"今日刷单表格文件夹里没有 .xlsx 文件：{SOURCE_DIR}")
    return files


def col_to_num(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        raise ValueError(f"单元格坐标异常：{cell_ref}")
    number = 0
    for char in match.group(1):
        number = number * 26 + ord(char) - 64
    return number


def read_shared_strings(zip_file: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    return ["".join(text.text or "" for text in item.findall(".//a:t", XML_NS)) for item in root.findall("a:si", XML_NS)]


def xml_cell_text(cell: ET.Element | None, shared_strings: list[str]) -> str:
    if cell is None:
        return ""
    cell_type = cell.get("t")
    if cell_type == "s":
        value = cell.find("a:v", XML_NS)
        if value is None or not value.text:
            return ""
        return shared_strings[int(value.text)]
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", XML_NS)).strip()
    value = cell.find("a:v", XML_NS)
    return value.text.strip() if value is not None and value.text is not None else ""


def xml_row_cells(row: ET.Element) -> dict[int, ET.Element]:
    return {col_to_num(cell.get("r", "")): cell for cell in row.findall("a:c", XML_NS)}


def read_product_names() -> dict[str, str]:
    if not PRODUCT_FILE.exists():
        raise FileNotFoundError(f"找不到聚水潭商品资料：{PRODUCT_FILE}")

    product_names: dict[str, str] = {}
    conflicts: dict[str, set[str]] = {}
    with ZipFile(PRODUCT_FILE) as zip_file:
        shared_strings = read_shared_strings(zip_file)
        root = ET.fromstring(zip_file.read("xl/worksheets/sheet1.xml"))
        rows = root.findall("a:sheetData/a:row", XML_NS)
        if not rows:
            raise RuntimeError(f"{PRODUCT_FILE.name} 没有可读取的数据。")
        headers = {
            xml_cell_text(cell, shared_strings): column
            for column, cell in xml_row_cells(rows[0]).items()
            if xml_cell_text(cell, shared_strings)
        }
        if "商品编码" not in headers or "商品名称" not in headers:
            raise RuntimeError(f"{PRODUCT_FILE.name} 找不到【商品编码】或【商品名称】字段。")

        code_col = headers["商品编码"]
        name_col = headers["商品名称"]
        for row in rows[1:]:
            cells = xml_row_cells(row)
            code = xml_cell_text(cells.get(code_col), shared_strings)
            name = xml_cell_text(cells.get(name_col), shared_strings)
            if not code or not name:
                continue
            if code in product_names and product_names[code] != name:
                conflicts.setdefault(code, {product_names[code]}).add(name)
                continue
            product_names[code] = name

    if conflicts:
        print(f"提示：聚水潭商品资料里有 {len(conflicts)} 个商品编码对应多个名称，已保留首次出现的名称。")
    return product_names


def read_brush_product_codes() -> dict[str, str]:
    if not BRUSH_PRODUCT_FILE.exists():
        raise FileNotFoundError(f"找不到今日刷单产品表：{BRUSH_PRODUCT_FILE}")

    wb = load_workbook(BRUSH_PRODUCT_FILE, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = {cell_text(ws.cell(1, column).value): column for column in range(1, ws.max_column + 1)}
    if "刷手" not in headers or "商品编码" not in headers:
        raise RuntimeError(f"{BRUSH_PRODUCT_FILE.name} 找不到【刷手】或【商品编码】字段。")

    code_by_brusher: dict[str, str] = {}
    empty_brushers: list[str] = []
    for row in range(2, ws.max_row + 1):
        brusher = cell_text(ws.cell(row, headers["刷手"]).value)
        product_code = cell_text(ws.cell(row, headers["商品编码"]).value)
        if not brusher:
            continue
        code_by_brusher[brusher] = product_code
        if not product_code:
            empty_brushers.append(brusher)

    if empty_brushers:
        print(f"提示：今日刷单产品表里这些刷手未填写商品编码：{', '.join(empty_brushers)}")
    return code_by_brusher


def product_info_for_brusher(
    brusher: str,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> tuple[str, str]:
    product_code = brush_product_codes.get(brusher, "")
    if not product_code:
        return "", ""
    return product_code, product_names.get(product_code, "")


def detect_source_type(headers: list[str]) -> str | None:
    normalized = [header.strip() for header in headers]
    header_set = set(normalized)
    if {"日期", "刷手", "订单编号", "订单金额", "佣金金额"}.issubset(header_set):
        return "标准刷单模板"
    if {"日期", "金额", "旺旺账号", "单号"}.issubset(header_set):
        return "唐杨格式"
    if {"日期", "金额", "单号"}.issubset(header_set) and ("佣金" in header_set or "拥金" in header_set):
        return "小蝴蝶格式"
    if normalized[:3] == ["日期", "金额", "单号"]:
        return "总表三列格式"
    return None


def latest_date_start(ws) -> tuple[int, int, int, str]:
    markers: list[tuple[int, int, int, str]] = []
    for row in range(2, ws.max_row + 1):
        marker = parse_date_marker(ws.cell(row, 1).value)
        if marker:
            month, day, date_text = marker
            markers.append((row, month, day, date_text))
    if not markers:
        raise RuntimeError(f"{ws.title} 没有找到类似“4月24”的日期。")
    return markers[-1]


def read_tangyang_batch(
    path: Path,
    ws,
    source_type: str,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> SourceBatch:
    start_row, month, day, date_text = latest_date_start(ws)
    records: list[SourceRecord] = []
    source_mtime = path.stat().st_mtime
    product_code, product_name = product_info_for_brusher(TANGYANG_BRUSHER, product_names, brush_product_codes)
    for row in range(start_row, ws.max_row + 1):
        amount = ws.cell(row, 2).value
        order_no = cell_text(ws.cell(row, 4).value)
        if amount in (None, "") or not order_no:
            continue
        records.append(
            SourceRecord(
                order_no=order_no,
                amount=to_decimal(amount),
                commission=TANGYANG_COMMISSION,
                brusher=TANGYANG_BRUSHER,
                product_code=product_code,
                product_name=product_name,
                source_file=path.name,
                source_mtime=source_mtime,
                source_row=row,
            )
        )
    if not records:
        raise RuntimeError(f"{path.name} 的 {date_text} 下没有找到“金额 + 单号”的有效记录。")
    return SourceBatch(path, source_type, ws.title, month, day, date_text, records)


def read_xiaohudie_batch(
    path: Path,
    ws,
    source_type: str,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> SourceBatch:
    start_row, month, day, date_text = latest_date_start(ws)
    records: list[SourceRecord] = []
    source_mtime = path.stat().st_mtime
    product_code, product_name = product_info_for_brusher(XIAOHUDIE_BRUSHER, product_names, brush_product_codes)
    for row in range(start_row, ws.max_row + 1):
        amount = ws.cell(row, 2).value
        order_no = cell_text(ws.cell(row, 3).value)
        commission = ws.cell(row, 4).value
        # 小蝴蝶表尾会有合计行；单号为空的行一律跳过。
        if amount in (None, "") or not order_no or commission in (None, ""):
            continue
        records.append(
            SourceRecord(
                order_no=order_no,
                amount=to_decimal(amount),
                commission=to_decimal(commission),
                brusher=XIAOHUDIE_BRUSHER,
                product_code=product_code,
                product_name=product_name,
                source_file=path.name,
                source_mtime=source_mtime,
                source_row=row,
            )
        )
    if not records:
        raise RuntimeError(f"{path.name} 的 {date_text} 下没有找到“金额 + 单号 + 佣金”的有效记录。")
    return SourceBatch(path, source_type, ws.title, month, day, date_text, records)


def parse_total_source_filename(path: Path) -> tuple[Decimal, int] | None:
    match = re.match(r"^【总表】天猫超市\d{1,2}\.\d{1,2}账([\d.]+)元(\d+)单", path.name)
    if not match:
        return None
    return Decimal(match.group(1)), int(match.group(2))


def read_total_table_batch(
    path: Path,
    ws,
    source_type: str,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> SourceBatch:
    start_row, month, day, date_text = latest_date_start(ws)
    raw_rows: list[tuple[Decimal, str]] = []
    for row in range(start_row, ws.max_row + 1):
        amount = ws.cell(row, 2).value
        order_no = cell_text(ws.cell(row, 3).value)
        if amount in (None, "") or not order_no:
            continue
        raw_rows.append((to_decimal(amount), order_no))
    if not raw_rows:
        raise RuntimeError(f"{path.name} 的 {date_text} 下没有找到“金额 + 单号”的有效记录。")

    filename_total = parse_total_source_filename(path)
    if filename_total:
        total_amount, expected_count = filename_total
        if expected_count != len(raw_rows):
            print(f"提示：{path.name} 文件名写的是 {expected_count} 单，表内最新日期识别到 {len(raw_rows)} 单。")
        order_total = sum((amount for amount, _ in raw_rows), Decimal("0"))
        commission_total = total_amount - order_total
        commission = (commission_total / Decimal(len(raw_rows))).quantize(Decimal("0.01"))
    else:
        commission = Decimal("0")
        print(f"提示：{path.name} 无法从文件名解析总额，三列表佣金按 0 处理。")

    product_code, product_name = product_info_for_brusher(TOTAL_SOURCE_BRUSHER, product_names, brush_product_codes)
    source_mtime = path.stat().st_mtime
    records = [
        SourceRecord(
            order_no=order_no,
            amount=amount,
            commission=commission,
            brusher=TOTAL_SOURCE_BRUSHER,
            product_code=product_code,
            product_name=product_name,
            source_file=path.name,
            source_mtime=source_mtime,
        )
        for amount, order_no in raw_rows
    ]
    return SourceBatch(path, source_type, ws.title, month, day, date_text, records)


def read_standard_brush_template_batch(
    path: Path,
    ws,
    source_type: str,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> list[SourceBatch]:
    headers = {cell_text(ws.cell(1, column).value): column for column in range(1, ws.max_column + 1)}
    required_headers = ("日期", "刷手", "订单编号", "订单金额", "佣金金额")
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise RuntimeError(f"{path.name} 缺少标准模板字段：{', '.join(missing)}")

    source_mtime = path.stat().st_mtime
    grouped_records: dict[tuple[int, int, str], list[SourceRecord]] = {}
    skipped_rows = 0
    for row in range(2, ws.max_row + 1):
        row_date = ws.cell(row, headers["日期"]).value
        brusher = cell_text(ws.cell(row, headers["刷手"]).value)
        order_no = cell_text(ws.cell(row, headers["订单编号"]).value)
        amount = ws.cell(row, headers["订单金额"]).value
        commission = ws.cell(row, headers["佣金金额"]).value

        if brusher in ("",) or order_no in ("",) or amount in (None, "") or commission in (None, ""):
            skipped_rows += 1
            print(f"跳过标准模板行：{path.name} 第 {row} 行，缺少 刷手/订单编号/订单金额/佣金金额 中的必填值")
            continue

        try:
            month, day, date_text = parse_standard_row_date(row_date)
            amount_decimal = to_decimal(amount)
            commission_decimal = to_decimal(commission)
        except Exception as exc:
            skipped_rows += 1
            print(f"跳过标准模板行：{path.name} 第 {row} 行，原因：{exc}")
            continue

        product_code = cell_text(ws.cell(row, headers["商品编码"]).value) if "商品编码" in headers else ""
        if not product_code:
            product_code, product_name = product_info_for_brusher(brusher, product_names, brush_product_codes)
        else:
            product_name = product_names.get(product_code, "")

        grouped_records.setdefault((month, day, date_text), []).append(
            SourceRecord(
                order_no=order_no,
                amount=amount_decimal,
                commission=commission_decimal,
                brusher=brusher,
                product_code=product_code,
                product_name=product_name,
                source_file=path.name,
                source_mtime=source_mtime,
                source_row=row,
            )
        )

    if not grouped_records:
        raise RuntimeError(f"{path.name} 没有读到任何有效标准模板记录。")
    if skipped_rows:
        print(f"提示：{path.name} 标准模板共跳过 {skipped_rows} 行缺字段记录。")

    batches = [
        SourceBatch(path, source_type, ws.title, month, day, date_text, records)
        for (month, day, date_text), records in sorted(grouped_records.items(), key=lambda item: (item[0][0], item[0][1], item[0][2]))
    ]
    return batches


def read_source_batch(
    path: Path,
    product_names: dict[str, str],
    brush_product_codes: dict[str, str],
) -> list[SourceBatch]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell_text(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    source_type = detect_source_type(headers)
    if source_type == "标准刷单模板":
        return read_standard_brush_template_batch(path, ws, source_type, product_names, brush_product_codes)
    if source_type == "唐杨格式":
        return [read_tangyang_batch(path, ws, source_type, product_names, brush_product_codes)]
    if source_type == "小蝴蝶格式":
        return [read_xiaohudie_batch(path, ws, source_type, product_names, brush_product_codes)]
    if source_type == "总表三列格式":
        return [read_total_table_batch(path, ws, source_type, product_names, brush_product_codes)]
    raise RuntimeError(f"{path.name} 无法识别格式，表头为：{headers}")


def read_all_source_batches() -> list[SourceBatch]:
    product_names = read_product_names()
    brush_product_codes = read_brush_product_codes()
    batches = []
    errors = []
    for path in source_files():
        try:
            batches.extend(read_source_batch(path, product_names, brush_product_codes))
        except Exception as exc:
            errors.append(f"{path.name}：{exc}")
            print(f"跳过不可解析源表：{path.name}，原因：{exc}")
    if not batches and errors:
        raise RuntimeError("没有可用源表；" + "；".join(errors))
    return dedupe_records_by_latest_source_mtime(batches)


def dedupe_records_by_latest_source_mtime(batches: list[SourceBatch]) -> list[SourceBatch]:
    latest_by_order: dict[str, SourceRecord] = {}
    for batch in batches:
        for record in batch.records:
            previous = latest_by_order.get(record.order_no)
            if previous is None:
                latest_by_order[record.order_no] = record
                continue
            if record.source_mtime >= previous.source_mtime:
                kept, dropped = record, previous
            else:
                kept, dropped = previous, record
            latest_by_order[record.order_no] = kept
            print(
                f"重复订单去重：{record.order_no} | 保留：{kept.source_file} "
                f"| 跳过：{dropped.source_file}"
            )

    deduped_batches: list[SourceBatch] = []
    for batch in batches:
        kept_records = [record for record in batch.records if latest_by_order.get(record.order_no) == record]
        if not kept_records:
            continue
        deduped_batches.append(
            SourceBatch(
                source_file=batch.source_file,
                source_type=batch.source_type,
                sheet_name=batch.sheet_name,
                month=batch.month,
                day=batch.day,
                date_text=batch.date_text,
                records=kept_records,
            )
        )
    return deduped_batches


def target_file_for_month(month: int) -> Path:
    return BASE_DIR / f"天猫超市{month}月刷单登记明细.xlsx"


def create_empty_target_from_scratch(target: Path, month: int) -> None:
    unified_name = f"天猫超市{month}月刷单登记明细"
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = unified_name
    ws["B1"] = unified_name
    ws["B1"].font = Font(bold=True, size=14)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, header in enumerate(TARGET_HEADERS, start=1):
        cell = ws.cell(row=2, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [8, 12, 14, 24, 12, 12, 22, 22, 18, 36, 18]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + column)].width = width
    ws.freeze_panes = "A3"
    wb.save(target)

    print(f"未找到任何历史月份登记表，已按固定字段创建新空表：{target}")


def create_empty_target_from_template(target: Path, month: int) -> str:
    templates = sorted(
        [path for path in BASE_DIR.glob("天猫超市*月刷单登记明细.xlsx") if path != target],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not templates:
        create_empty_target_from_scratch(target, month)
        return "scratch"

    template = templates[0]
    unified_name = f"天猫超市{month}月刷单登记明细"
    target.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_target = Path(tmp_dir) / target.name
        with ZipFile(template, "r") as zin, ZipFile(tmp_target, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r'<sheet name="[^"]+"',
                        f'<sheet name="{escape(unified_name)}"',
                        text,
                        count=1,
                    )
                    data = text.encode("utf-8")
                elif item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode("utf-8")
                    text = re.sub(r'<dimension ref="[^"]+"/>', '<dimension ref="A1:K2"/>', text, count=1)
                    text = re.sub(
                        r'<c r="B1"[^>]*>.*?</c>',
                        f'<c r="B1" s="4" t="inlineStr"><is><t>{escape(unified_name)}</t></is></c>',
                        text,
                        count=1,
                        flags=re.S,
                    )
                    text = re.sub(r'<row r="([3-9]|\d{2,})"[^>]*>.*?</row>', "", text, flags=re.S)
                    data = text.encode("utf-8")
                zout.writestr(item, data)
        shutil.copy2(tmp_target, target)

    print(f"未找到目标登记表，已根据模板创建空表：{target}")
    print(f"模板来源：{template}")
    return "template"


def inspect_target(path: Path) -> tuple[str, int, int, set[str]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    max_row = 2
    max_seq = 0
    existing_orders: set[str] = set()
    for row in range(3, ws.max_row + 1):
        seq = ws.cell(row, 1).value
        order_no = ws.cell(row, 4).value
        amount = ws.cell(row, 5).value
        if seq not in (None, "") or order_no not in (None, "") or amount not in (None, ""):
            max_row = row
        if isinstance(seq, (int, float)):
            max_seq = max(max_seq, int(seq))
        if order_no not in (None, ""):
            existing_orders.add(str(order_no).strip())

    return sheet_name, max_row, max_seq, existing_orders


def workbook_has_image_structure(path: Path) -> bool:
    with ZipFile(path) as zip_file:
        names = set(zip_file.namelist())
    return any(
        name.startswith(("xl/media/", "xl/drawings/")) or name == "xl/cellimages.xml"
        for name in names
    )


def print_records_summary(title: str, records: list[SourceRecord], *, show_brusher_totals: bool = False) -> None:
    order_total = sum((record.amount for record in records), Decimal("0"))
    commission_total = sum((record.commission for record in records), Decimal("0"))
    grand_total = order_total + commission_total

    print("")
    print(title)
    if not records:
        print("无")
    else:
        for index, record in enumerate(records, start=1):
            print(
                f"{index}. 刷手：{record.brusher} | 单号：{record.order_no} "
                f"| 订单金额：{money_text(record.amount)} | 佣金：{money_text(record.commission)} "
                f"| 小计：{money_text(record.amount + record.commission)} "
                f"| 商品编码：{record.product_code} | 名称：{record.product_name or '未匹配'} "
                f"| 来源：{record.source_file}"
            )
    if show_brusher_totals and records:
        grouped: dict[str, list[SourceRecord]] = {}
        for record in records:
            grouped.setdefault(record.brusher, []).append(record)
        print("按刷手统计：")
        for brusher in sorted(grouped):
            brusher_records = grouped[brusher]
            brusher_order_total = sum((record.amount for record in brusher_records), Decimal("0"))
            brusher_commission_total = sum((record.commission for record in brusher_records), Decimal("0"))
            brusher_grand_total = brusher_order_total + brusher_commission_total
            print(
                f"【{brusher}】订单金额合计：{money_text(brusher_order_total)} "
                f"佣金合计：{money_text(brusher_commission_total)} "
                f"订单金额+佣金合计：{money_text(brusher_grand_total)}"
            )
    print(f"订单金额合计：{money_text(order_total)}")
    print(f"佣金合计：{money_text(commission_total)}")
    print(f"订单金额+佣金合计：{money_text(grand_total)}")


def print_transfer_summary(records: list[SourceRecord], title: str = "本次刷手转账金额：") -> None:
    grouped: dict[str, list[SourceRecord]] = {}
    for record in records:
        grouped.setdefault(record.brusher, []).append(record)

    print("")
    print(title)
    if not grouped:
        print("无")
        return

    for brusher in sorted(grouped):
        brusher_records = grouped[brusher]
        principal = sum((record.amount for record in brusher_records), Decimal("0"))
        commission = sum((record.commission for record in brusher_records), Decimal("0"))
        total = principal + commission
        print(
            f"【{brusher}】今日刷单本金：{money_text(principal)} "
            f"刷单佣金：{money_text(commission)} 总金额：{money_text(total)}"
        )


def patch_workbook(
    target: Path,
    month: int,
    records_by_day: list[tuple[int, list[SourceRecord]]],
    append_start_row: int,
    first_seq: int,
) -> None:
    unified_name = f"天猫超市{month}月刷单登记明细"
    total_records = sum(len(records) for _, records in records_by_day)
    final_row = append_start_row + total_records - 1

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"{target.stem}-追加前备份-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(target, backup)

    date_style_id = detect_existing_date_style_id(target)

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_target = Path(tmp_dir) / target.name
        with ZipFile(target, "r") as zin, ZipFile(tmp_target, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r'<sheet name="[^"]+"',
                        f'<sheet name="{escape(unified_name)}"',
                        text,
                        count=1,
                    )
                    data = text.encode("utf-8")
                elif item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode("utf-8")
                    text = update_sheet_dimension(text, final_row)
                    text = re.sub(
                        r'<c r="B1"[^>]*>.*?</c>',
                        f'<c r="B1" s="4" t="inlineStr"><is><t>{escape(unified_name)}</t></is></c>',
                        text,
                        count=1,
                        flags=re.S,
                    )

                    rows = []
                    seq = first_seq
                    row_no = append_start_row
                    for day, records in records_by_day:
                        serial = excel_serial(datetime(datetime.now().year, month, day))
                        for record in records:
                            rows.append(
                                f'<row r="{row_no}" customHeight="1" spans="1:11">'
                                f'<c r="A{row_no}" s="2"><v>{seq}</v></c>'
                                f'<c r="B{row_no}" s="2" t="inlineStr"><is><t>{escape(record.brusher)}</t></is></c>'
                                f'<c r="C{row_no}" s="{date_style_id}"><v>{serial}</v></c>'
                                f'<c r="D{row_no}" s="17" t="inlineStr"><is><t>{escape(record.order_no)}</t></is></c>'
                                f'<c r="E{row_no}" s="2"><v>{money_text(record.amount)}</v></c>'
                                f'<c r="F{row_no}" s="2"><v>{money_text(record.commission)}</v></c>'
                                f'<c r="G{row_no}" s="2"/>'
                                f'<c r="H{row_no}" s="12"/>'
                                f'<c r="I{row_no}" s="2" t="inlineStr"><is><t>{escape(record.product_code)}</t></is></c>'
                                f'<c r="J{row_no}" s="2" t="inlineStr"><is><t>{escape(record.product_name)}</t></is></c>'
                                f"</row>"
                            )
                            seq += 1
                            row_no += 1

                    if "</sheetData>" not in text:
                        raise RuntimeError("目标表结构异常：找不到 </sheetData>。")
                    text = replace_rows(text, append_start_row, final_row, "".join(rows))
                    data = text.encode("utf-8")
                zout.writestr(item, data)
        shutil.copy2(tmp_target, target)

    print(f"已备份：{backup}")


def detect_existing_date_style_id(target: Path) -> int:
    workbook = load_workbook(target, read_only=False, data_only=False)
    try:
        sheet = workbook.active
        for row_no in range(3, min(sheet.max_row, 200) + 1):
            cell = sheet.cell(row=row_no, column=3)
            if cell.value in (None, ""):
                continue
            if cell.style_id:
                return int(cell.style_id)
    finally:
        workbook.close()
    return 8


def update_sheet_dimension(sheet_xml: str, final_row: int) -> str:
    def replacement(match: re.Match[str]) -> str:
        ref = match.group(1)
        row_matches = re.findall(r"[A-Z]+(\d+)", ref)
        current_max = max((int(row) for row in row_matches), default=final_row)
        return f'<dimension ref="A1:K{max(current_max, final_row)}"/>'

    return re.sub(r'<dimension ref="([^"]+)"/>', replacement, sheet_xml, count=1)


def replace_rows(sheet_xml: str, start_row: int, end_row: int, new_rows_xml: str) -> str:
    row_pattern = re.compile(r'<row r="(\d+)"(?=[\s>])[^>]*>.*?</row>', re.S)

    def remove_target_rows(match: re.Match[str]) -> str:
        row_no = int(match.group(1))
        if start_row <= row_no <= end_row:
            return ""
        return match.group(0)

    sheet_xml = row_pattern.sub(remove_target_rows, sheet_xml)

    insert_pos: int | None = None
    for match in row_pattern.finditer(sheet_xml):
        if int(match.group(1)) > end_row:
            insert_pos = match.start()
            break

    if insert_pos is None:
        return sheet_xml.replace("</sheetData>", new_rows_xml + "</sheetData>", 1)
    return sheet_xml[:insert_pos] + new_rows_xml + sheet_xml[insert_pos:]


def append_plain_workbook(
    target: Path,
    month: int,
    records_by_day: list[tuple[int, list[SourceRecord]]],
    append_start_row: int,
    first_seq: int,
) -> None:
    unified_name = f"天猫超市{month}月刷单登记明细"
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"{target.stem}-追加前备份-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(target, backup)

    wb = load_workbook(target)
    ws = wb[wb.sheetnames[0]]
    ws.title = unified_name
    ws["B1"] = unified_name

    seq = first_seq
    row_no = append_start_row
    for day, records in records_by_day:
        serial = excel_serial(datetime(datetime.now().year, month, day))
        for record in records:
            values = [
                seq,
                record.brusher,
                serial,
                record.order_no,
                float(record.amount),
                float(record.commission),
                None,
                None,
                record.product_code,
                record.product_name,
                None,
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=column, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_no, column=3).number_format = "m月d日"
            seq += 1
            row_no += 1

    wb.save(target)
    print(f"已备份：{backup}")


def group_batches_by_month(batches: list[SourceBatch]) -> dict[int, list[SourceBatch]]:
    grouped: dict[int, list[SourceBatch]] = {}
    for batch in batches:
        grouped.setdefault(batch.month, []).append(batch)
    return grouped


def clear_source_dir() -> None:
    removed_count = 0
    for path in SOURCE_DIR.iterdir():
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()
        removed_count += 1
    print(f"已清空今日刷单表格文件夹：{SOURCE_DIR}（删除 {removed_count} 项）")


def configure_paths(
    work_dir: Path,
    source_dir: Path | None,
    product_file: Path,
    brush_product_file: Path,
    wechat_file_dir: Path,
    wechat_target_dir: Path,
) -> None:
    global BASE_DIR, SOURCE_DIR, BACKUP_DIR, PRODUCT_FILE, BRUSH_PRODUCT_FILE, WECHAT_FILE_DIR, WECHAT_TARGET_DIR
    BASE_DIR = work_dir.expanduser().resolve()
    SOURCE_DIR = source_dir.expanduser().resolve() if source_dir else get_path("brush_orders_dir")
    BACKUP_DIR = get_path("backup_dir")
    PRODUCT_FILE = product_file.expanduser().resolve()
    BRUSH_PRODUCT_FILE = brush_product_file.expanduser().resolve()
    WECHAT_FILE_DIR = wechat_file_dir.expanduser().resolve()
    WECHAT_TARGET_DIR = wechat_target_dir.expanduser().resolve()


def write_latest_brush_orders(orders: list[str]) -> None:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "date": date.today().isoformat(),
        "orders": orders,
    }
    LATEST_BRUSH_ORDERS_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def python_has_modules(python_path: Path, modules: tuple[str, ...]) -> bool:
    if not python_path.exists():
        return False
    command = [
        str(python_path),
        "-c",
        "import importlib.util, sys; "
        "mods=sys.argv[1:]; "
        "missing=[m for m in mods if importlib.util.find_spec(m) is None]; "
        "raise SystemExit(1 if missing else 0)",
        *modules,
    ]
    result = subprocess.run(command, capture_output=True, text=True)
    return result.returncode == 0


def python_candidates() -> list[Path]:
    candidates = [
        BUNDLED_PYTHON,
        Path(sys.executable),
        Path("/usr/bin/python3"),
        Path("/usr/local/bin/python3"),
        Path("/opt/homebrew/bin/python3"),
    ]
    unique: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def choose_python_for_jst_tagging() -> str:
    required_modules: tuple[str, ...] = ()
    for candidate in python_candidates():
        if python_has_modules(candidate, required_modules):
            return str(candidate)
    return sys.executable


def trigger_jst_tagging() -> None:
    if not TAG_JST_TASK_PATH.exists():
        print(f"未找到聚水潭打标脚本：{TAG_JST_TASK_PATH}")
        return
    print("\n开始自动执行聚水潭刷单订单打标...")
    python_bin = choose_python_for_jst_tagging()
    command = [python_bin, str(TAG_JST_TASK_PATH), "--input", str(LATEST_BRUSH_ORDERS_PATH)]
    result = subprocess.run(command, text=True, capture_output=True)
    if result.stdout:
        print(result.stdout, end="" if result.stdout.endswith("\n") else "\n")
    if result.stderr:
        print(result.stderr, end="" if result.stderr.endswith("\n") else "\n", file=sys.stderr)
    if result.returncode == 0:
        print("聚水潭刷单订单打标完成。")
        return
    print(f"聚水潭刷单订单打标失败，返回码：{result.returncode}")
    print("刷单登记已完成；如需重试，可单独运行：")
    print(f"{python_bin} {TAG_JST_TASK_PATH}")


def run(
    dry_run: bool = False,
    *,
    auto_fetch_wechat: bool = True,
    wechat_month_day: tuple[int, int] | None = None,
    print_skipped_wechat: bool = False,
) -> dict[str, object]:
    global SOURCE_DIR
    appended_orders: list[str] = []
    if auto_fetch_wechat and not has_xlsx_files(SOURCE_DIR):
        month, day = wechat_month_day or (date.today().month, date.today().day)
        print(f"今日刷单表格文件夹暂无 Excel，开始从微信文件目录自动查找：{month}月{day}日")
        copied = copy_wechat_source_files(month, day, print_skipped=print_skipped_wechat)
        SOURCE_DIR = WECHAT_TARGET_DIR
        print(f"已切换今日刷单源表文件夹：{SOURCE_DIR}（本次复制 {len(copied)} 个）")

    batches = read_all_source_batches()
    print(f"今日刷单表格文件夹：{SOURCE_DIR}")
    print(f"识别到有效源表：{len(batches)} 个")
    for batch in batches:
        missing_name = " | 名称未匹配" if any(not record.product_name for record in batch.records) else ""
        product_codes = sorted({record.product_code for record in batch.records if record.product_code})
        product_code_text = "、".join(product_codes) if product_codes else "未填写"
        print(
            f"- {batch.source_file.name} | {batch.source_type} | sheet：{batch.sheet_name} "
            f"| 最新日期：{batch.date_text} | 记录数：{len(batch.records)} "
            f"| 商品编码：{product_code_text}{missing_name}"
        )
        print_records_summary(f"{batch.source_file.name} 最新日期源数据明细：", batch.records)

    for month, month_batches in group_batches_by_month(batches).items():
        target = target_file_for_month(month)
        target_creation_mode = "existing"
        if not target.exists():
            if dry_run:
                print(f"预览模式：目标登记表不存在，正式运行时会自动创建空表：{target}")
                max_row = 2
                max_seq = 0
                existing_orders: set[str] = set()
            else:
                target_creation_mode = create_empty_target_from_template(target, month)
                _, max_row, max_seq, existing_orders = inspect_target(target)
        else:
            _, max_row, max_seq, existing_orders = inspect_target(target)
        records_by_day: list[tuple[int, list[SourceRecord]]] = []
        all_new_records: list[SourceRecord] = []
        seen_orders = set(existing_orders)
        total_source_records = 0
        for batch in month_batches:
            new_records = []
            for record in batch.records:
                if record.order_no in seen_orders:
                    continue
                seen_orders.add(record.order_no)
                new_records.append(record)
            records_by_day.append((batch.day, new_records))
            all_new_records.extend(new_records)
            total_source_records += len(batch.records)

        print("")
        print(f"目标表：{target.name}")
        print(f"源记录数：{total_source_records}")
        print(f"可追加新记录：{len(all_new_records)}")
        print_records_summary("本次实际追加数据明细：", all_new_records, show_brusher_totals=True)

        if dry_run:
            print_transfer_summary(all_new_records, title="预览刷手转账金额：")
            print("预览模式：未写入登记表。")
            continue
        if not all_new_records:
            print("没有追加：源表单号已存在，避免重复写入。")
            continue

        if target_creation_mode == "scratch" and not workbook_has_image_structure(target):
            append_plain_workbook(
                target=target,
                month=month,
                records_by_day=records_by_day,
                append_start_row=max_row + 1,
                first_seq=max_seq + 1,
            )
        else:
            patch_workbook(
                target=target,
                month=month,
                records_by_day=records_by_day,
                append_start_row=max_row + 1,
                first_seq=max_seq + 1,
            )
        print(f"追加行：{max_row + 1}-{max_row + len(all_new_records)}")
        print(f"序号：{max_seq + 1}-{max_seq + len(all_new_records)}")
        print_transfer_summary(all_new_records)
        print("完成。")
        appended_orders.extend(record.order_no for record in all_new_records)

    if not dry_run:
        write_latest_brush_orders(appended_orders)
        print(f"本次新增订单号已写入：{LATEST_BRUSH_ORDERS_PATH}")
        if appended_orders:
            trigger_jst_tagging()
        clear_source_dir()
    return {
        "appended_orders": appended_orders,
        "appended_count": len(appended_orders),
        "source_dir": str(SOURCE_DIR),
        "latest_brush_orders_path": str(LATEST_BRUSH_ORDERS_PATH),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="自动识别今日刷单表格并追加到登记表")
    parser.add_argument("--work-dir", default=str(DEFAULT_WORK_DIR), help="刷单登记工作目录")
    parser.add_argument("--source-dir", help="今日刷单表格文件夹；默认读取 config/paths.yaml 的 brush_orders_dir")
    parser.add_argument("--product-file", default=str(DEFAULT_PRODUCT_FILE), help="聚水潭商品资料路径")
    parser.add_argument("--brush-product-file", default=str(DEFAULT_BRUSH_PRODUCT_FILE), help="刷手对应商品编码表")
    parser.add_argument("--no-auto-fetch-wechat", action="store_true", help="关闭从微信文件目录自动复制源表")
    parser.add_argument("--wechat-file-dir", default=str(DEFAULT_WECHAT_FILE_DIR), help="微信 file 总目录")
    parser.add_argument("--wechat-target-dir", default=str(DEFAULT_WECHAT_TARGET_DIR), help="微信源表复制目标目录")
    parser.add_argument(
        "--wechat-date",
        type=parse_month_day,
        help="要匹配的微信文件日期，支持 430、4.30、4-30、4月30；默认今天",
    )
    parser.add_argument("--print-skipped-wechat", action="store_true", help="打印微信扫描跳过文件")
    parser.add_argument("--dry-run", action="store_true", help="只预览识别和提取结果，不写入登记表")
    parser.add_argument("date_words", nargs="*", help="自然日期，例如 昨天的、4月29、4.29")
    args = parser.parse_args()
    context = TaskContext("append_brush_orders")
    context.add_input("dry_run", args.dry_run)
    context.add_input("work_dir", args.work_dir)
    context.add_input("source_dir", args.source_dir or str(DEFAULT_WECHAT_TARGET_DIR))
    context.add_input("product_file", args.product_file)
    context.add_input("brush_product_file", args.brush_product_file)
    context.add_input("auto_fetch_wechat", not args.no_auto_fetch_wechat)
    context.add_input("date_words", args.date_words)
    wechat_date = args.wechat_date or (parse_natural_month_day(args.date_words) if args.date_words else None)
    try:
        configure_paths(
            work_dir=Path(args.work_dir),
            source_dir=Path(args.source_dir) if args.source_dir else None,
            product_file=Path(args.product_file),
            brush_product_file=Path(args.brush_product_file),
            wechat_file_dir=Path(args.wechat_file_dir),
            wechat_target_dir=Path(args.wechat_target_dir),
        )
        summary = run(
            dry_run=args.dry_run,
            auto_fetch_wechat=not args.no_auto_fetch_wechat,
            wechat_month_day=wechat_date,
            print_skipped_wechat=args.print_skipped_wechat,
        )
        context.add_output("summary", summary)
        context.add_artifact(summary["latest_brush_orders_path"])
        if summary.get("appended_count"):
            context.add_next_task(
                "tag_jst_brush_orders",
                {"input": summary["latest_brush_orders_path"], "orders": summary["appended_orders"]},
            )
        context_path = context.finish("dry_run_success" if args.dry_run else "success")
        print(f"任务上下文：{context_path}")
    except Exception as exc:
        context.add_error(str(exc))
        context_path = context.finish("failed")
        print(f"任务上下文：{context_path}")
        raise


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"运行失败：{exc}") from exc
