#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import shutil
import sys
import tempfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请使用配套 Python 运行。") from exc

CURRENT_DIR = Path(__file__).resolve().parent
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from clients.ops_cli_client import run_ops_json  # noqa: E402
from core.config_loader import get_path  # noqa: E402


LOG_DIR = get_path("logs_dir")
RUNTIME_DIR = get_path("runtime_dir")
FAILED_EXPORT_TEMPLATE = "jst_brush_reimburse_failed_{date}.xlsx"
LOG_TEMPLATE = "jst_brush_reimburse_workorder_{date}.log"
DEFAULT_WORKBOOK_DIR = get_path("brush_register_dir")
DEFAULT_WORKBOOK_TEMPLATE = "天猫超市{month}月刷单登记明细.xlsx"
BACKUP_DIR = get_path("backup_dir")
MARKER_TEXT = "已经提交聚水潭报销"
MARKER_MATCH_TEXTS = (MARKER_TEXT, "上面的费用已经提交剧水潭报销", "上面的费用已经提交聚水潭报销")
XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", XML_NS)


@dataclass(frozen=True)
class BatchOrder:
    row_index: int
    brusher: str
    brush_date: Any
    order_no: str
    order_amount: Decimal
    commission_amount: Decimal
    product_code: str
    product_name: str


@dataclass(frozen=True)
class BatchInfo:
    workbook_path: Path
    start_row: int
    end_row: int
    orders: list[BatchOrder]
    principal_total: Decimal
    payout_total: Decimal


@dataclass
class CandidateResult:
    order: BatchOrder
    o_id: str = ""
    lp_order_no: str = ""
    item_name: str = ""
    has_existing_workorder: bool = False
    existing_detail: dict[str, Any] | None = None
    skip_reason: str = ""


@dataclass
class FailureRecord:
    order_no: str
    reason: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="聚水潭刷单报销工单")
    current_month = datetime.now().month
    parser.add_argument(
        "--input",
        default=str(DEFAULT_WORKBOOK_DIR / DEFAULT_WORKBOOK_TEMPLATE.format(month=current_month)),
        help="刷单登记表路径",
    )
    parser.add_argument("--dry-run", action="store_true", help="只识别批次和查询，不上传、不创建、不回写")
    parser.add_argument("--order-no", help="只在当前批次内处理指定订单号")
    return parser.parse_args()


def setup_logging() -> Path:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    path = LOG_DIR / LOG_TEMPLATE.format(date=datetime.now().strftime("%Y%m%d"))
    logger = logging.getLogger()
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    handler = logging.FileHandler(path, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(handler)
    return path


def to_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def money_text(value: Decimal) -> str:
    text = f"{value.quantize(Decimal('0.01'))}"
    return text.rstrip("0").rstrip(".") if "." in text else text


def cell_text(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def is_marker_text(text: str) -> bool:
    normalized = text.replace(" ", "")
    if any(marker in normalized for marker in MARKER_MATCH_TEXTS):
        return True
    return "已经提交" in normalized and "报销" in normalized and ("聚水潭" in normalized or "剧水潭" in normalized)


def read_current_batch(path: Path) -> BatchInfo:
    if not path.exists():
        raise FileNotFoundError(f"找不到刷单登记表：{path}")
    wb = load_workbook(path)
    ws = wb.active
    header_row = 2
    headers = {cell_text(ws.cell(header_row, col).value): col for col in range(1, ws.max_column + 1)}
    required = ["刷手", "刷单日期", "订单编号", "订单金额", "佣金金额", "商品编码", "名称"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"{path.name} 缺少字段：{', '.join(missing)}")

    last_marker_row = 0
    orders: list[BatchOrder] = []
    first_valid_row = 0
    last_valid_row = 0
    for row in range(header_row + 1, ws.max_row + 1):
        row_text = cell_text(ws.cell(row, 1).value)
        if row_text and is_marker_text(row_text):
            last_marker_row = row
            orders = []
            first_valid_row = 0
            last_valid_row = 0
            continue
        order_no = cell_text(ws.cell(row, headers["订单编号"]).value)
        product_code = cell_text(ws.cell(row, headers["商品编码"]).value)
        order_amount = to_decimal(ws.cell(row, headers["订单金额"]).value)
        commission_amount = to_decimal(ws.cell(row, headers["佣金金额"]).value)
        if row > last_marker_row and order_no and product_code and order_amount is not None and commission_amount is not None:
            if first_valid_row == 0:
                first_valid_row = row
            last_valid_row = row
            orders.append(
                BatchOrder(
                    row_index=row,
                    brusher=cell_text(ws.cell(row, headers["刷手"]).value),
                    brush_date=ws.cell(row, headers["刷单日期"]).value,
                    order_no=order_no,
                    order_amount=order_amount,
                    commission_amount=commission_amount,
                    product_code=product_code,
                    product_name=cell_text(ws.cell(row, headers["名称"]).value),
                )
            )
    if not orders:
        raise RuntimeError("当前没有可报销的未标记批次")
    principal_total = sum((item.order_amount for item in orders), Decimal("0"))
    payout_total = sum((item.commission_amount for item in orders), Decimal("0"))
    return BatchInfo(
        workbook_path=path,
        start_row=first_valid_row,
        end_row=last_valid_row,
        orders=orders,
        principal_total=principal_total,
        payout_total=payout_total,
    )


def ops_reimburse_payload(
    batch: BatchInfo,
    order: BatchOrder,
    *,
    execute: bool = False,
    interactive_recovery: bool = False,
) -> dict[str, Any]:
    command = [
        "--json",
        "jst",
        "order",
        "reimburse",
        "--outer-order-id",
        order.order_no,
        "--principal-total",
        money_text(batch.principal_total),
        "--payout-total",
        money_text(batch.payout_total),
        "--product-code",
        order.product_code,
        "--product-name",
        order.product_name,
        "--workbook-file",
        str(batch.workbook_path),
    ]
    if execute:
        command.append("--execute")
    payload = run_ops_json(command, interactive_recovery=interactive_recovery)
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        raise RuntimeError(f"Ops-Cli 返回结构异常：{payload}")
    return data


def backup_workbook(path: Path) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = BACKUP_DIR / f"{path.stem}_报销工单前备份_{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def _ns(tag: str) -> str:
    return f"{{{XML_NS}}}{tag}"


def _column_letters(column_index: int) -> str:
    letters = []
    current = column_index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _parse_ref(ref: str) -> tuple[int, int]:
    letters = "".join(ch for ch in ref if ch.isalpha())
    digits = "".join(ch for ch in ref if ch.isdigit())
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch.upper()) - 64)
    return col, int(digits)


def _update_dimension(sheet_root: ET.Element, target_row: int, max_col: int = 11) -> None:
    dimension = sheet_root.find(_ns("dimension"))
    if dimension is None:
        return
    current_ref = dimension.attrib.get("ref", f"A1:{_column_letters(max_col)}{target_row}")
    start_ref, _, end_ref = current_ref.partition(":")
    if not end_ref:
        end_ref = start_ref
    end_col, end_row = _parse_ref(end_ref)
    new_end_col = max(end_col, max_col)
    new_end_row = max(end_row, target_row)
    dimension.set("ref", f"{start_ref}:{_column_letters(new_end_col)}{new_end_row}")


def _ensure_marker_style(styles_xml: bytes) -> tuple[bytes, int]:
    root = ET.fromstring(styles_xml)
    cell_xfs = root.find(_ns("cellXfs"))
    if cell_xfs is None:
        raise RuntimeError("styles.xml 结构异常：缺少 cellXfs")

    for index, xf in enumerate(list(cell_xfs)):
        if (
            xf.attrib.get("fontId") == "5"
            and xf.attrib.get("fillId") == "2"
            and xf.attrib.get("borderId") == "1"
        ):
            alignment = xf.find(_ns("alignment"))
            if alignment is not None and alignment.attrib.get("horizontal") == "center" and alignment.attrib.get("vertical") == "center":
                return ET.tostring(root, encoding="utf-8", xml_declaration=True), index

    new_xf = ET.Element(
        _ns("xf"),
        {
            "numFmtId": "0",
            "fontId": "5",
            "fillId": "2",
            "borderId": "1",
            "xfId": "0",
            "applyFont": "1",
            "applyFill": "1",
            "applyBorder": "1",
            "applyAlignment": "1",
        },
    )
    ET.SubElement(new_xf, _ns("alignment"), {"horizontal": "center", "vertical": "center"})
    cell_xfs.append(new_xf)
    cell_xfs.set("count", str(len(list(cell_xfs))))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), len(list(cell_xfs)) - 1


def _build_marker_row(target_row: int, style_index: int) -> ET.Element:
    row = ET.Element(_ns("row"), {"r": str(target_row), "customHeight": "1", "spans": "1:11"})
    cell = ET.SubElement(row, _ns("c"), {"r": f"A{target_row}", "s": str(style_index), "t": "inlineStr"})
    inline = ET.SubElement(cell, _ns("is"))
    ET.SubElement(inline, _ns("t")).text = MARKER_TEXT
    return row


def _patch_marker_sheet(sheet_xml: bytes, target_row: int, style_index: int) -> bytes:
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_ns("sheetData"))
    if sheet_data is None:
        raise RuntimeError("sheet1.xml 结构异常：缺少 sheetData")

    existing_rows = {int(row.attrib["r"]): row for row in sheet_data.findall(_ns("row")) if row.attrib.get("r")}
    if target_row in existing_rows:
        row = existing_rows[target_row]
        row.clear()
        row.attrib.update({"r": str(target_row), "customHeight": "1", "spans": "1:11"})
        row.append(_build_marker_row(target_row, style_index)[0])
    else:
        marker_row = _build_marker_row(target_row, style_index)
        inserted = False
        for index, row in enumerate(list(sheet_data)):
            row_no = int(row.attrib.get("r", "0"))
            if row_no > target_row:
                sheet_data.insert(index, marker_row)
                inserted = True
                break
        if not inserted:
            sheet_data.append(marker_row)

    merge_cells = root.find(_ns("mergeCells"))
    merge_ref = f"A{target_row}:K{target_row}"
    if merge_cells is None:
        merge_cells = ET.Element(_ns("mergeCells"), {"count": "1"})
        insert_at = list(root).index(sheet_data) + 1
        root.insert(insert_at, merge_cells)
    exists = any(cell.attrib.get("ref") == merge_ref for cell in merge_cells.findall(_ns("mergeCell")))
    if not exists:
        ET.SubElement(merge_cells, _ns("mergeCell"), {"ref": merge_ref})
        merge_cells.set("count", str(len(merge_cells.findall(_ns("mergeCell")))))

    _update_dimension(root, target_row)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def write_marker_row(batch: BatchInfo) -> int:
    target_row = batch.end_row + 1
    with tempfile.TemporaryDirectory(prefix="jst_reimburse_marker_") as temp_dir:
        temp_path = Path(temp_dir) / batch.workbook_path.name
        with ZipFile(batch.workbook_path, "r") as zin, ZipFile(temp_path, "w", ZIP_DEFLATED) as zout:
            style_index: int | None = None
            patched_styles: bytes | None = None
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    patched_styles, style_index = _ensure_marker_style(data)
                    data = patched_styles
                elif item.filename == "xl/worksheets/sheet1.xml":
                    if style_index is None:
                        styles_xml = zin.read("xl/styles.xml")
                        patched_styles, style_index = _ensure_marker_style(styles_xml)
                    data = _patch_marker_sheet(data, target_row, style_index)
                zout.writestr(item, data)
        shutil.copy2(temp_path, batch.workbook_path)
    return target_row


def write_failed_export(rows: list[FailureRecord]) -> Path | None:
    if not rows:
        return None
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    path = RUNTIME_DIR / FAILED_EXPORT_TEMPLATE.format(date=datetime.now().strftime("%Y%m%d"))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "失败记录"
    sheet.append(["订单号", "失败原因"])
    for row in rows:
        sheet.append([row.order_no, row.reason])
    workbook.save(path)
    return path


def print_batch_summary(batch: BatchInfo) -> None:
    print(f"登记表：{batch.workbook_path}")
    print(f"当前批次范围：第 {batch.start_row} 行 - 第 {batch.end_row} 行")
    print(f"当前批次订单数：{len(batch.orders)}")
    print(f"本金合计：{money_text(batch.principal_total)}")
    print(f"打款金额合计：{money_text(batch.payout_total)}")


def choose_candidate(
    batch: BatchInfo,
    *,
    order_no: str | None = None,
    interactive_recovery: bool = False,
) -> tuple[CandidateResult | None, list[CandidateResult]]:
    results: list[CandidateResult] = []
    orders = [item for item in batch.orders if not order_no or item.order_no == order_no]
    if order_no and not orders:
        raise RuntimeError(f"指定订单不在当前批次内：{order_no}")

    for order in orders:
        result = CandidateResult(order=order)
        try:
            data = ops_reimburse_payload(batch, order, execute=False, interactive_recovery=interactive_recovery)
        except Exception as exc:
            result.skip_reason = str(exc)
            results.append(result)
            logging.warning("订单检查失败：%s | %s", order.order_no, exc)
            continue

        result.o_id = cell_text(data.get("internal_order_id"))
        result.lp_order_no = cell_text(data.get("online_order_id"))
        result.item_name = cell_text(data.get("item_name")) or order.product_name
        result.existing_detail = data.get("existing_detail") if isinstance(data.get("existing_detail"), dict) else {}
        result.has_existing_workorder = bool(data.get("has_existing_workorder"))
        results.append(result)
        logging.info(
            "订单检查：%s -> o_id=%s lp=%s existing=%s",
            order.order_no,
            result.o_id,
            result.lp_order_no,
            result.has_existing_workorder,
        )
        if not result.has_existing_workorder:
            return result, results
    return None, results


def _run_workflow(workflow_args: list[str]) -> int:
    from run import run_workflow

    return run_workflow(workflow_args)


def main(argv: list[str] | None = None) -> int:
    args = list(sys.argv[1:] if argv is None else argv)
    return _run_workflow(["jst_brush_reimburse_workorder", *args])


if __name__ == "__main__":
    raise SystemExit(main())
