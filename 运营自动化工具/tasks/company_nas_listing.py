#!/usr/bin/env python3
"""Download product listing materials from company NAS and build listing xlsx."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.config_loader import get_path  # noqa: E402


NAS_URL = "https://suolong.synology.me:5006"
NAS_MOUNT_NAME = "suolong.synology.me"
DEFAULT_NAS_MOUNT = get_path("company_nas_mount")
PRODUCT_LIBRARY = get_path("nas_product_library_dir")
JST_WORKBOOK = get_path("jst_product_master_file")
NAS_INDEX_PATH = get_path("nas_index_json")

BRAND_FOLDERS = {
    "奥克斯": "1.奥克斯",
    "志高": "2.志高",
    "苏泊尔": "4.苏泊尔",
    "QTQ": "5.QTQ",
    "佳健仕": "6.佳健仕",
    "蓝宝": "7.蓝宝",
    "名创优品": "8.名创优品",
    "礼品": "9.礼品图",
    "南极人": "10.南极人",
}

NAS_CATEGORIES = (
    "10.奥克斯500强修改",
    "11.南极人角标修改",
    "5.联想",
    "6.索隆",
    "7.俞兆林 北极绒",
    "8.按摩椅(旧)",
    "分销产品",
    "刮痧仪",
    "办公椅",
    "加热围巾",
    "品牌方苏泊尔详情",
    "按摩椅",
    "按摩床垫",
    "按摩垫",
    "按摩座垫",
    "按摩披肩",
    "按摩枕",
    "按摩棒",
    "按摩靠垫",
    "拔罐仪",
    "披肩",
    "揉腹仪",
    "榻榻米",
    "甩脂机",
    "甩脂腰带",
    "电竞椅",
    "盐袋",
    "筋膜枪",
    "腰腹按摩器",
    "腰部按摩器",
    "膝盖按摩器",
    "膝部按摩",
    "赠品PNG",
    "足部按摩器",
    "趴趴枕",
    "足疗机",
    "足浴盆",
    "护膝",
    "护颈仪",
    "护眼仪",
    "护腰带",
    "头部按摩器",
    "小腿按摩器",
    "手部按摩器",
)

NATURAL_TEXT_PREFIXES = (
    "从公司网盘下载",
    "公司网盘下载产品",
    "公司网盘下载",
    "NAS产品资料下载",
    "下载产品资料并生成上架数据",
)

TARGET_BRAND_DIRS = {
    "奥克斯": PRODUCT_LIBRARY / "奥克斯产品",
    "苏泊尔": PRODUCT_LIBRARY / "苏泊尔产品",
}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
SKIP_NAMES = {".DS_Store", "Thumbs.db", "desktop.ini"}
SKIP_EXTS = {".psd", ".mp4", ".mov", ".m4v", ".avi", ".db"}
NON_PRODUCT_WORDS = ("包材", "配件", "说明书", "网购箱", "彩盒", "泡沫", "适配器", "按摩头")
TITLE_NOISE_WORDS = ("优质", "爆款", "正品", "新款", "厂家", "批发", "专用", "适用")
WHITE_TRANSPARENT_ALIASES = {
    "白底透明",
    "白底图透明图",
    "透明图白底图",
    "白底透明图",
    "jpgpng",
}

# `company_nas_listing` is the single source of truth for NAS material selection.
# Docs and skills should summarize these rules instead of restating their own copies.
MAIN_IMAGE_PARENT_ALIASES = {"主图", "主附图"}
MAIN_IMAGE_CHILD_ALIASES = {"主图", "副主图", "功能性主图", "功能主图"}
SKU_DIR_ALIASES = {"sku"}
DETAIL_DIR_ALIASES = {"详情切片", "详情图"}
SCENE_DIR_ALIASES = {"场景图", "场景"}
SIZE_TOKENS_800 = ("800", "800x800", "800-800", "800x1200", "800x1000")
BLOCKED_PATH_KEYWORDS = ("视频", "旧版", "旧（厂家详情）")
DETAIL_DIR_NAME = "详情切片"
SCENE_DIR_NAME = "场景图"
MAIN_IMAGE_TARGET_DIR = "主图"
SKU_TARGET_DIR = "sku"

CATEGORY_TITLE_PROFILES = {
    "按摩椅": {
        "功能": ["全身", "全自动", "智能", "太空舱", "零重力", "电动", "多功能", "家用"],
        "属性": ["3D", "4D", "AI", "语音", "热敷", "气囊", "SL导轨", "免安装"],
        "人群": ["老人", "父母", "办公室", "家用"],
    },
    "按摩靠垫": {
        "功能": ["腰背", "颈部", "揉捏", "热敷", "多功能", "智能", "家用"],
        "属性": ["全身", "靠背", "坐垫", "电动", "便携"],
        "人群": ["老人", "父母", "办公室", "车载"],
    },
    "按摩床垫": {
        "功能": ["全身", "揉捏", "热敷", "智能", "多功能", "家用"],
        "属性": ["床垫", "电动", "便携", "舒适", "折叠"],
        "人群": ["老人", "父母", "办公室", "家庭"],
    },
    "足疗机": {
        "功能": ["足底", "揉捏", "热敷", "气囊", "全自动", "电动", "恒温"],
        "属性": ["小腿", "脚底", "滚轮", "家用", "智能", "多功能"],
        "人群": ["老人", "父母", "办公室", "男女"],
    },
    "足浴盆": {
        "功能": ["泡脚", "恒温", "加热", "按摩", "全自动", "电动", "冲浪"],
        "属性": ["深桶", "折叠", "家用", "智能", "排水", "多功能"],
        "人群": ["老人", "父母", "家庭", "送礼"],
    },
    "护眼仪": {
        "功能": ["眼部", "热敷", "按摩", "气囊", "震动", "智能", "护眼"],
        "属性": ["蓝牙", "折叠", "便携", "恒温", "多模式", "充电"],
        "人群": ["学生", "办公", "睡眠", "男女"],
    },
    "筋膜枪": {
        "功能": ["深层", "肌肉", "放松", "按摩", "静音", "多档", "震动"],
        "属性": ["便携", "迷你", "长续航", "专业", "充电"],
        "人群": ["运动", "健身", "男女", "办公室"],
    },
    "护腰带": {
        "功能": ["腰部", "热敷", "按摩", "支撑", "护腰", "保暖"],
        "属性": ["充电", "可调节", "家用", "智能"],
        "人群": ["老人", "久坐", "男女", "办公室"],
    },
}

OUTPUT_COLUMNS = [
    "产品型号",
    "颜色",
    "天猫搜索标题（30字限制）",
    "SKU命名",
    "卖点1",
    "卖点2",
    "卖点3",
    "天猫搜索标题（15字）",
    "商品编码",
    "商品重量kg",
    "长cm",
    "宽cm",
    "高cm",
    "体积立方米",
    "总库存",
    "可用数",
    "匹配商品名称",
    "匹配备注",
]


@dataclass(frozen=True)
class ModelSpec:
    raw: str
    path_text: str
    manual_code: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从公司 NAS 下载产品资料并生成上架数据.xlsx")
    parser.add_argument("--text", help="自然语言触发文本，例如：从公司网盘下载奥克斯足疗机AQA-JT-RFY06")
    parser.add_argument("--brand", help="品牌，例如：奥克斯、苏泊尔")
    parser.add_argument("--category", help="产品类别，例如：按摩靠垫、按摩椅、足浴盆")
    parser.add_argument("--models", nargs="*", default=[], help="型号列表；子目录用 / 或 \\ 表示，例如 AQA-12D-K10\\雾霾蓝")
    parser.add_argument("--models-file", help="型号文本文件，每行一个型号")
    parser.add_argument("--target-root", default=None, help="覆盖默认目标根目录")
    parser.add_argument("--jst-workbook", default=str(JST_WORKBOOK), help="聚水潭商品资料路径")
    parser.add_argument("--include-buyer-show", action="store_true", help="下载买家秀；默认不下载")
    parser.add_argument("--keep-mounted", action="store_true", help="执行结束后不卸载 NAS")
    parser.add_argument("--no-replace", action="store_true", help="不清理目标型号目录，直接增量复制")
    parser.add_argument("--skip-excel", action="store_true", help="只下载资料，不生成 Excel")
    parser.add_argument("--dry-run", action="store_true", help="只预览源目录、目标目录和预计文件数，不复制")
    return resolve_args(parser.parse_args())


def parse_natural_text(text: str) -> dict[str, Any]:
    normalized = re.sub(r"\s+", " ", text).strip()
    if not normalized:
        return {}
    brand = next((item for item in BRAND_FOLDERS if item in normalized), "")
    if not brand:
        return {}

    start = normalized.find(brand)
    remainder = normalized[start + len(brand):].strip(" ，,、")
    for prefix in NATURAL_TEXT_PREFIXES:
        if remainder.startswith(prefix):
            remainder = remainder[len(prefix):].strip(" ，,、")

    compact = re.sub(r"\s+", "", remainder)
    category = next((item for item in sorted(NAS_CATEGORIES, key=len, reverse=True) if compact.startswith(item)), "")
    if not category:
        return {"brand": brand}

    models_text = remainder
    if models_text.startswith(category):
        models_text = models_text[len(category):]
    else:
        models_text = re.sub(rf"^\s*{re.escape(category)}\s*", "", models_text, count=1)
    models_text = models_text.strip(" ，,、")
    models = [part.strip() for part in re.split(r"[，,、；;]+", models_text) if part.strip()]
    if len(models) == 1 and " " in models_text:
        models = [part.strip() for part in models_text.split() if part.strip()]
    return {"brand": brand, "category": category, "models": models}


def resolve_args(args: argparse.Namespace) -> argparse.Namespace:
    parsed = parse_natural_text(args.text or "")
    if not args.brand and parsed.get("brand"):
        args.brand = str(parsed["brand"])
    if not args.category and parsed.get("category"):
        args.category = str(parsed["category"])
    if not args.models and not args.models_file and parsed.get("models"):
        args.models = list(parsed["models"])

    missing = []
    if not args.brand:
        missing.append("--brand 或 --text 中的品牌")
    if not args.category:
        missing.append("--category 或 --text 中的类目")
    if not args.models and not args.models_file:
        missing.append("--models/--models-file 或 --text 中的型号")
    if missing:
        raise SystemExit(
            "公司 NAS 参数不足，缺少："
            + "、".join(missing)
            + "\n示例：python3 run.py \"从公司网盘下载奥克斯足疗机AQA-JT-RFY06\" --dry-run"
            + "\n或：python3 run.py company_nas_listing --brand 奥克斯 --category 足疗机 --models AQA-JT-RFY06 --dry-run"
        )
    return args


def run_command(command: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(command, text=True, capture_output=True)


def active_nas_mount() -> Path | None:
    if DEFAULT_NAS_MOUNT.exists():
        return DEFAULT_NAS_MOUNT
    volumes = DEFAULT_NAS_MOUNT.parent
    matches = sorted(
        path for path in volumes.glob(f"{NAS_MOUNT_NAME}*")
        if path.exists() and path.is_dir()
    )
    return matches[0] if matches else None


def nas_product_root() -> Path:
    configured = get_path("company_nas_product_root")
    if configured.exists():
        return configured
    mount = active_nas_mount() or DEFAULT_NAS_MOUNT
    return mount / "产品资料（运营）" / "1.产品资料"


def is_mounted() -> bool:
    return active_nas_mount() is not None


def mount_nas() -> None:
    if is_mounted():
        return
    result = run_command(["osascript", "-e", f'tell application "Finder" to mount volume "{NAS_URL}"'])
    if result.returncode != 0:
        raise SystemExit(f"NAS 挂载失败，请确认公司网络和钥匙串认证：{result.stderr.strip() or result.stdout.strip()}")
    mount = active_nas_mount()
    if not mount:
        raise SystemExit(f"NAS 挂载后仍找不到挂载点：{DEFAULT_NAS_MOUNT}")


def unmount_nas() -> dict[str, Any]:
    mount = active_nas_mount()
    if not mount:
        return {"attempted": False, "success": True, "message": "NAS 未挂载"}
    result = run_command(["umount", str(mount)])
    if result.returncode != 0 and "Resource busy" in (result.stderr or result.stdout):
        fallback = run_command(["diskutil", "unmount", str(mount)])
        return {
            "attempted": True,
            "success": fallback.returncode == 0,
            "message": (
                (fallback.stderr.strip() or fallback.stdout.strip())
                if fallback.returncode == 0
                else f"{result.stderr.strip() or result.stdout.strip()} | diskutil: {fallback.stderr.strip() or fallback.stdout.strip()}"
            ),
        }
    return {
        "attempted": True,
        "success": result.returncode == 0,
        "message": result.stderr.strip() or result.stdout.strip(),
    }


def normalize_model(raw: str) -> tuple[str, list[str]]:
    text = raw.strip().strip("\"'")
    parts = [p for p in re.split(r"[\\/]+", text) if p]
    if not parts:
        raise SystemExit("型号不能为空")
    return parts[0], parts[1:]


def parse_model_spec(raw: str) -> ModelSpec:
    text = raw.strip().strip("\"'")
    match = re.search(r"^(.*?)[【\[]\s*([^\]】]+?)\s*[】\]]\s*$", text)
    if match:
        path_text = match.group(1).strip()
        manual_code = match.group(2).strip()
    else:
        path_text = text
        manual_code = ""
    if not path_text:
        raise SystemExit(f"型号不能为空：{raw}")
    return ModelSpec(raw=text, path_text=path_text, manual_code=manual_code)


def load_models(args: argparse.Namespace) -> list[ModelSpec]:
    models = list(args.models)
    if args.models_file:
        path = Path(args.models_file).expanduser()
        if not path.is_file():
            raise SystemExit(f"型号文件不存在：{path}")
        models.extend([line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()])
    if not models:
        raise SystemExit("请通过 --models 或 --models-file 提供型号")
    return [parse_model_spec(model) for model in models]


def brand_source_dir(brand: str, category: str) -> Path:
    folder = BRAND_FOLDERS.get(brand)
    if not folder:
        raise SystemExit(f"暂未配置品牌目录：{brand}")
    return nas_product_root() / folder / category


def target_base_dir(brand: str, category: str, override: str | None) -> Path:
    if override:
        return Path(override).expanduser()
    return TARGET_BRAND_DIRS.get(brand, PRODUCT_LIBRARY / f"{brand}产品") / category


def model_source(base: Path, path_text: str) -> tuple[str, Path]:
    model, subdirs = normalize_model(path_text)
    src = base / model
    for part in subdirs:
        src = src / part
    display = "/".join([model, *subdirs])
    return display, src


def load_nas_index() -> dict[str, Any] | None:
    if not NAS_INDEX_PATH.exists():
        return None
    try:
        return json.loads(NAS_INDEX_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def score_index_record(record: dict[str, Any], *, brand: str, category: str, model: str) -> int:
    if record.get("type") != "dir":
        return 0
    if str(record.get("brand") or "") != brand:
        return 0
    if str(record.get("category") or "") != category:
        return 0

    needle = normalize_code_text(model)
    folder = normalize_code_text(record.get("model_or_folder", ""))
    filename = normalize_code_text(Path(str(record.get("path") or "")).name)
    keywords = " ".join(normalize_code_text(item) for item in (record.get("keywords") or []))
    score = 0
    if needle and needle == folder:
        score += 100
    elif needle and needle == filename:
        score += 95
    elif needle and needle in folder:
        score += 70
    elif needle and needle in filename:
        score += 65
    elif needle and needle in keywords:
        score += 40
    if int(record.get("depth") or 0) == 3:
        score += 15
    return score


def indexed_model_source(brand: str, category: str, base: Path, path_text: str) -> tuple[str, Path, str]:
    display, fallback = model_source(base, path_text)
    _, subdirs = normalize_model(path_text)
    if subdirs and fallback.is_dir():
        return display, fallback, "direct_path"

    payload = load_nas_index()
    if not payload:
        return display, fallback, "fallback_no_index"

    model, subdirs = normalize_model(path_text)
    candidates = []
    for record in payload.get("records") or []:
        score = score_index_record(record, brand=brand, category=category, model=model)
        if score <= 0:
            continue
        src = Path(str(record.get("path") or ""))
        for part in subdirs:
            src = src / part
        if src.is_dir():
            candidates.append((score, src))

    if not candidates:
        return display, fallback, "fallback_index_miss"

    candidates.sort(key=lambda item: (-item[0], len(item[1].parts), str(item[1])))
    return display, candidates[0][1], "nas_index"


def model_target(base: Path, path_text: str) -> tuple[str, Path]:
    model, subdirs = normalize_model(path_text)
    dst = base / model
    for part in subdirs:
        dst = dst / part
    display = "/".join([model, *subdirs])
    return display, dst


def is_800(path: Path) -> bool:
    s = str(path).lower()
    return any(token in s for token in SIZE_TOKENS_800)


def normalize_dir_name(name: str) -> str:
    return re.sub(r"[\W_]+", "", name).lower()


def dir_matches(name: str, aliases: set[str]) -> bool:
    return normalize_dir_name(name) in aliases


ChildDirCache = dict[Path, list[Path]]


def safe_child_dirs(
    base: Path,
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
) -> list[Path]:
    if cache is not None and base in cache:
        return cache[base]
    try:
        children = list(base.iterdir())
    except OSError:
        if cache is not None:
            cache[base] = []
        return []
    dirs: list[Path] = []
    for child in children:
        try:
            is_dir = child.is_dir()
        except OSError:
            continue
        if is_dir and not should_skip_path(child, include_buyer_show):
            dirs.append(child)
    if cache is not None:
        cache[base] = dirs
    return dirs


def iter_matching_child_dirs(
    base: Path,
    aliases: set[str],
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
) -> list[Path]:
    return [child for child in safe_child_dirs(base, include_buyer_show, cache) if dir_matches(child.name, aliases)]


def material_roots(
    base: Path,
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
) -> list[Path]:
    roots = [base]
    roots.extend(safe_child_dirs(base, include_buyer_show, cache))
    return roots


def iter_main_image_dirs(
    base: Path,
    include_buyer_show: bool,
    roots: list[Path] | None = None,
    cache: ChildDirCache | None = None,
) -> list[Path]:
    matches: list[Path] = []
    for root in roots or material_roots(base, include_buyer_show, cache):
        for parent in iter_matching_child_dirs(root, MAIN_IMAGE_PARENT_ALIASES, include_buyer_show, cache):
            matches.extend(iter_matching_child_dirs(parent, MAIN_IMAGE_CHILD_ALIASES, include_buyer_show, cache))
    return matches


def iter_category_dirs(
    base: Path,
    aliases: set[str],
    include_buyer_show: bool,
    roots: list[Path] | None = None,
    cache: ChildDirCache | None = None,
) -> list[Path]:
    matches: list[Path] = []
    for root in roots or material_roots(base, include_buyer_show, cache):
        matches.extend(iter_matching_child_dirs(root, aliases, include_buyer_show, cache))
    return matches


def should_skip_path(path: Path, include_buyer_show: bool) -> bool:
    # This gate owns the global skip policy: system junk, PSD/video, old dirs,
    # and buyer-show exclusion unless the caller explicitly opts in.
    text = str(path)
    if path.name in SKIP_NAMES or path.suffix.lower() in SKIP_EXTS:
        return True
    if not include_buyer_show and "买家秀" in text:
        return True
    if any(keyword in text for keyword in BLOCKED_PATH_KEYWORDS):
        return True
    return False


def collect_under(base: Path, rel_dir: Path | str, include_buyer_show: bool, predicate=lambda p: True) -> list[Path]:
    start = base / rel_dir
    out: list[Path] = []
    for root, dirs, files in os.walk(start):
        dirs[:] = [d for d in dirs if not should_skip_path(Path(root) / d, include_buyer_show)]
        for name in files:
            p = Path(root) / name
            if should_skip_path(p, include_buyer_show):
                continue
            if p.suffix.lower() not in IMAGE_EXTS:
                continue
            if predicate(p):
                out.append(p)
    return out


def selected_files(base: Path, include_buyer_show: bool) -> list[Path]:
    # Selection order is intentional and defines the canonical download contract:
    # main images/SKU/scene images use the 800-token rule; details only take
    # `790*`; white-transparent assets are flattened to the product root; buyer
    # show stays off unless `--include-buyer-show` is set.
    files: list[Path] = []
    child_dir_cache: ChildDirCache = {}
    roots = material_roots(base, include_buyer_show, child_dir_cache)
    for start in iter_main_image_dirs(base, include_buyer_show, roots, child_dir_cache):
        files += collect_under(base, start, include_buyer_show, lambda p: is_800(p) and p.suffix.lower() in {".jpg", ".jpeg"})
    for start in iter_category_dirs(base, SKU_DIR_ALIASES, include_buyer_show, roots, child_dir_cache):
        files += collect_under(base, start, include_buyer_show, lambda p: is_800(p) and p.suffix.lower() in {".jpg", ".jpeg"})

    for detail in iter_category_dirs(base, DETAIL_DIR_ALIASES, include_buyer_show, roots, child_dir_cache):
        # Keep both historical layouts:
        # 1. detail/790/*.jpg
        # 2. detail/*.jpg|*.gif with "790" in filename
        # One walk covers both layouts; WebDAV traversal is the slow path.
        def is_detail_790(p: Path, detail_dir: Path = detail) -> bool:
            if "790" in p.name:
                return True
            try:
                rel_parts = p.relative_to(detail_dir).parts[:-1]
            except ValueError:
                return False
            return any(part.startswith("790") for part in rel_parts)

        files += collect_under(
            base,
            detail,
            include_buyer_show,
            is_detail_790,
        )

    for start in iter_category_dirs(base, WHITE_TRANSPARENT_ALIASES, include_buyer_show, roots, child_dir_cache):
        files += collect_under(base, start, include_buyer_show, lambda p: is_800(p))

    for start in iter_category_dirs(base, SCENE_DIR_ALIASES, include_buyer_show, roots, child_dir_cache):
        files += collect_under(base, start, include_buyer_show, lambda p: is_800(p) and p.suffix.lower() in {".jpg", ".jpeg"})

    if include_buyer_show:
        files += collect_under(base, "买家秀", include_buyer_show)

    seen: set[Path] = set()
    unique: list[Path] = []
    for p in files:
        if p not in seen:
            unique.append(p)
            seen.add(p)
    return unique


def copy_relative_path(src: Path, item: Path) -> Path:
    rel = item.relative_to(src)
    alias_parts = [part for part in rel.parts[:-1] if dir_matches(part, WHITE_TRANSPARENT_ALIASES)]
    if alias_parts:
        return Path(item.name)
    if any(dir_matches(part, SCENE_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(SCENE_DIR_NAME) / item.name
    if any(dir_matches(part, DETAIL_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(DETAIL_DIR_NAME) / item.name
    if any(dir_matches(part, SKU_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(SKU_TARGET_DIR) / item.name
    for idx, part in enumerate(rel.parts[:-1]):
        if not dir_matches(part, MAIN_IMAGE_PARENT_ALIASES):
            continue
        next_idx = idx + 1
        if next_idx < len(rel.parts) - 1 and dir_matches(rel.parts[next_idx], MAIN_IMAGE_CHILD_ALIASES):
            return Path(MAIN_IMAGE_TARGET_DIR) / item.name
    if (
        len(rel.parts) >= 2
        and dir_matches(rel.parts[0], MAIN_IMAGE_PARENT_ALIASES)
        and dir_matches(rel.parts[1], MAIN_IMAGE_CHILD_ALIASES)
    ):
        return Path(MAIN_IMAGE_TARGET_DIR) / item.name
    return rel


def copy_product(src: Path, dst: Path, files: list[Path], replace: bool, dry_run: bool) -> tuple[int, list[str]]:
    if dry_run:
        return 0, []
    if replace and dst.exists():
        shutil.rmtree(dst)
    dst.mkdir(parents=True, exist_ok=True)
    copied = 0
    missing_files: list[str] = []
    for item in files:
        rel = copy_relative_path(src, item)
        out = dst / rel
        if not replace and out.exists():
            continue
        if out.exists():
            out = out.with_name(f"{out.stem}_{copied + 1}{out.suffix}")
        out.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(item, out)
        except FileNotFoundError as exc:
            missing_files.append(f"{item} :: {exc}")
            continue
        except OSError as exc:
            missing_files.append(f"{item} :: {exc}")
            continue
        copied += 1
    return copied, missing_files


def row_text(row: tuple[Any, ...], idx: dict[str, int], columns: list[str]) -> str:
    values = []
    for col in columns:
        i = idx.get(col)
        values.append(str(row[i] or "") if i is not None and i < len(row) else "")
    return " ".join(values)


def load_jst_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    if not path.is_file():
        raise SystemExit(f"聚水潭商品资料不存在：{path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def jst_cell(row: tuple[Any, ...], idx: dict[str, int], name: str) -> Any:
    i = idx.get(name)
    return row[i] if i is not None and i < len(row) else None


def model_tokens(display_model: str) -> list[str]:
    parts = [p for p in display_model.split("/") if p]
    tokens = [parts[0], parts[0].replace("-", "")]
    tokens.extend(parts[1:])
    return [t.lower() for t in tokens if t]


def normalize_code_text(value: Any) -> str:
    return re.sub(r"[\s\-/_.]+", "", str(value or "").strip()).lower()


def match_jst(display_model: str, manual_code: str, headers: list[str], rows: list[tuple[Any, ...]]) -> tuple[tuple[Any, ...] | None, str]:
    idx = {h: i for i, h in enumerate(headers)}
    code_idx = idx.get("商品编码")
    if code_idx is None:
        return None, "聚水潭资料缺少【商品编码】字段，需人工确认"

    if manual_code:
        target = normalize_code_text(manual_code)
        matches = [row for row in rows if normalize_code_text(jst_cell(row, idx, "商品编码")) == target]
        if not matches:
            return None, f"指定商品编码未匹配到聚水潭资料：{manual_code}"
        if len(matches) > 1:
            return matches[0], f"指定商品编码匹配到多条，已取第一条：{manual_code}"
        return matches[0], f"按指定商品编码精确匹配：{manual_code}"

    tokens = [normalize_code_text(token) for token in model_tokens(display_model)]
    tokens = [token for token in tokens if token]
    candidates = []
    for row in rows:
        product_code = str(jst_cell(row, idx, "商品编码") or "").strip()
        if not product_code:
            continue
        norm_code = normalize_code_text(product_code)
        score = 0
        for token in tokens:
            if norm_code == token:
                score = max(score, 100)
            elif token and token in norm_code:
                score = max(score, 60 + len(token))
            elif norm_code and norm_code in token:
                score = max(score, 40 + len(norm_code))
        if score:
            candidates.append((score, row))
    if not candidates:
        return None, "未提供商品编码，且按商品编码字段未匹配到聚水潭资料，需人工确认"
    candidates.sort(key=lambda item: item[0], reverse=True)
    top_score = candidates[0][0]
    same = [row for score, row in candidates if score == top_score]
    if len(same) > 1:
        return same[0], f"按商品编码字段模糊匹配到多条，已取第一条；候选数 {len(same)}"
    return candidates[0][1], "按商品编码字段模糊匹配"


def unique_terms(terms: list[str], category: str) -> list[str]:
    seen = set()
    out = []
    for term in terms:
        word = str(term).strip()
        if not word or word in seen or word == category or any(noise in word for noise in TITLE_NOISE_WORDS):
            continue
        seen.add(word)
        out.append(word)
    return out


def extract_title_terms(display_model: str, match: tuple[Any, ...] | None, idx: dict[str, int], category: str) -> list[str]:
    fields = ["商品名称", "商品简称", "分类", "颜色及规格", "颜色", "规格", "备注", "虚拟分类", "款式编码", "商品编码"]
    text = row_text(match, idx, fields) if match else display_model
    profile_words = []
    for profile in CATEGORY_TITLE_PROFILES.values():
        profile_words.extend(profile["功能"])
        profile_words.extend(profile["属性"])
        profile_words.extend(profile["人群"])
    candidates = [word for word in profile_words if word and word in text]
    color = infer_color(display_model, match, idx)
    if color:
        candidates.append(color)
    return unique_terms(candidates, category)


def category_profile(category: str) -> dict[str, list[str]]:
    return CATEGORY_TITLE_PROFILES.get(
        category,
        {
            "功能": ["智能", "多功能", "家用", "电动", "便携", "舒适"],
            "属性": ["大功率", "轻便", "实用", "升级"],
            "人群": ["家庭", "老人", "办公室", "送礼"],
        },
    )


def append_title_terms(base: str, terms: list[str], max_len: int) -> str:
    title = base
    for term in unique_terms(terms, ""):
        if term in title:
            continue
        if len(title) + len(term) <= max_len:
            title += term
    return title


def build_search_title(display_model: str, brand: str, category: str, match: tuple[Any, ...] | None, idx: dict[str, int]) -> str:
    profile = category_profile(category)
    extracted = extract_title_terms(display_model, match, idx, category)
    priority_terms = [*profile["功能"], *extracted, *profile["属性"], *profile["人群"]]
    title = append_title_terms(f"{brand}{category}", priority_terms, 30)
    title = title.replace(f"{category}{category}", category)
    if title.count(category) > 1:
        first = title.find(category)
        title = title[: first + len(category)] + title[first + len(category) :].replace(category, "")
    if len(title) < 28:
        title = append_title_terms(title, [*profile["属性"], *profile["人群"], "旗舰款", "送礼"], 30)
    return title[:30]


def compact_title(title: str, length: int = 15) -> str:
    return title[:length] if len(title) >= length else fit_len(title, length)


def fit_len(text: str, length: int) -> str:
    if len(text) >= length:
        return text[:length]
    return text + ("款" * (length - len(text)))


def infer_color(display_model: str, match: tuple[Any, ...] | None, idx: dict[str, int]) -> str:
    parts = display_model.split("/")
    if len(parts) > 1:
        return parts[-1]
    if match:
        name = str(jst_cell(match, idx, "商品名称") or "")
        m = re.search(r"[（(【](.*?色).*?[）)】]", name)
        if m:
            return m.group(1)
    return ""


def infer_brand(match: tuple[Any, ...] | None, idx: dict[str, int], fallback: str) -> str:
    if match:
        brand = str(jst_cell(match, idx, "品牌") or "").strip()
        if brand:
            return brand
    return fallback


def selling_points(category: str) -> list[str]:
    if category == "按摩靠垫":
        return ["腰背按摩", "热敷舒压", "揉捏放松"]
    if category == "按摩床垫":
        return ["全身按摩", "热敷舒压", "家用便携"]
    if category == "按摩椅":
        return ["AI语控", "零重体验", "全身拉伸"]
    return ["智能操控", "热敷舒压", "家用便携"]


def listing_row(display_model: str, brand: str, category: str, match: tuple[Any, ...] | None, remark: str, headers: list[str]) -> list[Any]:
    idx = {h: i for i, h in enumerate(headers)}
    title_brand = infer_brand(match, idx, brand)
    title30 = build_search_title(display_model, title_brand, category, match, idx)
    title15 = compact_title(title30)
    color = infer_color(display_model, match, idx)
    points = selling_points(category)
    sku = f"【旗舰款】{color or '标准款'}{points[0]}"
    if match is None:
        return [display_model, color, title30, sku, *points, title15, "", "", "", "", "", "", "", "", "", remark]

    volume = jst_cell(match, idx, "体积")
    try:
        volume_m3 = round(float(volume) / 1000000, 6) if volume not in (None, "") else ""
    except Exception:
        volume_m3 = ""
    return [
        display_model,
        color,
        title30,
        sku,
        *points,
        title15,
        jst_cell(match, idx, "商品编码") or "",
        jst_cell(match, idx, "重量") or "",
        jst_cell(match, idx, "长") or "",
        jst_cell(match, idx, "宽") or "",
        jst_cell(match, idx, "高") or "",
        volume_m3,
        jst_cell(match, idx, "总库存") or 0,
        jst_cell(match, idx, "可用数") or 0,
        jst_cell(match, idx, "商品名称") or "",
        remark,
    ]


def save_listing(path: Path, rows: list[list[Any]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "上架数据"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_COLUMNS))
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(OUTPUT_COLUMNS)):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [18, 10, 34, 28, 10, 10, 10, 22, 16, 12, 10, 10, 10, 12, 10, 10, 42, 26]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(OUTPUT_COLUMNS))}{ws.max_row}"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def validate_outputs(base: Path | list[Path], listing_files: list[Path], include_buyer_show: bool) -> dict[str, Any]:
    bases = base if isinstance(base, list) else [base]
    invalid_files = []
    for item_base in bases:
        for root, _, files in os.walk(item_base):
            for name in files:
                p = Path(root) / name
                if p.suffix.lower() in SKIP_EXTS or p.name in SKIP_NAMES:
                    invalid_files.append(str(p))
    buyer_show_dirs = []
    if not include_buyer_show:
        for item_base in bases:
            buyer_show_dirs.extend(str(p) for p in item_base.rglob("买家秀") if p.is_dir())

    rule_errors = []
    for listing_file in listing_files:
        if listing_file.exists():
            wb = load_workbook(listing_file, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[2]]
            ci = {h: i + 1 for i, h in enumerate(headers)}
            for row_num in range(3, ws.max_row + 1):
                title30 = ws.cell(row_num, ci["天猫搜索标题（30字限制）"]).value or ""
                title15 = ws.cell(row_num, ci["天猫搜索标题（15字）"]).value or ""
                if len(title30) > 30 or len(title30) < 28:
                    rule_errors.append({"file": str(listing_file), "row": row_num, "field": "30字标题", "length": len(title30)})
                if len(title15) != 15:
                    rule_errors.append({"file": str(listing_file), "row": row_num, "field": "15字标题", "length": len(title15)})
                for field in ("卖点1", "卖点2", "卖点3"):
                    value = ws.cell(row_num, ci[field]).value or ""
                    if len(value) > 4:
                        rule_errors.append({"file": str(listing_file), "row": row_num, "field": field, "length": len(value)})

    return {
        "listing_files_exist": all(p.exists() for p in listing_files),
        "buyer_show_dirs": buyer_show_dirs,
        "invalid_files": invalid_files,
        "rule_errors": rule_errors,
    }


def main() -> None:
    args = parse_args()
    model_specs = load_models(args)
    mount_nas()
    resolved_mount = active_nas_mount()

    source_base = brand_source_dir(args.brand, args.category)
    if not source_base.is_dir() and not load_nas_index():
        raise SystemExit(f"源类目目录不存在：{source_base}")

    target_base = target_base_dir(args.brand, args.category, args.target_root)
    plan = []
    listing_files: list[Path] = []
    target_dirs: list[Path] = []
    jst_headers: list[str] = []
    jst_rows: list[tuple[Any, ...]] = []
    if not args.skip_excel:
        jst_headers, jst_rows = load_jst_rows(Path(args.jst_workbook).expanduser())

    try:
        for spec in model_specs:
            display, src, source_resolver = indexed_model_source(args.brand, args.category, source_base, spec.path_text)
            _, dst = model_target(target_base, spec.path_text)
            if not src.is_dir():
                plan.append({
                    "model": display,
                    "manual_code": spec.manual_code,
                    "source": str(src),
                    "source_resolver": source_resolver,
                    "target": str(dst),
                    "status": "源目录不存在",
                    "selected_files": 0,
                    "copied_files": 0,
                })
                continue
            selection_start = time.monotonic()
            files = selected_files(src, args.include_buyer_show)
            selection_seconds = round(time.monotonic() - selection_start, 3)
            copy_start = time.monotonic()
            copied, missing_files = copy_product(src, dst, files, replace=not args.no_replace, dry_run=args.dry_run)
            copy_seconds = round(time.monotonic() - copy_start, 3)
            record = {
                "model": display,
                "manual_code": spec.manual_code,
                "source": str(src),
                "source_resolver": source_resolver,
                "target": str(dst),
                "status": "ok",
                "selected_files": len(files),
                "copied_files": copied,
                "selection_seconds": selection_seconds,
                "copy_seconds": copy_seconds,
                "missing_files": missing_files,
            }
            plan.append(record)
            target_dirs.append(dst)
            if not args.skip_excel and not args.dry_run:
                match, remark = match_jst(display, spec.manual_code, jst_headers, jst_rows)
                row = listing_row(display, args.brand, args.category, match, remark, jst_headers)
                listing_path = dst / "上架数据.xlsx"
                save_listing(listing_path, [row], f"{display} 上架数据")
                listing_files.append(listing_path)

        if not args.skip_excel and not args.dry_run:
            validation = validate_outputs(target_dirs, listing_files, args.include_buyer_show)
        else:
            validation = {}

        unmount = {"attempted": False, "success": True, "message": "keep-mounted"}
        if args.keep_mounted:
            unmount = {"attempted": False, "success": True, "message": "keep-mounted"}
        if not args.keep_mounted:
            unmount = unmount_nas()

        print(json.dumps({
            "task": "company_nas_listing",
            "brand": args.brand,
            "category": args.category,
            "target_base": str(target_base),
            "include_buyer_show": args.include_buyer_show,
            "dry_run": args.dry_run,
            "mount_path": str(resolved_mount) if resolved_mount else "",
            "used_existing_mount": bool(resolved_mount),
            "items": plan,
            "listing_workbooks": [str(p) for p in listing_files],
            "validation": validation,
            "unmount": unmount,
            "finished_at": datetime.now().isoformat(timespec="seconds"),
        }, ensure_ascii=False, indent=2))
    finally:
        if not args.keep_mounted and is_mounted():
            # Best-effort safety if an exception happened after mounting.
            unmount_nas()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("已中断", file=sys.stderr)
        raise SystemExit(130)
