from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ITEM_ID_HEADERS = {"平台商品ID", "商品ID", "item_id", "itemId", "platform_item_id"}


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = value.strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def parse_item_ids(raw: str | None) -> list[str]:
    if not raw:
        return []
    return _dedupe([part.strip() for part in raw.split(",")])


def load_item_ids_from_excel(path: str | Path) -> list[str]:
    file_path = Path(path).expanduser()
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{file_path}")
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Excel 文件为空：{file_path}")
    headers = [_to_text(value) for value in rows[0]]
    column_index = next((idx for idx, header in enumerate(headers) if header in ITEM_ID_HEADERS), None)
    if column_index is None:
        raise RuntimeError(f"Excel 中找不到商品ID字段，支持字段：{', '.join(sorted(ITEM_ID_HEADERS))}")
    return _dedupe([_to_text(row[column_index] if column_index < len(row) else "") for row in rows[1:]])


def resolve_item_ids(*, item_ids: str | None, input_file: str | None) -> list[str]:
    merged: list[str] = []
    merged.extend(parse_item_ids(item_ids))
    if input_file:
        merged.extend(load_item_ids_from_excel(input_file))
    result = _dedupe(merged)
    if not result:
        raise RuntimeError("没有输入商品ID。请使用 --item-ids 或 --input-file。")
    return result
